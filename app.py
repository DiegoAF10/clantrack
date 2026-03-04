#!/usr/bin/env python3
"""
WALMART INTELLIGENCE — Clan Cervecero
Streamlit Dashboard
"""

import sqlite3
import os
import sys
import shutil
import tempfile
import subprocess
from datetime import datetime, timedelta

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ═══════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "walmart.db")
IMPORTS_DIR = os.path.join(BASE_DIR, "imports")
ARCHIVE_DIR = os.path.join(BASE_DIR, "imports", "_processed")
REPORT_DIR = os.path.join(BASE_DIR, "reports")

# Environment detection: local (Windows) vs Streamlit Cloud (Linux)
IS_LOCAL = sys.platform == "win32"

# Ensure directories
os.makedirs(IMPORTS_DIR, exist_ok=True)
os.makedirs(ARCHIVE_DIR, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)

# Colors
CLAN_BLUE = "#2D4A7A"
CLAN_DARK = "#1B2A4A"
CLAN_LIGHT = "#E8EDF5"
GREEN = "#27AE60"
RED = "#E74C3C"
ORANGE = "#F39C12"
PURPLE = "#8E24AA"

# Business thresholds & defaults
CONFIG = {
    # Alertas
    "oos_pct_threshold": 0.30,
    "coverage_critical_days": 3,
    "coverage_warning_days": 7,
    "overstock_days": 60,
    "sales_drop_critical_pct": 0.50,
    "sales_drop_warning_pct": 0.25,
    "store_drop_critical_pct": 0.60,
    "store_drop_warning_pct": 0.30,
    "claims_critical_q": 2000,
    "claims_warning_q": 500,
    "claims_store_critical": 8,
    "claims_store_warning": 3,
    "min_sales_for_oos": 5,
    # Sugerido
    "default_target_days": 15,
    "default_lead_time_days": 10,
    "priority_p2_multiplier": 0.7,
    "priority_p3_multiplier": 1.5,
    "velocity_weeks": 5,
    # Centralization
    "centralization_rate": 0.04,
    # Dashboard
    "default_comparison_weeks": 4,
}

# Margin labels
MARGIN_SELLOUT = "Margen Sell-Out — precio consumidor menos costo Clan a Walmart"
MARGIN_SELLIN = "Margen Sell-In — lo que Clan factura a Walmart menos costos"


# ═══════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════

def _build_wm_week_map():
    """Build mapping of semana_wm -> first real date from the database.

    Walmart fiscal weeks do NOT follow ISO weeks. The only source of truth
    is the 'diario' column in retail_link. This function reads the actual
    min(diario) per semana_wm and caches the result.
    """
    try:
        conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        rows = conn.execute("""
            SELECT semana_wm, MIN(diario)
            FROM retail_link
            WHERE diario IS NOT NULL AND diario != '' AND diario != 'nan'
            GROUP BY semana_wm
            ORDER BY semana_wm
        """).fetchall()
        conn.close()
        mapping = {}
        for sem, diario in rows:
            try:
                # diario format: "2026/02/21"
                d = datetime.strptime(str(diario).strip(), "%Y/%m/%d")
                mapping[str(sem).strip()] = d
            except Exception:
                pass
        return mapping
    except Exception:
        return {}


@st.cache_data(ttl=300)
def _get_wm_week_map():
    """Cached version — refreshes every 5 min or on new import."""
    return _build_wm_week_map()


def wm_week_to_date(semana_wm: str) -> str:
    """Convert Walmart week '202604' to readable date 'Feb 21, 2026'.

    Uses real dates from the database (diario column), NOT ISO week math.
    """
    s = str(semana_wm).strip()
    if len(s) != 6 or s == '—':
        return s
    mapping = _get_wm_week_map()
    d = mapping.get(s)
    if d:
        return d.strftime("%b %d, %Y")
    # Fallback: interpolate from nearest known week
    return _wm_week_interpolate(s, "%b %d, %Y")


def wm_week_to_date_short(semana_wm: str) -> str:
    """Convert '202604' to 'Feb 21'."""
    s = str(semana_wm).strip()
    if len(s) != 6 or s == '—':
        return s
    mapping = _get_wm_week_map()
    d = mapping.get(s)
    if d:
        return d.strftime("%b %d")
    return _wm_week_interpolate(s, "%b %d")


def wm_week_to_datetime(semana_wm: str):
    """Convert '202604' to datetime object."""
    s = str(semana_wm).strip()
    if len(s) != 6:
        return None
    mapping = _get_wm_week_map()
    d = mapping.get(s)
    if d:
        return d
    # Interpolation fallback
    return _wm_week_interpolate_dt(s)


def _wm_week_interpolate(semana_wm: str, fmt: str) -> str:
    """Interpolate date for a WM week not in the database.

    Uses nearest known week and offsets by 7 days per week difference.
    """
    d = _wm_week_interpolate_dt(semana_wm)
    if d:
        return d.strftime(fmt)
    return semana_wm


def _wm_week_interpolate_dt(semana_wm: str):
    """Return interpolated datetime for unknown WM week."""
    mapping = _get_wm_week_map()
    if not mapping:
        return None
    try:
        target_year = int(semana_wm[:4])
        target_week = int(semana_wm[4:])
        # Find closest known week
        best_key = None
        best_dist = 9999
        for k in mapping:
            ky = int(k[:4])
            kw = int(k[4:])
            # Convert to absolute week number for distance
            abs_target = target_year * 53 + target_week
            abs_known = ky * 53 + kw
            dist = abs(abs_target - abs_known)
            if dist < best_dist:
                best_dist = dist
                best_key = k
        if best_key:
            ref_date = mapping[best_key]
            ref_year = int(best_key[:4])
            ref_week = int(best_key[4:])
            week_diff = (target_year * 53 + target_week) - (ref_year * 53 + ref_week)
            return ref_date + timedelta(days=week_diff * 7)
    except Exception:
        pass
    return None


def format_period_label(period_val, granularity):
    """Format a period value for display based on granularity."""
    if granularity == "Diario":
        try:
            d = datetime.strptime(str(period_val), "%Y/%m/%d")
            return d.strftime("%b %d")
        except Exception:
            return str(period_val)
    elif granularity == "Semanal":
        return wm_week_to_date_short(period_val)
    elif granularity == "Mensual":
        try:
            parts = str(period_val).split('-')
            d = datetime(int(parts[0]), int(parts[1]), 1)
            return d.strftime("%b %Y")
        except Exception:
            return str(period_val)
    return str(period_val)


def build_trend_query(granularity, where_extra="", extra_select="", extra_joins=""):
    """Build SQL for aggregating sales by granularity (Diario/Semanal/Mensual)."""
    if granularity == "Diario":
        return f"""
            SELECT diario as periodo,
                   SUM(venta_und) as und,
                   SUM(venta_q) as q,
                   SUM(venta_costo_q) as costo_q,
                   COUNT(DISTINCT store_nbr) as tiendas,
                   COUNT(DISTINCT item_nbr) as skus
                   {extra_select}
            FROM retail_link rl {extra_joins}
            WHERE rl.tipo_registro='VENTA' AND rl.diario IS NOT NULL {where_extra}
            GROUP BY diario ORDER BY diario
        """
    elif granularity == "Mensual":
        return f"""
            SELECT rl.anio || '-' || printf('%02d', rl.mes) as periodo,
                   SUM(venta_und) as und,
                   SUM(venta_q) as q,
                   SUM(venta_costo_q) as costo_q,
                   COUNT(DISTINCT store_nbr) as tiendas,
                   COUNT(DISTINCT item_nbr) as skus
                   {extra_select}
            FROM retail_link rl {extra_joins}
            WHERE rl.tipo_registro='VENTA' AND rl.anio IS NOT NULL {where_extra}
            GROUP BY rl.anio, rl.mes ORDER BY rl.anio, rl.mes
        """
    else:  # Semanal
        return f"""
            SELECT semana_wm as periodo,
                   SUM(venta_und) as und,
                   SUM(venta_q) as q,
                   SUM(venta_costo_q) as costo_q,
                   COUNT(DISTINCT store_nbr) as tiendas,
                   COUNT(DISTINCT item_nbr) as skus
                   {extra_select}
            FROM retail_link rl {extra_joins}
            WHERE rl.tipo_registro='VENTA' {where_extra}
            GROUP BY semana_wm ORDER BY semana_wm
        """


def fmt_q(val):
    """Format as GTQ currency."""
    if val is None:
        return "Q0"
    return f"Q{val:,.0f}"


def fmt_pct(val):
    """Format as percentage."""
    if val is None:
        return "0%"
    return f"{val:.1%}"


def fmt_int(val):
    """Format as integer with commas."""
    if val is None:
        return "0"
    return f"{int(val):,}"


def download_df(df, filename, label="Descargar CSV"):
    csv = df.to_csv(index=False).encode('utf-8')
    st.download_button(label, csv, f"{filename}.csv", "text/csv", use_container_width=True)


def delta_color(val):
    """Return 'normal' or 'inverse' for st.metric delta."""
    return "normal"


@st.cache_resource
def get_db():
    """Get DB connection (cached)."""
    return sqlite3.connect(DB_PATH, check_same_thread=False)


def query_df(sql, params=None):
    """Execute SQL and return DataFrame."""
    conn = get_db()
    return pd.read_sql_query(sql, conn, params=params)


def query_one(sql, params=None):
    """Execute SQL and return single row."""
    conn = get_db()
    cur = conn.execute(sql, params or [])
    return cur.fetchone()


def query_val(sql, params=None):
    """Execute SQL and return single value."""
    row = query_one(sql, params)
    return row[0] if row else None


# ═══════════════════════════════════════════
# PAGE: ALERTAS
# ═══════════════════════════════════════════

def page_alertas():
    st.markdown("## Centro de Alertas")

    # ── Week windows ──
    weeks = query_df("""
        SELECT DISTINCT semana_wm FROM retail_link
        WHERE tipo_registro='VENTA' ORDER BY semana_wm DESC LIMIT 8
    """)['semana_wm'].tolist()

    if not weeks:
        st.warning("No hay datos de venta. Subí archivos en la pestaña 'Upload'.")
        return

    last_week = weeks[0]
    prev_week = weeks[1] if len(weeks) >= 2 else None
    curr_4 = weeks[:4] if len(weeks) >= 4 else weeks
    prev_4 = weeks[4:8] if len(weeks) >= 8 else []

    inv_week = query_val(
        "SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO'"
    )

    # ── Data freshness check ──
    last_import = query_one(
        "SELECT imported_at FROM import_log WHERE rows_added > 0 ORDER BY id DESC LIMIT 1"
    )
    if last_import and last_import[0]:
        try:
            last_dt = datetime.strptime(last_import[0][:10], "%Y-%m-%d")
            days_stale = (datetime.now() - last_dt).days
            if days_stale > 7:
                st.warning(
                    f"Datos desactualizados — última importación con datos nuevos: "
                    f"{last_dt.strftime('%d/%m/%Y')} ({days_stale} días)."
                )
        except Exception:
            pass

    # ── Collect alerts ──
    # severity: 0=CRÍTICO, 1=ATENCIÓN, 2=INFO
    alerts = []

    # --- ALERT 1: OOS Crítico (>30% tiendas a 0 inventario) ---
    # Only flag products with recent sales (excludes seasonals like HB Oktoberfest)
    if inv_week:
        ph_c4 = ','.join(['?'] * len(curr_4))
        df_oos = query_df(f"""
            SELECT
                r.item_nbr, r.producto,
                COUNT(DISTINCT r.store_nbr) as tiendas_total,
                SUM(CASE WHEN r.inv_actual = 0 THEN 1 ELSE 0 END) as tiendas_oos
            FROM retail_link r
            WHERE r.tipo_registro = 'INVENTARIO' AND r.semana_wm = ?
              AND r.item_nbr IN (
                SELECT item_nbr FROM retail_link
                WHERE tipo_registro='VENTA' AND semana_wm IN ({ph_c4})
                GROUP BY item_nbr HAVING SUM(venta_und) >= {CONFIG["min_sales_for_oos"]}
              )
            GROUP BY r.item_nbr, r.producto
            HAVING tiendas_oos > 0
        """, [inv_week] + curr_4)
        if not df_oos.empty:
            df_oos['pct_oos'] = df_oos['tiendas_oos'] / df_oos['tiendas_total']
            df_crit = df_oos[df_oos['pct_oos'] >= CONFIG["oos_pct_threshold"]].copy()
            if not df_crit.empty:
                df_crit = df_crit.sort_values('pct_oos', ascending=False)
                alerts.append({
                    'severity': 0,
                    'title': f"OOS Crítico — {len(df_crit)} productos con >{CONFIG['oos_pct_threshold']:.0%} tiendas sin stock",
                    'caption': f"Inventario: {wm_week_to_date(inv_week)}",
                    'df': df_crit[['producto', 'tiendas_oos', 'tiendas_total', 'pct_oos']].rename(columns={
                        'producto': 'Producto', 'tiendas_oos': 'Tiendas OOS',
                        'tiendas_total': 'Tiendas Total', 'pct_oos': '% OOS'
                    }),
                    'fmt': {'% OOS': '{:.0%}'},
                })

    # --- ALERT 2: Cobertura Baja (<7 días) ---
    if inv_week:
        ph_c4 = ','.join(['?'] * len(curr_4))
        df_cov = query_df(f"""
            WITH inv AS (
                SELECT item_nbr, producto, SUM(inv_actual) as inv_total
                FROM retail_link
                WHERE tipo_registro='INVENTARIO' AND semana_wm = ?
                GROUP BY item_nbr
            ),
            sales AS (
                SELECT item_nbr, SUM(venta_und) as venta_total,
                       COUNT(DISTINCT semana_wm) as n_weeks
                FROM retail_link
                WHERE tipo_registro='VENTA' AND semana_wm IN ({ph_c4})
                GROUP BY item_nbr
            )
            SELECT inv.item_nbr, inv.producto, inv.inv_total,
                   COALESCE(sales.venta_total, 0) as venta_4w,
                   sales.n_weeks
            FROM inv LEFT JOIN sales ON inv.item_nbr = sales.item_nbr
        """, [inv_week] + curr_4)
        if not df_cov.empty:
            df_cov['avg_weekly'] = df_cov.apply(
                lambda r: r['venta_4w'] / r['n_weeks'] if r['n_weeks'] and r['n_weeks'] > 0 else 0, axis=1
            )
            df_cov['dias_cobertura'] = df_cov.apply(
                lambda r: (r['inv_total'] * 7 / r['avg_weekly']) if r['avg_weekly'] > 0 else 999, axis=1
            )
            df_low = df_cov[df_cov['dias_cobertura'] < CONFIG["coverage_warning_days"]].copy()
            if not df_low.empty:
                df_low = df_low.sort_values('dias_cobertura')
                for _, row in df_low.iterrows():
                    sev = 0 if row['dias_cobertura'] <= CONFIG["coverage_critical_days"] else 1
                    # Collect per-severity batch below
                df_crit_cov = df_low[df_low['dias_cobertura'] <= CONFIG["coverage_critical_days"]]
                df_att_cov = df_low[(df_low['dias_cobertura'] > CONFIG["coverage_critical_days"]) & (df_low['dias_cobertura'] < CONFIG["coverage_warning_days"])]
                for sev, subset, label in [
                    (0, df_crit_cov, f"≤{CONFIG['coverage_critical_days']} días"),
                    (1, df_att_cov, f"{CONFIG['coverage_critical_days']}-{CONFIG['coverage_warning_days']} días"),
                ]:
                    if not subset.empty:
                        alerts.append({
                            'severity': sev,
                            'title': f"Cobertura Baja ({label}) — {len(subset)} productos",
                            'caption': f"Inventario al {wm_week_to_date(inv_week)}, ventas promedio últimas {len(curr_4)} semanas",
                            'df': subset[['producto', 'inv_total', 'avg_weekly', 'dias_cobertura']].rename(columns={
                                'producto': 'Producto', 'inv_total': 'Inventario',
                                'avg_weekly': 'Venta/Semana', 'dias_cobertura': 'Días Cobertura'
                            }),
                            'fmt': {'Venta/Semana': '{:,.0f}', 'Inventario': '{:,.0f}', 'Días Cobertura': '{:.1f}'},
                        })

    # --- ALERT 3: Sobre-stock (>60 días cobertura) ---
    if inv_week and 'df_cov' in locals() and not df_cov.empty:
        df_over = df_cov[df_cov['dias_cobertura'] > CONFIG["overstock_days"]].copy()
        df_over = df_over[df_over['dias_cobertura'] < 999]  # exclude zero-sales infinite
        if not df_over.empty:
            df_over = df_over.sort_values('dias_cobertura', ascending=False)
            alerts.append({
                'severity': 2,
                'title': f"Sobre-stock — {len(df_over)} productos con >{CONFIG['overstock_days']} días cobertura",
                'caption': "Capital inmovilizado en exceso de inventario",
                'df': df_over[['producto', 'inv_total', 'avg_weekly', 'dias_cobertura']].rename(columns={
                    'producto': 'Producto', 'inv_total': 'Inventario',
                    'avg_weekly': 'Venta/Semana', 'dias_cobertura': 'Días Cobertura'
                }),
                'fmt': {'Venta/Semana': '{:,.0f}', 'Inventario': '{:,.0f}', 'Días Cobertura': '{:.0f}'},
            })

    # --- ALERT 4: Caída ventas producto (>25% vs semana anterior) ---
    if prev_week:
        df_drop = query_df("""
            WITH curr AS (
                SELECT item_nbr, producto, SUM(venta_und) as und_curr
                FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm = ?
                GROUP BY item_nbr
            ),
            prev AS (
                SELECT item_nbr, SUM(venta_und) as und_prev
                FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm = ?
                GROUP BY item_nbr
            )
            SELECT curr.item_nbr, curr.producto, curr.und_curr,
                   prev.und_prev
            FROM curr JOIN prev ON curr.item_nbr = prev.item_nbr
            WHERE prev.und_prev > 0
        """, [last_week, prev_week])
        if not df_drop.empty:
            df_drop['cambio_pct'] = (df_drop['und_curr'] - df_drop['und_prev']) / df_drop['und_prev']
            df_fell = df_drop[df_drop['cambio_pct'] <= -CONFIG["sales_drop_warning_pct"]].copy()
            if not df_fell.empty:
                df_fell = df_fell.sort_values('cambio_pct')
                df_crit_drop = df_fell[df_fell['cambio_pct'] <= -CONFIG["sales_drop_critical_pct"]]
                df_att_drop = df_fell[df_fell['cambio_pct'] > -CONFIG["sales_drop_critical_pct"]]
                for sev, subset, label in [
                    (0, df_crit_drop, f">{CONFIG['sales_drop_critical_pct']:.0%} caída"),
                    (1, df_att_drop, f"{CONFIG['sales_drop_warning_pct']:.0%}-{CONFIG['sales_drop_critical_pct']:.0%} caída"),
                ]:
                    if not subset.empty:
                        alerts.append({
                            'severity': sev,
                            'title': f"Caída Ventas Producto ({label}) — {len(subset)} SKUs",
                            'caption': f"{wm_week_to_date(last_week)} vs {wm_week_to_date(prev_week)}",
                            'df': subset[['producto', 'und_curr', 'und_prev', 'cambio_pct']].rename(columns={
                                'producto': 'Producto', 'und_curr': f'Und {wm_week_to_date_short(last_week)}',
                                'und_prev': f'Und {wm_week_to_date_short(prev_week)}', 'cambio_pct': 'Cambio %'
                            }),
                            'fmt': {'Cambio %': '{:+.0%}'},
                        })

    # --- ALERT 5: Caída ventas tienda (>30% vs promedio 4 semanas) ---
    if prev_4:
        ph_c4 = ','.join(['?'] * len(curr_4))
        ph_p4 = ','.join(['?'] * len(prev_4))
        df_store_drop = query_df(f"""
            WITH curr AS (
                SELECT store_nbr, tienda, SUM(venta_und) as und_curr
                FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm IN ({ph_c4})
                GROUP BY store_nbr
            ),
            prev AS (
                SELECT store_nbr, SUM(venta_und) as und_prev
                FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm IN ({ph_p4})
                GROUP BY store_nbr
            )
            SELECT curr.store_nbr, curr.tienda, curr.und_curr, prev.und_prev
            FROM curr JOIN prev ON curr.store_nbr = prev.store_nbr
            WHERE prev.und_prev > 0
        """, curr_4 + prev_4)
        if not df_store_drop.empty:
            df_store_drop['cambio_pct'] = (
                df_store_drop['und_curr'] - df_store_drop['und_prev']
            ) / df_store_drop['und_prev']
            df_sf = df_store_drop[df_store_drop['cambio_pct'] <= -CONFIG["store_drop_warning_pct"]].copy()
            if not df_sf.empty:
                df_sf = df_sf.sort_values('cambio_pct')
                df_crit_sf = df_sf[df_sf['cambio_pct'] <= -CONFIG["store_drop_critical_pct"]]
                df_att_sf = df_sf[df_sf['cambio_pct'] > -CONFIG["store_drop_critical_pct"]]
                for sev, subset, label in [
                    (0, df_crit_sf, f">{CONFIG['store_drop_critical_pct']:.0%} caída"),
                    (1, df_att_sf, f"{CONFIG['store_drop_warning_pct']:.0%}-{CONFIG['store_drop_critical_pct']:.0%} caída"),
                ]:
                    if not subset.empty:
                        alerts.append({
                            'severity': sev,
                            'title': f"Caída Ventas Tienda ({label}) — {len(subset)} tiendas",
                            'caption': f"Últimas 4 sem vs 4 sem anteriores",
                            'df': subset[['tienda', 'und_curr', 'und_prev', 'cambio_pct']].rename(columns={
                                'tienda': 'Tienda', 'und_curr': 'Und Recientes',
                                'und_prev': 'Und Anteriores', 'cambio_pct': 'Cambio %'
                            }),
                            'fmt': {'Cambio %': '{:+.0%}'},
                        })

    # --- ALERT 6: Tienda muerta (tenía ventas en prev_4, cero en curr_4) ---
    if prev_4:
        ph_c4 = ','.join(['?'] * len(curr_4))
        ph_p4 = ','.join(['?'] * len(prev_4))
        df_dead = query_df(f"""
            WITH prev AS (
                SELECT DISTINCT store_nbr, tienda
                FROM retail_link
                WHERE tipo_registro='VENTA' AND semana_wm IN ({ph_p4}) AND venta_und > 0
            ),
            curr AS (
                SELECT store_nbr, SUM(venta_und) as und_curr
                FROM retail_link
                WHERE tipo_registro='VENTA' AND semana_wm IN ({ph_c4})
                GROUP BY store_nbr
            )
            SELECT prev.store_nbr, prev.tienda,
                   COALESCE(curr.und_curr, 0) as und_curr
            FROM prev LEFT JOIN curr ON prev.store_nbr = curr.store_nbr
            WHERE COALESCE(curr.und_curr, 0) = 0
        """, prev_4 + curr_4)
        if not df_dead.empty:
            alerts.append({
                'severity': 0,
                'title': f"Tienda Muerta — {len(df_dead)} tiendas sin ventas en 4 semanas",
                'caption': "Tenían ventas en periodo anterior, ahora 0",
                'df': df_dead[['tienda']].rename(columns={'tienda': 'Tienda'}),
                'fmt': {},
            })

    # --- ALERT 7: Claims nuevos (desde última importación) ---
    claims_count = query_val("SELECT COUNT(*) FROM claims_rl")
    if claims_count and claims_count > 0:
        df_claims_new = query_df("""
            SELECT descripcion as Producto, tienda as Tienda,
                   SUM(cant_reclamo) as Unidades,
                   SUM(precio_reclamo_q) as "Monto Q",
                   semana_wm
            FROM claims_rl
            WHERE semana_wm = (SELECT MAX(semana_wm) FROM claims_rl)
            GROUP BY descripcion, tienda
            ORDER BY SUM(precio_reclamo_q) DESC
        """)
        if not df_claims_new.empty:
            _claims_week = df_claims_new['semana_wm'].iloc[0] if 'semana_wm' in df_claims_new.columns else None
            df_claims_new = df_claims_new.drop(columns=['semana_wm'], errors='ignore')
            total_q = df_claims_new['Monto Q'].sum()
            if total_q > CONFIG["claims_critical_q"]:
                sev = 0
            elif total_q >= CONFIG["claims_warning_q"]:
                sev = 1
            else:
                sev = 2
            alerts.append({
                'severity': sev,
                'title': f"Claims Última Semana — {fmt_q(total_q)} en {len(df_claims_new)} registros",
                'caption': f"Claims al {wm_week_to_date(str(_claims_week)) if _claims_week else 'última semana'}",
                'df': df_claims_new,
                'fmt': {'Monto Q': 'Q{:,.0f}', 'Unidades': '{:,.0f}'},
            })

    # --- ALERT 8: Tienda problema (>3 claims en últimas 4 semanas) ---
    if claims_count and claims_count > 0 and curr_4:
        ph_c4 = ','.join(['?'] * len(curr_4))
        df_prob = query_df(f"""
            SELECT tienda as Tienda, store_nbr,
                   COUNT(*) as "Total Claims",
                   SUM(cant_reclamo) as "Unidades",
                   SUM(precio_reclamo_q) as "Monto Q"
            FROM claims_rl
            WHERE semana_wm IN ({ph_c4})
            GROUP BY store_nbr
            HAVING COUNT(*) > {CONFIG["claims_store_warning"]}
            ORDER BY COUNT(*) DESC
        """, curr_4)
        if not df_prob.empty:
            for _, row in df_prob.iterrows():
                pass  # just checking non-empty
            max_claims = df_prob['Total Claims'].max()
            if max_claims > CONFIG["claims_store_critical"]:
                sev = 0
            else:
                sev = 1
            alerts.append({
                'severity': sev,
                'title': f"Tienda Problema — {len(df_prob)} tiendas con >{CONFIG['claims_store_warning']} claims recientes",
                'caption': f"Últimas {len(curr_4)} semanas de datos",
                'df': df_prob[['Tienda', 'Total Claims', 'Unidades', 'Monto Q']],
                'fmt': {'Monto Q': 'Q{:,.0f}', 'Unidades': '{:,.0f}'},
            })

    # ── Sort by severity ──
    alerts.sort(key=lambda a: a['severity'])

    # ── Summary counters ──
    n_crit = sum(1 for a in alerts if a['severity'] == 0)
    n_att = sum(1 for a in alerts if a['severity'] == 1)
    n_info = sum(1 for a in alerts if a['severity'] == 2)
    n_total = len(alerts)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Críticas", n_crit, delta=None)
    c2.metric("Atención", n_att, delta=None)
    c3.metric("Info", n_info, delta=None)
    c4.metric("Total", n_total, delta=None)

    st.caption(
        f"Ventas al {wm_week_to_date(last_week)}  •  "
        f"Inventario al {wm_week_to_date(inv_week) if inv_week else '—'}"
    )
    st.divider()

    # ── Empty state ──
    if not alerts:
        st.success("Todo en orden — sin alertas activas.")
        return

    # ── Download alerts summary ──
    if alerts:
        summary_rows = []
        for a in alerts:
            sev_label = ["CRÍTICA", "ATENCIÓN", "INFO"][a['severity']]
            summary_rows.append({'Severidad': sev_label, 'Alerta': a['title'], 'Detalle': a['caption']})
        alerts_summary_df = pd.DataFrame(summary_rows)
        download_df(alerts_summary_df, "alertas_resumen")

    # ── Render by severity group ──
    sev_config = [
        (0, "CRÍTICAS", "#C62828", "#FFEBEE"),
        (1, "ATENCIÓN", "#F57F17", "#FFF8E1"),
        (2, "INFO", "#1565C0", "#E3F2FD"),
    ]
    for sev_code, sev_label, fg, bg in sev_config:
        group = [a for a in alerts if a['severity'] == sev_code]
        if not group:
            continue
        icon = ["🔴", "🟡", "🔵"][sev_code]
        st.markdown(f"### {icon} {sev_label} ({len(group)})")

        for alert in group:
            expanded = sev_code == 0
            with st.expander(alert['title'], expanded=expanded):
                st.caption(alert['caption'])
                df_display = alert['df'].reset_index(drop=True)
                try:
                    styled = df_display.style.format(alert.get('fmt', {}), na_rep='—')
                    st.dataframe(styled, use_container_width=True, hide_index=True)
                except Exception:
                    st.dataframe(df_display, use_container_width=True, hide_index=True)
                # Action hints
                title_lower = alert['title'].lower()
                if 'oos' in title_lower or 'cobertura' in title_lower:
                    st.caption("→ Consultá la página **Sugerido** para ver el pedido recomendado")
                elif 'claims' in title_lower or 'tienda problema' in title_lower:
                    st.caption("→ Consultá la página **Claims** para ver detalle por fuente")
                elif 'caída' in title_lower or 'muerta' in title_lower:
                    st.caption("→ Consultá **Tendencias** para ver el contexto histórico")


# ═══════════════════════════════════════════
# PAGE: DASHBOARD
# ═══════════════════════════════════════════

def page_dashboard():
    st.markdown("## Dashboard Ejecutivo")

    # Date context
    max_week = query_val("SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='VENTA'")
    min_week = query_val("SELECT MIN(semana_wm) FROM retail_link WHERE tipo_registro='VENTA'")

    if not max_week:
        st.warning("No hay datos en la base. Subí archivos en la pestaña 'Upload'.")
        return

    st.caption(f"Datos: {wm_week_to_date(min_week)} — {wm_week_to_date(max_week)}")

    # Flexible period selector
    n_weeks = st.selectbox("Periodo", [4, 8, 12], index=0,
                           format_func=lambda x: f"Últimas {x} semanas", key="dash_period")

    # Get weeks for comparison
    weeks = query_df(f"""
        SELECT DISTINCT semana_wm FROM retail_link
        WHERE tipo_registro='VENTA' ORDER BY semana_wm DESC LIMIT {n_weeks * 2}
    """)['semana_wm'].tolist()

    curr_n = weeks[:n_weeks] if len(weeks) >= n_weeks else weeks
    prev_n = weeks[n_weeks:n_weeks*2] if len(weeks) >= n_weeks*2 else []

    # Aliases for backward compat within this function
    curr_4 = curr_n
    prev_4 = prev_n

    ph_curr = ','.join(['?'] * len(curr_4))

    # Current period KPIs
    curr = query_one(f"""
        SELECT SUM(venta_und), SUM(venta_q), SUM(venta_costo_q),
               COUNT(DISTINCT store_nbr), COUNT(DISTINCT item_nbr)
        FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm IN ({ph_curr})
    """, curr_4)

    prev = (0, 0, 0, 0, 0)
    if prev_4:
        ph_prev = ','.join(['?'] * len(prev_4))
        prev = query_one(f"""
            SELECT SUM(venta_und), SUM(venta_q), SUM(venta_costo_q),
                   COUNT(DISTINCT store_nbr), COUNT(DISTINCT item_nbr)
            FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm IN ({ph_prev})
        """, prev_4)

    inv = query_one("""
        SELECT SUM(inv_actual), COUNT(DISTINCT item_nbr)
        FROM retail_link WHERE tipo_registro='INVENTARIO'
        AND semana_wm = (SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO')
    """)

    si = query_one("""
        SELECT SUM(ingreso_bruto), SUM(costo_total), SUM(centralizacion),
               SUM(ingreso_neto), SUM(margen_bruto), SUM(cantidad_facturada)
        FROM sell_in
    """)

    # ── KPI CARDS ──
    st.markdown("### Sell-Out (Walmart vende al consumidor)")
    c1, c2, c3, c4, c5 = st.columns(5)

    venta_und = curr[0] or 0
    venta_q = curr[1] or 0
    prev_und = prev[0] or 0
    prev_q = prev[1] or 0
    delta_und = venta_und - prev_und
    delta_q = venta_q - prev_q

    c1.metric(f"Venta Und ({n_weeks} sem)", fmt_int(venta_und),
              delta=f"{delta_und:+,}" if prev_und else None)
    c2.metric(f"Venta Q ({n_weeks} sem)", fmt_q(venta_q),
              delta=f"{fmt_pct(delta_q / prev_q) if prev_q else ''}" if prev_q else None)
    c3.metric("Tiendas Activas", fmt_int(curr[3]),
              delta=f"{(curr[3] or 0) - (prev[3] or 0):+d}" if prev[3] else None)
    c4.metric("SKUs Activos", fmt_int(curr[4]))
    c5.metric("Inventario Actual", fmt_int(inv[0] if inv else 0),
              delta=f"{inv[1] or 0} SKUs")

    st.markdown("### Sell-In (Clan factura a Walmart)")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Ingreso Bruto", fmt_q(si[0] if si else 0))
    c2.metric("Centraliz. 4%", fmt_q(si[2] if si else 0))
    c3.metric("Ingreso Neto", fmt_q(si[3] if si else 0))
    margen = si[4] or 0
    ingreso = si[0] or 1
    c4.metric("Margen Bruto", fmt_q(margen), delta=f"{margen/ingreso:.1%}")

    # Sell-Through Rate (aggregate)
    st.markdown("### Sell-Through")
    st_data = query_one(f"""
        WITH so AS (
            SELECT SUM(venta_und) as und_out
            FROM retail_link
            WHERE tipo_registro='VENTA' AND semana_wm IN ({ph_curr})
        ),
        si_und AS (
            SELECT SUM(cantidad_facturada) as und_in FROM sell_in
        )
        SELECT so.und_out, si_und.und_in FROM so, si_und
    """, curr_4)
    if st_data and st_data[0] and st_data[1] and st_data[1] > 0:
        st_rate = st_data[0] / st_data[1] * 100
        st.metric("Sell-Through Rate", f"{st_rate:.1f}%",
                  delta=f"{fmt_int(st_data[0])} sell-out / {fmt_int(st_data[1])} sell-in")
    else:
        st.caption("Sin datos cruzados de sell-in/sell-out")
    st.caption(MARGIN_SELLOUT)

    st.divider()

    # ── TREND CHART: Sales with granularity toggle ──
    st.markdown("### Tendencia de Ventas")

    dash_gran_col, dash_n_col = st.columns([1, 2])
    with dash_gran_col:
        dash_granularity = st.radio("Vista", ["Diario", "Semanal", "Mensual"],
                                     index=1, horizontal=True, key="dash_gran")

    trend = query_df(build_trend_query(dash_granularity))

    if not trend.empty:
        trend['label'] = trend['periodo'].apply(lambda x: format_period_label(x, dash_granularity))

        with dash_n_col:
            n_points = st.select_slider(
                "Puntos a mostrar",
                options=[14, 30, 60, 90, len(trend)] if dash_granularity == "Diario"
                    else [8, 12, 16, 24, 52, len(trend)] if dash_granularity == "Semanal"
                    else [3, 6, 12, len(trend)],
                value=min(30 if dash_granularity == "Diario" else 24 if dash_granularity == "Semanal" else 12, len(trend)),
                format_func=lambda x: f"{x}d" if dash_granularity == "Diario" and x < len(trend)
                    else f"{x} sem" if dash_granularity == "Semanal" and x < len(trend)
                    else f"{x}m" if dash_granularity == "Mensual" and x < len(trend)
                    else "Todo",
                key="dash_n"
            )
        trend_show = trend.tail(n_points)

        fig = make_subplots(specs=[[{"secondary_y": True}]])

        fig.add_trace(
            go.Bar(x=trend_show['label'], y=trend_show['q'],
                   name="Venta Q", marker_color=CLAN_BLUE, opacity=0.7),
            secondary_y=False
        )
        fig.add_trace(
            go.Scatter(x=trend_show['label'], y=trend_show['und'],
                       name="Venta Und", line=dict(color=GREEN, width=2),
                       mode='lines+markers'),
            secondary_y=True
        )

        fig.update_layout(
            height=400,
            margin=dict(l=20, r=20, t=30, b=20),
            legend=dict(orientation="h", y=1.12),
            hovermode="x unified"
        )
        fig.update_yaxes(title_text="Venta Q", secondary_y=False, tickformat=",")
        fig.update_yaxes(title_text="Unidades", secondary_y=True, tickformat=",")

        st.plotly_chart(fig, use_container_width=True)

    st.divider()

    # ── TOP PRODUCTS & STORES (side by side) ──
    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown(f"### Top 10 Productos ({n_weeks} sem)")
        top_prods = query_df(f"""
            SELECT rl.producto as Producto, p.marca as Marca,
                   SUM(rl.venta_und) as Und, SUM(rl.venta_q) as Q
            FROM retail_link rl
            LEFT JOIN productos p ON rl.producto = p.producto
            WHERE rl.tipo_registro='VENTA' AND rl.semana_wm IN ({ph_curr})
            GROUP BY rl.producto
            ORDER BY Und DESC
            LIMIT 10
        """, curr_4)

        if not top_prods.empty:
            fig_p = px.bar(top_prods, x='Und', y='Producto', orientation='h',
                          color='Marca', color_discrete_sequence=px.colors.qualitative.Set2,
                          height=350)
            fig_p.update_layout(margin=dict(l=10, r=10, t=10, b=10),
                               yaxis={'categoryorder': 'total ascending'},
                               showlegend=True, legend=dict(orientation="h", y=-0.15))
            st.plotly_chart(fig_p, use_container_width=True)

    with col_right:
        st.markdown(f"### Top 10 Tiendas ({n_weeks} sem)")
        top_stores = query_df(f"""
            SELECT rl.tienda as Tienda, t.formato as Formato,
                   SUM(rl.venta_und) as Und, SUM(rl.venta_q) as Q
            FROM retail_link rl
            LEFT JOIN tiendas t ON rl.store_nbr = t.no_tienda
            WHERE rl.tipo_registro='VENTA' AND rl.semana_wm IN ({ph_curr})
            GROUP BY rl.tienda
            ORDER BY Und DESC
            LIMIT 10
        """, curr_4)

        if not top_stores.empty:
            fig_s = px.bar(top_stores, x='Und', y='Tienda', orientation='h',
                          color='Formato', color_discrete_sequence=px.colors.qualitative.Pastel,
                          height=350)
            fig_s.update_layout(margin=dict(l=10, r=10, t=10, b=10),
                               yaxis={'categoryorder': 'total ascending'},
                               showlegend=True, legend=dict(orientation="h", y=-0.15))
            st.plotly_chart(fig_s, use_container_width=True)

    # ── PARETO: Revenue concentration ──
    st.divider()
    col_pareto, col_formato = st.columns(2)

    with col_pareto:
        st.markdown("### Concentracion de Revenue (Pareto)")
        pareto = query_df(f"""
            SELECT rl.producto as Producto, p.marca as Marca,
                   SUM(rl.venta_q) as Q
            FROM retail_link rl
            LEFT JOIN productos p ON rl.producto = p.producto
            WHERE rl.tipo_registro='VENTA' AND rl.semana_wm IN ({ph_curr})
            GROUP BY rl.producto ORDER BY Q DESC
        """, curr_4)

        if not pareto.empty:
            total_revenue = pareto['Q'].sum()
            pareto['% Acum'] = pareto['Q'].cumsum() / total_revenue * 100
            pareto['rank'] = range(1, len(pareto) + 1)

            # Find how many SKUs make 80%
            sku_80 = pareto[pareto['% Acum'] <= 80].shape[0] + 1
            st.caption(f"**{sku_80} de {len(pareto)} SKUs** generan el 80% del revenue ({len(curr_4)} sem)")

            fig_pareto = make_subplots(specs=[[{"secondary_y": True}]])
            fig_pareto.add_trace(
                go.Bar(x=pareto['Producto'].head(10), y=pareto['Q'].head(10),
                       name="Venta Q", marker_color=CLAN_BLUE, opacity=0.8),
                secondary_y=False)
            fig_pareto.add_trace(
                go.Scatter(x=pareto['Producto'].head(10), y=pareto['% Acum'].head(10),
                           name="% Acumulado", line=dict(color=RED, width=2.5),
                           mode='lines+markers'),
                secondary_y=True)
            fig_pareto.add_hline(y=80, line_dash="dash", line_color="gray",
                                annotation_text="80%", secondary_y=True)
            fig_pareto.update_layout(height=350, margin=dict(l=10, r=10, t=10, b=10),
                                     legend=dict(orientation="h", y=1.12))
            fig_pareto.update_yaxes(title_text="Venta Q", tickformat=",", secondary_y=False)
            fig_pareto.update_yaxes(title_text="% Acumulado", tickformat=".0f", secondary_y=True)
            st.plotly_chart(fig_pareto, use_container_width=True)

    with col_formato:
        st.markdown("### Ventas por Formato de Tienda")
        by_format = query_df(f"""
            SELECT COALESCE(t.formato, 'Sin formato') as Formato,
                   SUM(rl.venta_und) as Und, SUM(rl.venta_q) as Q,
                   COUNT(DISTINCT rl.store_nbr) as Tiendas
            FROM retail_link rl
            LEFT JOIN tiendas t ON rl.store_nbr = t.no_tienda
            WHERE rl.tipo_registro='VENTA' AND rl.semana_wm IN ({ph_curr})
            GROUP BY t.formato ORDER BY Q DESC
        """, curr_4)

        if not by_format.empty:
            fig_fmt = px.pie(by_format, values='Q', names='Formato',
                            color_discrete_sequence=px.colors.qualitative.Set2, height=300)
            fig_fmt.update_layout(margin=dict(l=10, r=10, t=10, b=10))
            st.plotly_chart(fig_fmt, use_container_width=True)

            by_format['Q/Tienda'] = by_format['Q'] / by_format['Tiendas'].replace(0, 1)
            by_format['Und/Tienda'] = by_format['Und'] / by_format['Tiendas'].replace(0, 1)
            st.dataframe(by_format.style.format({
                'Und': '{:,.0f}', 'Q': 'Q{:,.0f}', 'Q/Tienda': 'Q{:,.0f}', 'Und/Tienda': '{:,.0f}'
            }), use_container_width=True, hide_index=True)

    # ── ALERTS SUMMARY (compact — full detail on Alertas page) ──
    st.divider()
    st.markdown("### Alertas")

    # Quick severity counts reusing alertas logic
    _alert_weeks = query_df("""
        SELECT DISTINCT semana_wm FROM retail_link
        WHERE tipo_registro='VENTA' ORDER BY semana_wm DESC LIMIT 8
    """)['semana_wm'].tolist()
    _inv_week = query_val("SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO'")
    _n_crit, _n_att, _n_info = 0, 0, 0

    if _inv_week and _alert_weeks:
        _a4 = _alert_weeks[:4]
        _ph4 = ','.join(['?'] * len(_a4))
        # OOS check
        _oos_df = query_df(f"""
            SELECT item_nbr, COUNT(DISTINCT store_nbr) as t,
                   SUM(CASE WHEN inv_actual=0 THEN 1 ELSE 0 END) as z
            FROM retail_link
            WHERE tipo_registro='INVENTARIO' AND semana_wm=?
              AND item_nbr IN (SELECT item_nbr FROM retail_link WHERE tipo_registro='VENTA'
                               AND semana_wm IN ({_ph4}) GROUP BY item_nbr HAVING SUM(venta_und)>={CONFIG["min_sales_for_oos"]})
            GROUP BY item_nbr HAVING z>0
        """, [_inv_week] + _a4)
        if not _oos_df.empty:
            _oos_df['p'] = _oos_df['z'] / _oos_df['t']
            _n_crit += len(_oos_df[_oos_df['p'] >= CONFIG["oos_pct_threshold"]])

        # Coverage check
        _cov_df = query_df(f"""
            WITH inv AS (SELECT item_nbr, SUM(inv_actual) as i FROM retail_link
                         WHERE tipo_registro='INVENTARIO' AND semana_wm=? GROUP BY item_nbr),
                 sal AS (SELECT item_nbr, SUM(venta_und) as v, COUNT(DISTINCT semana_wm) as n
                         FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm IN ({_ph4})
                         GROUP BY item_nbr)
            SELECT inv.item_nbr, inv.i, COALESCE(sal.v,0) as v, sal.n
            FROM inv LEFT JOIN sal ON inv.item_nbr=sal.item_nbr
        """, [_inv_week] + _a4)
        if not _cov_df.empty:
            _cov_df['aw'] = _cov_df.apply(lambda r: r['v']/r['n'] if r['n'] and r['n']>0 else 0, axis=1)
            _cov_df['dc'] = _cov_df.apply(lambda r: (r['i']*7/r['aw']) if r['aw']>0 else 999, axis=1)
            _n_crit += len(_cov_df[_cov_df['dc'] <= CONFIG["coverage_critical_days"]])
            _n_att += len(_cov_df[(_cov_df['dc'] > CONFIG["coverage_critical_days"]) & (_cov_df['dc'] < CONFIG["coverage_warning_days"])])

    _claims_q = query_val("SELECT SUM(precio_reclamo_q) FROM claims_rl WHERE semana_wm=(SELECT MAX(semana_wm) FROM claims_rl)") or 0
    if _claims_q > CONFIG["claims_critical_q"]:
        _n_crit += 1
    elif _claims_q >= CONFIG["claims_warning_q"]:
        _n_att += 1
    elif _claims_q > 0:
        _n_info += 1

    ac1, ac2, ac3 = st.columns(3)
    ac1.metric("Críticas", _n_crit)
    ac2.metric("Atención", _n_att)
    ac3.metric("Info", _n_info)
    st.caption("Ver detalle completo en la página **Alertas**")


# ═══════════════════════════════════════════
# PAGE: WEEKLY REPORT
# ═══════════════════════════════════════════

def page_semanal():
    st.markdown("## Reporte Semanal")

    weeks_df = query_df("""
        SELECT DISTINCT semana_wm FROM retail_link
        WHERE tipo_registro='VENTA' ORDER BY semana_wm DESC
    """)
    all_weeks = weeks_df['semana_wm'].tolist()

    if not all_weeks:
        st.warning("No hay datos de ventas.")
        return

    # Week selector
    week_labels = {w: wm_week_to_date(w) for w in all_weeks[:20]}
    selected = st.selectbox("Semana", options=list(week_labels.keys()),
                            format_func=lambda w: week_labels[w])

    idx = all_weeks.index(selected)
    prev_week = all_weeks[idx + 1] if idx + 1 < len(all_weeks) else None

    # Week comparison
    curr = query_one("""
        SELECT SUM(venta_und), SUM(venta_q), SUM(venta_costo_q),
               COUNT(DISTINCT store_nbr), COUNT(DISTINCT item_nbr)
        FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm=?
    """, [selected])

    prev = (0, 0, 0, 0, 0)
    if prev_week:
        prev = query_one("""
            SELECT SUM(venta_und), SUM(venta_q), SUM(venta_costo_q),
                   COUNT(DISTINCT store_nbr), COUNT(DISTINCT item_nbr)
            FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm=?
        """, [prev_week])

    # YoY comparison
    week_num = selected[4:]  # "04" from "202604"
    prior_year = str(int(selected[:4]) - 1)
    yoy_week = prior_year + week_num
    yoy = query_one("""
        SELECT SUM(venta_und), SUM(venta_q)
        FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm=?
    """, [yoy_week])

    st.markdown(f"### {wm_week_to_date(selected)} vs {wm_week_to_date(prev_week) if prev_week else 'N/A'}")

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Venta Und", fmt_int(curr[0]),
              delta=f"{(curr[0] or 0) - (prev[0] or 0):+,}" if prev[0] else None)
    c2.metric("Venta Q", fmt_q(curr[1]),
              delta=f"{fmt_pct(((curr[1] or 0) - (prev[1] or 0)) / (prev[1] or 1))}" if prev[1] else None)
    c3.metric("Tiendas", fmt_int(curr[3]),
              delta=f"{(curr[3] or 0) - (prev[3] or 0):+d}" if prev[3] else None)
    c4.metric("SKUs", fmt_int(curr[4]))
    if yoy and yoy[0]:
        yoy_delta = ((curr[0] or 0) - (yoy[0] or 0)) / (yoy[0] or 1)
        c5.metric(f"vs {prior_year}", fmt_int(yoy[0]),
                  delta=f"{yoy_delta:+.1%}")
    else:
        c5.metric(f"vs {prior_year}", "—", delta="Sin datos")

    st.divider()

    # Products this week
    st.markdown(f"### Ventas por Producto — {wm_week_to_date(selected)}")

    prods = query_df("""
        SELECT rl.producto as Producto, p.marca as Marca,
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_und ELSE 0 END) as "Und Actual",
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_q ELSE 0 END) as "Q Actual",
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_und ELSE 0 END) as "Und Anterior"
        FROM retail_link rl
        LEFT JOIN productos p ON rl.producto = p.producto
        WHERE rl.tipo_registro='VENTA' AND rl.semana_wm IN (?,?)
        GROUP BY rl.producto
        HAVING "Und Actual" > 0 OR "Und Anterior" > 0
        ORDER BY "Und Actual" DESC
    """, [selected, selected, prev_week or '', selected, prev_week or ''])

    if not prods.empty:
        prods['Var %'] = prods.apply(
            lambda r: (r['Und Actual'] - r['Und Anterior']) / r['Und Anterior']
            if r['Und Anterior'] > 0 else 0, axis=1)

        # Chart
        fig = px.bar(prods.head(15), x='Producto', y=['Und Actual', 'Und Anterior'],
                     barmode='group', color_discrete_sequence=[CLAN_BLUE, '#B0BEC5'],
                     height=350)
        fig.update_layout(margin=dict(l=10, r=10, t=10, b=10),
                         legend=dict(orientation="h", y=1.1))
        st.plotly_chart(fig, use_container_width=True)

        # Table
        st.dataframe(
            prods.style.format({
                'Und Actual': '{:,.0f}', 'Q Actual': 'Q{:,.0f}',
                'Und Anterior': '{:,.0f}', 'Var %': '{:+.1%}'
            }),
            use_container_width=True, hide_index=True
        )
        download_df(prods, f"semanal_productos_{selected}")

    # Stores this week
    st.divider()
    st.markdown(f"### Top 15 Tiendas — {wm_week_to_date(selected)}")

    stores = query_df("""
        SELECT rl.tienda as Tienda, t.formato as Formato,
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_und ELSE 0 END) as "Und Actual",
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_q ELSE 0 END) as "Q Actual",
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_und ELSE 0 END) as "Und Anterior"
        FROM retail_link rl
        LEFT JOIN tiendas t ON rl.store_nbr = t.no_tienda
        WHERE rl.tipo_registro='VENTA' AND rl.semana_wm IN (?,?)
        GROUP BY rl.tienda
        ORDER BY "Und Actual" DESC
        LIMIT 15
    """, [selected, selected, prev_week or '', selected, prev_week or ''])

    if not stores.empty:
        stores['Var %'] = stores.apply(
            lambda r: (r['Und Actual'] - r['Und Anterior']) / r['Und Anterior']
            if r['Und Anterior'] > 0 else 0, axis=1)

        st.dataframe(
            stores.style.format({
                'Und Actual': '{:,.0f}', 'Q Actual': 'Q{:,.0f}',
                'Und Anterior': '{:,.0f}', 'Var %': '{:+.1%}'
            }),
            use_container_width=True, hide_index=True
        )
        download_df(stores, f"semanal_tiendas_{selected}")


# ═══════════════════════════════════════════
# PAGE: TRENDS
# ═══════════════════════════════════════════

def page_tendencias():
    st.markdown("## Tendencias Históricas")

    period = st.radio("Período", ["Diario", "Semanal", "Mensual", "Trimestral", "Anual"],
                      index=1, horizontal=True)

    if period == "Diario":
        data = query_df("""
            SELECT diario as periodo,
                   SUM(venta_und) as und,
                   SUM(venta_q) as q,
                   SUM(venta_costo_q) as costo_q,
                   COUNT(DISTINCT store_nbr) as tiendas,
                   COUNT(DISTINCT item_nbr) as skus
            FROM retail_link
            WHERE tipo_registro='VENTA' AND diario IS NOT NULL
            GROUP BY diario ORDER BY diario
        """)
        if data.empty:
            st.info("Sin datos.")
            return
        data['label'] = data['periodo'].apply(lambda x: format_period_label(x, "Diario"))
        x_col = 'label'

        n_days = st.select_slider("Dias a mostrar", options=[14, 30, 60, 90, 180, len(data)],
                                   value=min(30, len(data)),
                                   format_func=lambda x: f"{x}d" if x < len(data) else "Todo",
                                   key="tend_daily_n")
        data = data.tail(n_days)

    elif period == "Semanal":
        data = query_df("""
            SELECT semana_wm,
                   SUM(venta_und) as und,
                   SUM(venta_q) as q,
                   SUM(venta_costo_q) as costo_q,
                   COUNT(DISTINCT store_nbr) as tiendas,
                   COUNT(DISTINCT item_nbr) as skus
            FROM retail_link
            WHERE tipo_registro='VENTA'
            GROUP BY semana_wm ORDER BY semana_wm
        """)
        if data.empty:
            st.info("Sin datos.")
            return
        data['periodo'] = data['semana_wm'].apply(wm_week_to_date_short)
        data['periodo_dt'] = data['semana_wm'].apply(wm_week_to_datetime)
        x_col = 'periodo'

    elif period == "Mensual":
        data = query_df("""
            SELECT anio || '-' || printf('%02d', mes) as periodo,
                   SUM(venta_und) as und,
                   SUM(venta_q) as q,
                   SUM(venta_costo_q) as costo_q,
                   COUNT(DISTINCT store_nbr) as tiendas,
                   COUNT(DISTINCT item_nbr) as skus
            FROM retail_link
            WHERE tipo_registro='VENTA' AND anio IS NOT NULL
            GROUP BY anio, mes ORDER BY anio, mes
        """)
        x_col = 'periodo'

    elif period == "Trimestral":
        data = query_df("""
            SELECT anio || '-Q' || ((mes - 1) / 3 + 1) as periodo,
                   SUM(venta_und) as und,
                   SUM(venta_q) as q,
                   SUM(venta_costo_q) as costo_q,
                   COUNT(DISTINCT store_nbr) as tiendas
            FROM retail_link
            WHERE tipo_registro='VENTA' AND anio IS NOT NULL
            GROUP BY anio, (mes - 1) / 3
            ORDER BY anio, (mes - 1) / 3
        """)
        x_col = 'periodo'

    else:  # Anual
        data = query_df("""
            SELECT CAST(anio AS TEXT) as periodo,
                   SUM(venta_und) as und,
                   SUM(venta_q) as q,
                   SUM(venta_costo_q) as costo_q,
                   COUNT(DISTINCT store_nbr) as tiendas
            FROM retail_link
            WHERE tipo_registro='VENTA' AND anio IS NOT NULL
            GROUP BY anio ORDER BY anio
        """)
        x_col = 'periodo'

    if data.empty:
        st.info("Sin datos para este período.")
        return

    # Calculate margins and variations
    data['margen_q'] = data['q'] - data['costo_q'].fillna(0)
    data['margen_pct'] = data['margen_q'] / data['q'].replace(0, 1)
    data['var_q'] = data['q'].pct_change()
    data['var_und'] = data['und'].pct_change()

    # ── Main chart: Revenue + Units ──
    st.markdown("### Ventas")
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Bar(x=data[x_col], y=data['q'], name="Venta Q",
               marker_color=CLAN_BLUE, opacity=0.8),
        secondary_y=False
    )
    fig.add_trace(
        go.Scatter(x=data[x_col], y=data['und'], name="Unidades",
                   line=dict(color=GREEN, width=2.5), mode='lines+markers'),
        secondary_y=True
    )
    fig.update_layout(height=400, margin=dict(l=20, r=20, t=30, b=20),
                     legend=dict(orientation="h", y=1.1), hovermode="x unified")
    fig.update_yaxes(title_text="Venta Q", tickformat=",", secondary_y=False)
    fig.update_yaxes(title_text="Unidades", tickformat=",", secondary_y=True)
    st.plotly_chart(fig, use_container_width=True)

    # ── Variation chart ──
    st.markdown("### Variación %")
    fig2 = go.Figure()
    colors = [GREEN if v >= 0 else RED for v in data['var_q'].fillna(0)]
    fig2.add_trace(go.Bar(x=data[x_col], y=data['var_q'], name="Var % Q",
                          marker_color=colors))
    fig2.update_layout(height=250, margin=dict(l=20, r=20, t=10, b=20),
                      yaxis_tickformat=".0%")
    st.plotly_chart(fig2, use_container_width=True)

    # ── Margin trend ──
    st.markdown("### Margen Walmart")
    st.caption(MARGIN_SELLOUT)
    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=data[x_col], y=data['margen_pct'],
                              name="Margen %", line=dict(color=PURPLE, width=2),
                              mode='lines+markers', fill='tozeroy',
                              fillcolor='rgba(142,36,170,0.1)'))
    fig3.update_layout(height=250, margin=dict(l=20, r=20, t=10, b=20),
                      yaxis_tickformat=".0%")
    st.plotly_chart(fig3, use_container_width=True)

    # ── YoY Comparison ──
    if period == "Mensual" and len(data) > 12:
        st.divider()
        st.markdown("### Comparativo Year-over-Year")

        data['anio'] = data['periodo'].str[:4]
        data['mes_num'] = data['periodo'].str[5:].astype(int)

        years = sorted(data['anio'].unique())
        if len(years) >= 2:
            fig_yoy = go.Figure()
            month_names = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                          'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
            colors_yoy = [CLAN_BLUE, GREEN, ORANGE, RED]
            for i, yr in enumerate(years):
                yr_data = data[data['anio'] == yr].sort_values('mes_num')
                fig_yoy.add_trace(go.Scatter(
                    x=[month_names[m-1] for m in yr_data['mes_num']],
                    y=yr_data['q'], name=yr,
                    line=dict(width=2.5, color=colors_yoy[i % len(colors_yoy)]),
                    mode='lines+markers'
                ))
            fig_yoy.update_layout(height=350, margin=dict(l=20, r=20, t=10, b=20),
                                 yaxis_tickformat=",", hovermode="x unified")
            st.plotly_chart(fig_yoy, use_container_width=True)

    # Data table
    with st.expander("Ver datos"):
        st.dataframe(data[[x_col, 'und', 'q', 'margen_q', 'margen_pct']].style.format({
            'und': '{:,.0f}', 'q': 'Q{:,.0f}',
            'margen_q': 'Q{:,.0f}', 'margen_pct': '{:.1%}'
        }), use_container_width=True, hide_index=True)

    # ── Sell-Through Rate (Sell-Out / Sell-In) ──
    st.divider()
    st.markdown("### Sell-Through Rate (Sell-Out vs Sell-In)")
    st.caption("Qué porcentaje de lo facturado a Walmart ya se vendió al consumidor")

    sell_through = query_df("""
        WITH so_data AS (
            SELECT
                REPLACE(SUBSTR(diario, 1, 7), '/', '-') as mes,
                p.codigo as codigo,
                COALESCE(p.producto, rl.producto) as nombre,
                SUM(rl.venta_und) as und_out
            FROM retail_link rl
            LEFT JOIN productos p ON rl.item_nbr = p.item_wm
            WHERE rl.tipo_registro = 'VENTA' AND rl.diario IS NOT NULL
                AND p.codigo IS NOT NULL
            GROUP BY REPLACE(SUBSTR(diario, 1, 7), '/', '-'), p.codigo
        ),
        si_data AS (
            SELECT
                SUBSTR(fecha, 1, 7) as mes,
                codigo_producto as codigo,
                SUM(cantidad_facturada) as und_in
            FROM sell_in
            WHERE fecha IS NOT NULL
            GROUP BY SUBSTR(fecha, 1, 7), codigo_producto
        )
        SELECT
            COALESCE(so.mes, si.mes) as Mes,
            COALESCE(so.nombre, si.codigo) as Producto,
            COALESCE(si.und_in, 0) as "Sell-In (Und)",
            COALESCE(so.und_out, 0) as "Sell-Out (Und)",
            CASE WHEN COALESCE(si.und_in, 0) > 0
                THEN ROUND(COALESCE(so.und_out, 0) * 100.0 / si.und_in, 1)
                ELSE NULL END as "ST Rate %"
        FROM so_data so
        LEFT JOIN si_data si ON so.mes = si.mes AND so.codigo = si.codigo
        WHERE COALESCE(si.und_in, 0) > 0 OR COALESCE(so.und_out, 0) > 0
        UNION
        SELECT
            si.mes, si.codigo, si.und_in, 0, NULL
        FROM si_data si
        LEFT JOIN so_data so ON so.mes = si.mes AND so.codigo = si.codigo
        WHERE so.mes IS NULL
        ORDER BY 1, 2
    """)

    if not sell_through.empty:
        # Aggregate by month
        st_monthly = sell_through.groupby('Mes').agg({
            'Sell-In (Und)': 'sum',
            'Sell-Out (Und)': 'sum'
        }).reset_index()
        st_monthly['ST Rate %'] = (st_monthly['Sell-Out (Und)'] / st_monthly['Sell-In (Und)'].replace(0, 1) * 100).round(1)

        fig_st = make_subplots(specs=[[{"secondary_y": True}]])
        fig_st.add_trace(
            go.Bar(x=st_monthly['Mes'], y=st_monthly['Sell-In (Und)'],
                   name="Sell-In", marker_color=GREEN, opacity=0.7),
            secondary_y=False
        )
        fig_st.add_trace(
            go.Bar(x=st_monthly['Mes'], y=st_monthly['Sell-Out (Und)'],
                   name="Sell-Out", marker_color=CLAN_BLUE, opacity=0.7),
            secondary_y=False
        )
        fig_st.add_trace(
            go.Scatter(x=st_monthly['Mes'], y=st_monthly['ST Rate %'],
                       name="ST Rate %", line=dict(color=RED, width=2.5),
                       mode='lines+markers'),
            secondary_y=True
        )
        fig_st.update_layout(height=350, barmode='group',
                             margin=dict(l=20, r=20, t=30, b=20),
                             legend=dict(orientation="h", y=1.1),
                             hovermode="x unified")
        fig_st.update_yaxes(title_text="Unidades", tickformat=",", secondary_y=False)
        fig_st.update_yaxes(title_text="ST Rate %", tickformat=".0f", secondary_y=True)
        st.plotly_chart(fig_st, use_container_width=True)

        # ST by product
        with st.expander("Sell-Through por Producto"):
            st_prod = sell_through.groupby('Producto').agg({
                'Sell-In (Und)': 'sum',
                'Sell-Out (Und)': 'sum'
            }).reset_index()
            st_prod['ST Rate %'] = (st_prod['Sell-Out (Und)'] / st_prod['Sell-In (Und)'].replace(0, 1) * 100).round(1)
            st_prod = st_prod.sort_values('ST Rate %', ascending=False)
            st.dataframe(st_prod.style.format({
                'Sell-In (Und)': '{:,.0f}', 'Sell-Out (Und)': '{:,.0f}',
                'ST Rate %': '{:.1f}%'
            }), use_container_width=True, hide_index=True)
            download_df(st_prod, "sell_through_producto")
    else:
        st.info("No hay datos cruzados de Sell-In y Sell-Out para calcular Sell-Through.")

    # ── Seasonality Index ──
    st.divider()
    st.markdown("### Indice de Estacionalidad")
    st.caption("Qué meses son los más fuertes vs el promedio. >100 = arriba del promedio.")

    seasonality = query_df("""
        SELECT mes,
               SUM(venta_und) as total_und,
               SUM(venta_q) as total_q,
               COUNT(DISTINCT anio) as anios
        FROM retail_link
        WHERE tipo_registro = 'VENTA' AND mes IS NOT NULL AND anio IS NOT NULL
        GROUP BY mes
        ORDER BY mes
    """)

    if not seasonality.empty and len(seasonality) > 1:
        avg_und = seasonality['total_und'].mean()
        avg_q = seasonality['total_q'].mean()
        seasonality['idx_und'] = (seasonality['total_und'] / avg_und * 100).round(1)
        seasonality['idx_q'] = (seasonality['total_q'] / avg_q * 100).round(1)
        month_names = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                      'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        seasonality['Mes'] = seasonality['mes'].apply(
            lambda m: month_names[int(m)-1] if m and 1 <= int(m) <= 12 else str(m))

        fig_seas = go.Figure()
        colors_s = [GREEN if v >= 100 else RED for v in seasonality['idx_q']]
        fig_seas.add_trace(go.Bar(
            x=seasonality['Mes'], y=seasonality['idx_q'],
            marker_color=colors_s, text=seasonality['idx_q'].apply(lambda x: f"{x:.0f}"),
            textposition='outside'
        ))
        fig_seas.add_hline(y=100, line_dash="dash", line_color="gray", opacity=0.5)
        fig_seas.update_layout(height=300, margin=dict(l=20, r=20, t=30, b=20),
                               yaxis_title="Indice (100 = promedio)",
                               showlegend=False)
        st.plotly_chart(fig_seas, use_container_width=True)

        with st.expander("Datos de estacionalidad"):
            seas_display = seasonality[['Mes', 'total_und', 'total_q', 'anios', 'idx_und', 'idx_q']].rename(columns={
                'total_und': 'Und Total', 'total_q': 'Q Total',
                'anios': 'Años', 'idx_und': 'Idx Und', 'idx_q': 'Idx Q'
            })
            st.dataframe(seas_display.style.format({
                'Und Total': '{:,.0f}', 'Q Total': 'Q{:,.0f}',
                'Idx Und': '{:.1f}', 'Idx Q': '{:.1f}'
            }), use_container_width=True, hide_index=True)
            download_df(seas_display, "estacionalidad")


# ═══════════════════════════════════════════
# PAGE: PRODUCTS
# ═══════════════════════════════════════════

def page_productos():
    st.markdown("## Análisis por Producto")

    # Get all products with sales
    products = query_df("""
        SELECT DISTINCT rl.producto, p.marca
        FROM retail_link rl
        LEFT JOIN productos p ON rl.producto = p.producto
        WHERE rl.tipo_registro='VENTA' AND rl.venta_und > 0
        ORDER BY rl.producto
    """)

    if products.empty:
        st.info("Sin datos de productos.")
        return

    # Brand filter
    marcas = ['Todas'] + sorted(products['marca'].dropna().unique().tolist())
    marca_sel = st.selectbox("Filtrar por marca", marcas)

    if marca_sel != 'Todas':
        products = products[products['marca'] == marca_sel]

    selected_prod = st.selectbox("Producto", products['producto'].tolist())

    if not selected_prod:
        return

    st.divider()

    # Product detail
    trend = query_df("""
        SELECT semana_wm,
               SUM(venta_und) as und,
               SUM(venta_q) as q,
               SUM(inv_actual) as inv,
               COUNT(DISTINCT store_nbr) as tiendas
        FROM retail_link
        WHERE producto=? AND tipo_registro='VENTA'
        GROUP BY semana_wm ORDER BY semana_wm
    """, [selected_prod])

    inv_trend = query_df("""
        SELECT semana_wm, SUM(inv_actual) as inv
        FROM retail_link
        WHERE producto=? AND tipo_registro='INVENTARIO'
        GROUP BY semana_wm ORDER BY semana_wm
    """, [selected_prod])

    if trend.empty:
        st.info(f"Sin datos de venta para {selected_prod}")
        return

    trend['fecha'] = trend['semana_wm'].apply(wm_week_to_date_short)

    # KPIs
    total_und = trend['und'].sum()
    total_q = trend['q'].sum()
    last_inv = inv_trend['inv'].iloc[-1] if not inv_trend.empty else 0
    avg_weekly = trend['und'].mean()
    coverage = last_inv / (avg_weekly / 7) if avg_weekly > 0 else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Venta Total Und", fmt_int(total_und))
    c2.metric("Venta Total Q", fmt_q(total_q))
    c3.metric("Inventario Actual", fmt_int(last_inv))
    c4.metric("Cobertura", f"{coverage:.0f} dias")

    # Sales trend chart with granularity toggle
    st.markdown(f"### Ventas — {selected_prod}")

    prod_gran_col, prod_n_col = st.columns([1, 2])
    with prod_gran_col:
        prod_gran = st.radio("Vista", ["Diario", "Semanal", "Mensual"],
                              index=1, horizontal=True, key="prod_gran")

    if prod_gran == "Diario":
        prod_trend = query_df("""
            SELECT diario as periodo, SUM(venta_und) as und, SUM(venta_q) as q
            FROM retail_link
            WHERE producto=? AND tipo_registro='VENTA' AND diario IS NOT NULL
            GROUP BY diario ORDER BY diario
        """, [selected_prod])
    elif prod_gran == "Mensual":
        prod_trend = query_df("""
            SELECT anio || '-' || printf('%02d', mes) as periodo,
                   SUM(venta_und) as und, SUM(venta_q) as q
            FROM retail_link
            WHERE producto=? AND tipo_registro='VENTA' AND anio IS NOT NULL
            GROUP BY anio, mes ORDER BY anio, mes
        """, [selected_prod])
    else:
        prod_trend = trend.rename(columns={'semana_wm': 'periodo'})[['periodo', 'und', 'q']].copy()

    if not prod_trend.empty:
        prod_trend['label'] = prod_trend['periodo'].apply(lambda x: format_period_label(x, prod_gran))

        with prod_n_col:
            opts = [14, 30, 60, 90, len(prod_trend)] if prod_gran == "Diario" \
                else [8, 12, 24, 52, len(prod_trend)] if prod_gran == "Semanal" \
                else [3, 6, 12, len(prod_trend)]
            n = st.select_slider("Puntos", opts,
                                  value=min(opts[-2] if len(opts) > 1 else opts[0], len(prod_trend)),
                                  key="prod_n",
                                  format_func=lambda x: "Todo" if x >= len(prod_trend) else str(x))
        show = prod_trend.tail(n)

        fig = make_subplots(specs=[[{"secondary_y": True}]])
        fig.add_trace(go.Bar(x=show['label'], y=show['q'], name="Venta Q",
                            marker_color=CLAN_BLUE, opacity=0.7), secondary_y=False)
        fig.add_trace(go.Scatter(x=show['label'], y=show['und'], name="Und",
                                line=dict(color=GREEN, width=2), mode='lines+markers'),
                      secondary_y=True)
        fig.update_layout(height=350, margin=dict(l=20, r=20, t=10, b=20),
                         legend=dict(orientation="h", y=1.1), hovermode="x unified")
        st.plotly_chart(fig, use_container_width=True)

    # Inventory trend — combine inv_historico (historical) + inv_actual (recent)
    st.markdown("### Inventario Historico")

    # inv_historico from VENTA records (weekly aggregate)
    inv_hist = query_df("""
        SELECT semana_wm, SUM(inv_historico) as inv
        FROM retail_link
        WHERE producto=? AND tipo_registro='VENTA' AND inv_historico > 0
        GROUP BY semana_wm ORDER BY semana_wm
    """, [selected_prod])

    # inv_actual from INVENTARIO records (for recent period where inv_historico = 0)
    inv_act = query_df("""
        SELECT semana_wm, SUM(inv_actual) as inv
        FROM retail_link
        WHERE producto=? AND tipo_registro='INVENTARIO'
        GROUP BY semana_wm ORDER BY semana_wm
    """, [selected_prod])

    if not inv_hist.empty or not inv_act.empty:
        fig_inv = go.Figure()

        if not inv_hist.empty:
            inv_hist['fecha'] = inv_hist['semana_wm'].apply(wm_week_to_date_short)
            fig_inv.add_trace(go.Scatter(
                x=inv_hist['fecha'], y=inv_hist['inv'],
                fill='tozeroy', fillcolor='rgba(45,74,122,0.1)',
                line=dict(color=CLAN_BLUE, width=2), name="Inv Historico"))

        if not inv_act.empty:
            inv_act['fecha'] = inv_act['semana_wm'].apply(wm_week_to_date_short)
            fig_inv.add_trace(go.Scatter(
                x=inv_act['fecha'], y=inv_act['inv'],
                fill='tozeroy', fillcolor='rgba(39,174,96,0.15)',
                line=dict(color=GREEN, width=2, dash='dot'), name="Inv Actual"))

        fig_inv.update_layout(height=250, margin=dict(l=20, r=20, t=10, b=20),
                             legend=dict(orientation="h", y=1.1))
        st.plotly_chart(fig_inv, use_container_width=True)
    else:
        st.info("Sin datos de inventario para este producto.")

    # Store breakdown
    st.markdown("### Ventas por Tienda (todo el período)")
    by_store = query_df("""
        SELECT rl.tienda as Tienda, t.formato as Formato,
               SUM(rl.venta_und) as Und, SUM(rl.venta_q) as Q
        FROM retail_link rl
        LEFT JOIN tiendas t ON rl.store_nbr = t.no_tienda
        WHERE rl.producto=? AND rl.tipo_registro='VENTA' AND rl.venta_und > 0
        GROUP BY rl.tienda
        ORDER BY Und DESC
    """, [selected_prod])

    if not by_store.empty:
        st.dataframe(by_store.style.format({'Und': '{:,.0f}', 'Q': 'Q{:,.0f}'}),
                    use_container_width=True, hide_index=True)
        download_df(by_store, f"producto_tiendas_{selected_prod[:20]}")


# ═══════════════════════════════════════════
# PAGE: STORES
# ═══════════════════════════════════════════

def page_tiendas():
    st.markdown("## Análisis por Tienda")

    stores = query_df("""
        SELECT DISTINCT rl.store_nbr, rl.tienda, t.formato, t.region
        FROM retail_link rl
        LEFT JOIN tiendas t ON rl.store_nbr = t.no_tienda
        WHERE rl.tipo_registro='VENTA' AND rl.venta_und > 0
        ORDER BY rl.tienda
    """)

    if stores.empty:
        st.info("Sin datos de tiendas.")
        return

    # Format filter
    formatos = ['Todos'] + sorted(stores['formato'].dropna().unique().tolist())
    fmt_sel = st.selectbox("Formato", formatos)
    if fmt_sel != 'Todos':
        stores = stores[stores['formato'] == fmt_sel]

    selected = st.selectbox("Tienda", stores['tienda'].tolist())
    store_nbr = stores[stores['tienda'] == selected]['store_nbr'].iloc[0]

    st.divider()

    # Store trend
    trend = query_df("""
        SELECT semana_wm,
               SUM(venta_und) as und,
               SUM(venta_q) as q,
               COUNT(DISTINCT item_nbr) as skus
        FROM retail_link
        WHERE store_nbr=? AND tipo_registro='VENTA'
        GROUP BY semana_wm ORDER BY semana_wm
    """, [int(store_nbr)])

    if trend.empty:
        st.info("Sin datos para esta tienda.")
        return

    trend['fecha'] = trend['semana_wm'].apply(wm_week_to_date_short)

    total_und = trend['und'].sum()
    total_q = trend['q'].sum()

    c1, c2, c3 = st.columns(3)
    c1.metric("Venta Total Und", fmt_int(total_und))
    c2.metric("Venta Total Q", fmt_q(total_q))
    c3.metric("SKUs Vendidos", fmt_int(trend['skus'].max()))

    st.markdown(f"### Ventas — {selected}")

    store_gran_col, store_n_col = st.columns([1, 2])
    with store_gran_col:
        store_gran = st.radio("Vista", ["Diario", "Semanal", "Mensual"],
                               index=1, horizontal=True, key="store_gran")

    if store_gran == "Diario":
        store_trend = query_df("""
            SELECT diario as periodo, SUM(venta_und) as und, SUM(venta_q) as q,
                   COUNT(DISTINCT item_nbr) as skus
            FROM retail_link
            WHERE store_nbr=? AND tipo_registro='VENTA' AND diario IS NOT NULL
            GROUP BY diario ORDER BY diario
        """, [int(store_nbr)])
    elif store_gran == "Mensual":
        store_trend = query_df("""
            SELECT anio || '-' || printf('%02d', mes) as periodo,
                   SUM(venta_und) as und, SUM(venta_q) as q,
                   COUNT(DISTINCT item_nbr) as skus
            FROM retail_link
            WHERE store_nbr=? AND tipo_registro='VENTA' AND anio IS NOT NULL
            GROUP BY anio, mes ORDER BY anio, mes
        """, [int(store_nbr)])
    else:
        store_trend = trend.rename(columns={'semana_wm': 'periodo'})[['periodo', 'und', 'q', 'skus']].copy()

    if not store_trend.empty:
        store_trend['label'] = store_trend['periodo'].apply(lambda x: format_period_label(x, store_gran))

        with store_n_col:
            opts = [14, 30, 60, 90, len(store_trend)] if store_gran == "Diario" \
                else [8, 12, 24, 52, len(store_trend)] if store_gran == "Semanal" \
                else [3, 6, 12, len(store_trend)]
            n = st.select_slider("Puntos", opts,
                                  value=min(opts[-2] if len(opts) > 1 else opts[0], len(store_trend)),
                                  key="store_n",
                                  format_func=lambda x: "Todo" if x >= len(store_trend) else str(x))
        show = store_trend.tail(n)

        fig = make_subplots(specs=[[{"secondary_y": True}]])
        fig.add_trace(go.Bar(x=show['label'], y=show['q'], name="Venta Q",
                            marker_color=CLAN_BLUE, opacity=0.7), secondary_y=False)
        fig.add_trace(go.Scatter(x=show['label'], y=show['und'], name="Und",
                                line=dict(color=GREEN, width=2), mode='lines+markers'),
                      secondary_y=True)
        fig.update_layout(height=350, margin=dict(l=20, r=20, t=10, b=20),
                         legend=dict(orientation="h", y=1.1), hovermode="x unified")
        st.plotly_chart(fig, use_container_width=True)

    # Product breakdown
    st.markdown("### Ventas por Producto")
    by_prod = query_df("""
        SELECT rl.producto as Producto, p.marca as Marca,
               SUM(rl.venta_und) as Und, SUM(rl.venta_q) as Q
        FROM retail_link rl
        LEFT JOIN productos p ON rl.producto = p.producto
        WHERE rl.store_nbr=? AND rl.tipo_registro='VENTA' AND rl.venta_und > 0
        GROUP BY rl.producto
        ORDER BY Und DESC
    """, [int(store_nbr)])

    if not by_prod.empty:
        fig_p = px.pie(by_prod.head(10), values='Und', names='Producto',
                      color_discrete_sequence=px.colors.qualitative.Set2, height=350)
        fig_p.update_layout(margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig_p, use_container_width=True)

        st.dataframe(by_prod.style.format({'Und': '{:,.0f}', 'Q': 'Q{:,.0f}'}),
                    use_container_width=True, hide_index=True)
        download_df(by_prod, f"tienda_productos_{selected[:20]}")


# ═══════════════════════════════════════════
# PAGE: PROFITABILITY
# ═══════════════════════════════════════════

def page_rentabilidad():
    st.markdown("## Rentabilidad")
    st.caption(MARGIN_SELLIN)

    si = query_one("""
        SELECT SUM(ingreso_bruto), SUM(costo_total), SUM(centralizacion),
               SUM(ingreso_neto), SUM(margen_bruto), SUM(cantidad_facturada)
        FROM sell_in
    """)

    if not si or not si[0]:
        st.info("Sin datos de Sell-In. Importar facturas de Odoo.")
        return

    rent_view = st.radio("Vista", ["Total Histórico", "Mensual"], horizontal=True, key="rent_view")

    dev_q = query_val("SELECT SUM(precio_reclamo_q + precio_devolucion_q) FROM devoluciones") or 0

    # ── Monthly P&L view ──
    if rent_view == "Mensual":
        monthly = query_df("""
            SELECT SUBSTR(fecha, 1, 7) as mes,
                   SUM(ingreso_bruto) as ingreso_bruto,
                   SUM(costo_total) as costo_total,
                   SUM(centralizacion) as centralizacion,
                   SUM(ingreso_neto) as ingreso_neto,
                   SUM(margen_bruto) as margen_bruto,
                   SUM(cantidad_facturada) as und
            FROM sell_in
            WHERE fecha IS NOT NULL
            GROUP BY SUBSTR(fecha, 1, 7)
            ORDER BY mes
        """)
        if monthly.empty:
            st.info("Sin datos mensuales.")
            return

        monthly['Margen %'] = monthly['margen_bruto'] / monthly['ingreso_bruto'].replace(0, 1)
        monthly['Central. %'] = monthly['centralizacion'] / monthly['ingreso_bruto'].replace(0, 1)

        # Line chart
        st.markdown("### Tendencia Mensual")
        fig_m = make_subplots(specs=[[{"secondary_y": True}]])
        fig_m.add_trace(go.Bar(x=monthly['mes'], y=monthly['ingreso_neto'],
                               name="Ingreso Neto", marker_color=CLAN_BLUE, opacity=0.7),
                        secondary_y=False)
        fig_m.add_trace(go.Scatter(x=monthly['mes'], y=monthly['margen_bruto'],
                                   name="Margen Bruto", line=dict(color=GREEN, width=2.5),
                                   mode='lines+markers'),
                        secondary_y=False)
        fig_m.add_trace(go.Scatter(x=monthly['mes'], y=monthly['Margen %'],
                                   name="Margen %", line=dict(color=PURPLE, width=2, dash='dot'),
                                   mode='lines+markers'),
                        secondary_y=True)
        fig_m.update_layout(height=400, margin=dict(l=20, r=20, t=30, b=20),
                           legend=dict(orientation="h", y=1.1), hovermode="x unified")
        fig_m.update_yaxes(title_text="Monto Q", tickformat=",", secondary_y=False)
        fig_m.update_yaxes(title_text="Margen %", tickformat=".0%", secondary_y=True)
        st.plotly_chart(fig_m, use_container_width=True)

        # Monthly table
        monthly_display = monthly.rename(columns={
            'mes': 'Mes', 'ingreso_bruto': 'Ingreso Bruto', 'costo_total': 'Costo',
            'centralizacion': 'Central.', 'ingreso_neto': 'Ingreso Neto',
            'margen_bruto': 'Margen', 'und': 'Und'
        })
        st.dataframe(monthly_display[['Mes', 'Ingreso Bruto', 'Costo', 'Central.', 'Ingreso Neto', 'Margen', 'Margen %', 'Und']].style.format({
            'Ingreso Bruto': 'Q{:,.0f}', 'Costo': 'Q{:,.0f}', 'Central.': 'Q{:,.0f}',
            'Ingreso Neto': 'Q{:,.0f}', 'Margen': 'Q{:,.0f}', 'Margen %': '{:.1%}', 'Und': '{:,.0f}'
        }), use_container_width=True, hide_index=True)
        download_df(monthly_display, "rentabilidad_mensual")
        st.divider()

    # P&L Waterfall (always shown for Total, after monthly for Mensual)
    st.markdown("### P&L Simplificado — Clan Cervecero x Walmart")

    ingreso_bruto = si[0] or 0
    costo = si[1] or 0
    centraliz = si[2] or 0
    ingreso_neto = si[3] or 0
    margen = si[4] or 0
    margen_neto = margen - dev_q

    actual_central_pct = centraliz / ingreso_bruto if ingreso_bruto else 0
    waterfall_data = {
        'Concepto': ['Ingreso Bruto', f'Centraliz. ({actual_central_pct:.1%})', 'Ingreso Neto',
                     'Costo Producto', 'Margen Bruto', 'Devoluciones', 'Margen Neto'],
        'Monto': [ingreso_bruto, -centraliz, ingreso_neto, -costo, margen, -dev_q, margen_neto],
        'Tipo': ['Ingreso', 'Descuento', 'Subtotal', 'Costo', 'Subtotal', 'Costo', 'Total']
    }

    fig_wf = go.Figure(go.Waterfall(
        x=waterfall_data['Concepto'],
        y=waterfall_data['Monto'],
        measure=['absolute', 'relative', 'total', 'relative', 'total', 'relative', 'total'],
        connector={"line": {"color": CLAN_BLUE}},
        increasing={"marker": {"color": GREEN}},
        decreasing={"marker": {"color": RED}},
        totals={"marker": {"color": CLAN_BLUE}},
        textposition="outside",
        text=[fmt_q(v) for v in waterfall_data['Monto']]
    ))
    fig_wf.update_layout(height=400, margin=dict(l=20, r=20, t=30, b=20),
                         showlegend=False)
    st.plotly_chart(fig_wf, use_container_width=True)

    # KPIs
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Margen Bruto %", f"{margen/ingreso_bruto:.1%}")
    c2.metric("Margen Neto %", f"{margen_neto/ingreso_bruto:.1%}")
    c3.metric("Dev/Ingreso %", f"{dev_q/ingreso_bruto:.1%}")
    c4.metric("Centraliz. %", f"{centraliz/ingreso_bruto:.1%}")

    st.divider()

    # By Brand
    st.markdown("### Rentabilidad por Marca")
    brands = query_df("""
        SELECT marca as Marca,
               SUM(cantidad_facturada) as Und,
               SUM(ingreso_bruto) as "Ingreso Bruto",
               SUM(costo_total) as Costo,
               SUM(margen_bruto) as Margen
        FROM sell_in
        GROUP BY marca
        ORDER BY Margen DESC
    """)

    if not brands.empty:
        brands['Margen %'] = brands['Margen'] / brands['Ingreso Bruto'].replace(0, 1)

        fig_b = px.bar(brands, x='Marca', y=['Ingreso Bruto', 'Costo', 'Margen'],
                      barmode='group', color_discrete_sequence=[CLAN_BLUE, RED, GREEN],
                      height=350)
        fig_b.update_layout(margin=dict(l=20, r=20, t=10, b=20),
                           legend=dict(orientation="h", y=1.1))
        st.plotly_chart(fig_b, use_container_width=True)

        st.dataframe(brands.style.format({
            'Und': '{:,.0f}', 'Ingreso Bruto': 'Q{:,.0f}',
            'Costo': 'Q{:,.0f}', 'Margen': 'Q{:,.0f}', 'Margen %': '{:.1%}'
        }), use_container_width=True, hide_index=True)
        download_df(brands, "rentabilidad_marcas")

    # By Product
    st.divider()
    st.markdown("### Rentabilidad por Producto")
    prods = query_df("""
        SELECT nombre_producto as Producto, marca as Marca,
               SUM(cantidad_facturada) as Und,
               SUM(ingreso_bruto) as "Ingreso Bruto",
               SUM(costo_total) as Costo,
               SUM(centralizacion) as "Central.",
               SUM(margen_bruto) as Margen
        FROM sell_in
        GROUP BY codigo_producto
        ORDER BY Margen DESC
    """)

    if not prods.empty:
        prods['Margen %'] = prods['Margen'] / prods['Ingreso Bruto'].replace(0, 1)

        st.dataframe(prods.style.format({
            'Und': '{:,.0f}', 'Ingreso Bruto': 'Q{:,.0f}', 'Costo': 'Q{:,.0f}',
            'Central.': 'Q{:,.0f}', 'Margen': 'Q{:,.0f}', 'Margen %': '{:.1%}'
        }),
        use_container_width=True, hide_index=True)
        download_df(prods, "rentabilidad_productos")


# ═══════════════════════════════════════════
# PAGE: SUGGESTED ORDERS
# ═══════════════════════════════════════════

def page_sugerido():
    st.markdown("## Pedido Sugerido")

    inv_week = query_val("SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO'")
    if not inv_week:
        st.warning("Sin datos de inventario.")
        return

    col_t, col_l = st.columns(2)
    with col_t:
        target_days = st.slider("Dias de cobertura objetivo", 7, 30, CONFIG["default_target_days"])
    with col_l:
        lead_time = st.slider("Lead time (días de espera del pedido)", 0, 21, CONFIG["default_lead_time_days"])

    # Get last N weeks of sales
    sale_weeks = query_df(f"""
        SELECT DISTINCT semana_wm FROM retail_link
        WHERE tipo_registro='VENTA' ORDER BY semana_wm DESC LIMIT {CONFIG["velocity_weeks"]}
    """)['semana_wm'].tolist()

    ph = ','.join(['?'] * len(sale_weeks))
    days_period = len(sale_weeks) * 7

    st.caption(f"Inventario al {wm_week_to_date(inv_week)} | "
               f"Ventas: Últimas {len(sale_weeks)} semanas | Objetivo: {target_days} + {lead_time} días lead time")

    # Calculate per product (aggregated across stores)
    suggestions = query_df(f"""
        WITH ventas AS (
            SELECT producto,
                   SUM(venta_und) as venta_total
            FROM retail_link
            WHERE tipo_registro='VENTA' AND semana_wm IN ({ph})
            GROUP BY producto
        ),
        inventario AS (
            SELECT producto, SUM(inv_actual) as inv_actual
            FROM retail_link
            WHERE tipo_registro='INVENTARIO' AND semana_wm=?
            GROUP BY producto
        )
        SELECT
            COALESCE(v.producto, i.producto) as Producto,
            p.marca as Marca,
            p.und_x_caja as "Und/Caja",
            COALESCE(v.venta_total, 0) as "Venta {days_period}d",
            COALESCE(i.inv_actual, 0) as Inventario,
            CASE
                WHEN COALESCE(v.venta_total, 0) = 0 THEN 999
                ELSE COALESCE(i.inv_actual, 0) * {days_period}.0 / COALESCE(v.venta_total, 1)
            END as "Cobertura (dias)"
        FROM ventas v
        FULL OUTER JOIN inventario i ON v.producto = i.producto
        LEFT JOIN productos p ON COALESCE(v.producto, i.producto) = p.producto
        WHERE p.estado = 'Activo'
        ORDER BY "Cobertura (dias)" ASC
    """, [*sale_weeks, inv_week])

    if suggestions.empty:
        st.info("Sin datos suficientes.")
        return

    # Calculate suggested boxes
    def calc_suggestion(row):
        venta = row[f'Venta {days_period}d']
        inv = row['Inventario']
        und_caja = row['Und/Caja'] or 1
        daily_rate = venta / days_period if venta > 0 else 0
        needed = max(0, ((target_days + lead_time) * daily_rate) - inv)
        cajas = max(1, int((needed + und_caja - 1) // und_caja)) if needed > 0 else 0
        return pd.Series({'Cajas Sug.': cajas, 'Und Sug.': cajas * und_caja,
                         'Rate/dia': daily_rate})

    extra = suggestions.apply(calc_suggestion, axis=1)
    suggestions = pd.concat([suggestions, extra], axis=1)

    # Priority
    def get_priority(row):
        cob = row['Cobertura (dias)']
        inv = row['Inventario']
        venta = row[f'Venta {days_period}d']
        if (cob <= 0 or inv == 0) and venta > 0:
            return 'P1'
        elif cob < target_days * CONFIG["priority_p2_multiplier"]:
            return 'P2'
        elif cob < target_days * CONFIG["priority_p3_multiplier"]:
            return 'P3'
        return 'OK'

    suggestions['Prioridad'] = suggestions.apply(get_priority, axis=1)

    # Filter to actionable
    actionable = suggestions[suggestions['Prioridad'] != 'OK'].copy()

    # Summary
    p1 = actionable[actionable['Prioridad'] == 'P1']
    p2 = actionable[actionable['Prioridad'] == 'P2']
    p3 = actionable[actionable['Prioridad'] == 'P3']

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("P1 Urgente", f"{len(p1)} items", delta=f"{p1['Cajas Sug.'].sum():.0f} cajas")
    c2.metric("P2 Media", f"{len(p2)} items", delta=f"{p2['Cajas Sug.'].sum():.0f} cajas")
    c3.metric("P3 Baja", f"{len(p3)} items", delta=f"{p3['Cajas Sug.'].sum():.0f} cajas")
    c4.metric("Total Cajas", fmt_int(actionable['Cajas Sug.'].sum()))

    st.divider()

    # Detailed table
    def fmt_coverage(val):
        return "Sin ventas" if val >= 999 else f"{val:.1f}"

    if not actionable.empty:
        def color_priority(val):
            if val == 'P1':
                return 'background-color: #FFEBEE; color: #C62828; font-weight: bold'
            elif val == 'P2':
                return 'background-color: #FFF8E1; color: #F57F17; font-weight: bold'
            elif val == 'P3':
                return 'background-color: #E8F5E9; color: #2E7D32; font-weight: bold'
            return ''

        display_cols = ['Prioridad', 'Producto', 'Marca', 'Und/Caja',
                       f'Venta {days_period}d', 'Inventario', 'Cobertura (dias)',
                       'Cajas Sug.', 'Und Sug.']
        styled = actionable[display_cols].style.map(
            color_priority, subset=['Prioridad']
        ).format({
            f'Venta {days_period}d': '{:,.0f}', 'Inventario': '{:,.0f}',
            'Cobertura (dias)': fmt_coverage, 'Cajas Sug.': '{:,.0f}', 'Und Sug.': '{:,.0f}'
        })
        st.dataframe(styled, use_container_width=True, hide_index=True)
        download_df(actionable[display_cols], "pedido_sugerido")

    # All products overview
    with st.expander("Ver todos los productos (incluye OK)"):
        st.dataframe(suggestions[['Prioridad', 'Producto', 'Marca',
                                  f'Venta {days_period}d', 'Inventario',
                                  'Cobertura (dias)', 'Cajas Sug.']].style.format({
            f'Venta {days_period}d': '{:,.0f}', 'Inventario': '{:,.0f}',
            'Cobertura (dias)': fmt_coverage, 'Cajas Sug.': '{:,.0f}'
        }), use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════
# PAGE: CLAIMS / DEVOLUCIONES
# ═══════════════════════════════════════════

def page_claims():
    st.markdown("## Devoluciones & Claims")

    # ── 3 fuentes de claims ──
    cc = query_one("""
        SELECT COUNT(*), SUM(COALESCE(importe_q, 0))
        FROM claims_contact_center WHERE clase_documento = 'Z2'
    """) or (0, 0)
    correo = query_one("""
        SELECT COUNT(*), SUM(COALESCE(valor_total, 0)),
               SUM(COALESCE(unidades, 0))
        FROM claims_correo
    """) or (0, 0, 0)
    rl_claims = query_one("""
        SELECT COUNT(*), SUM(COALESCE(cant_reclamo, 0)),
               SUM(COALESCE(precio_reclamo_q, 0)),
               SUM(COALESCE(costo_reclamo_q, 0))
        FROM claims_rl
    """) or (0, 0, 0, 0)

    st.markdown("### Fuentes de Claims")
    c1, c2, c3 = st.columns(3)
    c1.metric("Contact Center (SAP)", f"{cc[0]} docs Z2",
              delta=fmt_q(cc[1]))
    c2.metric("Correo (Portal)", f"{correo[0]} claims",
              delta=f"{fmt_q(correo[1])}  ({fmt_int(correo[2])} und)")
    c3.metric("Retail Link", f"{rl_claims[0]} registros",
              delta=f"{fmt_q(rl_claims[2])}  ({fmt_int(rl_claims[1])} und)")

    st.divider()

    # ── Conciliación resumen ──
    st.markdown("### Conciliación")
    conc_data = {
        'Fuente': ['Contact Center (SAP)', 'Correo (Portal)', 'Retail Link'],
        'Registros': [cc[0], correo[0], rl_claims[0]],
        'Monto Q': [cc[1], correo[1], rl_claims[2]],
        'Unidades': ['-', correo[2], rl_claims[1]],
        'Qué mide': [
            'Deducciones contables en cuenta Clan',
            'Claims formales con detalle de tienda',
            'Devoluciones en sistema sell-out por SKU/tienda'
        ],
    }
    conc_df = pd.DataFrame(conc_data)
    st.dataframe(conc_df.style.format({
        'Registros': '{:,.0f}',
        'Monto Q': 'Q{:,.0f}',
    }), use_container_width=True, hide_index=True)
    download_df(conc_df, "claims_conciliacion")

    st.info("""
**¿Cuál fuente usar?**
- **Para contabilidad** → Contact Center (SAP): deducciones reales en cuenta
- **Para negociar con Walmart** → Correo (Portal): claims formales con detalle
- **Para análisis operativo** → Retail Link: nivel SKU/tienda, más granular

Los montos NO deben sumar — cada fuente mide una etapa diferente del proceso.
""")

    st.divider()

    # ── Claims RL por producto (la fuente más granular) ──
    st.markdown("### Claims por Producto (Retail Link)")
    by_prod = query_df("""
        SELECT descripcion as Producto,
               SUM(cant_reclamo) as "Unidades",
               SUM(precio_reclamo_q) as "Precio Q",
               SUM(costo_reclamo_q) as "Costo Q",
               COUNT(DISTINCT store_nbr) as "Tiendas"
        FROM claims_rl
        GROUP BY item_nbr, descripcion
        ORDER BY SUM(precio_reclamo_q) DESC
    """)

    if not by_prod.empty:
        fig = px.bar(by_prod, x='Producto', y='Precio Q',
                    color_discrete_sequence=[RED], height=300)
        fig.update_layout(margin=dict(l=20, r=20, t=10, b=20))
        st.plotly_chart(fig, use_container_width=True)

        st.dataframe(by_prod.style.format({
            'Unidades': '{:,.0f}', 'Precio Q': 'Q{:,.0f}',
            'Costo Q': 'Q{:,.0f}', 'Tiendas': '{:,.0f}'
        }), use_container_width=True, hide_index=True)
        download_df(by_prod, "claims_por_producto")

        # Claims per 1000 units sold
        total_claims_und = by_prod['Unidades'].sum()
        total_sellout_und = query_val("SELECT SUM(venta_und) FROM retail_link WHERE tipo_registro='VENTA'") or 1
        claims_rate = total_claims_und / total_sellout_und * 1000
        st.metric("Claims por 1,000 und vendidas", f"{claims_rate:.1f}")

    # ── Claims por tienda (top offenders) ──
    st.divider()
    st.markdown("### Top Tiendas con Claims (RL)")
    by_store = query_df("""
        SELECT tienda as Tienda,
               SUM(cant_reclamo) as "Unidades",
               SUM(precio_reclamo_q) as "Precio Q",
               COUNT(DISTINCT item_nbr) as "SKUs"
        FROM claims_rl
        GROUP BY store_nbr, tienda
        ORDER BY SUM(precio_reclamo_q) DESC
        LIMIT 15
    """)

    if not by_store.empty:
        fig = px.bar(by_store, x='Tienda', y='Precio Q',
                    color_discrete_sequence=[ORANGE], height=300)
        fig.update_layout(margin=dict(l=20, r=20, t=10, b=20))
        st.plotly_chart(fig, use_container_width=True)

        st.dataframe(by_store.style.format({
            'Unidades': '{:,.0f}', 'Precio Q': 'Q{:,.0f}', 'SKUs': '{:,.0f}'
        }), use_container_width=True, hide_index=True)
        download_df(by_store, "claims_por_tienda_rl")

    # ── Correo por tienda ──
    st.divider()
    st.markdown("### Claims por Tienda (Correo)")
    correo_store = query_df("""
        SELECT nombre_tienda as Tienda,
               SUM(unidades) as "Unidades",
               SUM(valor_neto) as "Valor Neto Q",
               SUM(valor_total) as "Valor Total Q",
               COUNT(*) as "Docs"
        FROM claims_correo
        GROUP BY codigo_tienda, nombre_tienda
        ORDER BY SUM(valor_total) DESC
        LIMIT 15
    """)

    if not correo_store.empty:
        st.dataframe(correo_store.style.format({
            'Unidades': '{:,.0f}', 'Valor Neto Q': 'Q{:,.0f}',
            'Valor Total Q': 'Q{:,.0f}', 'Docs': '{:,.0f}'
        }), use_container_width=True, hide_index=True)
        download_df(correo_store, "claims_por_tienda_correo")

    # ── Tendencia semanal (RL claims) ──
    st.divider()
    st.markdown("### Tendencia Semanal de Claims")
    dev_trend = query_df("""
        SELECT semana_wm,
               SUM(cant_reclamo) as und,
               SUM(precio_reclamo_q) as q
        FROM claims_rl
        WHERE semana_wm IS NOT NULL
        GROUP BY semana_wm
        ORDER BY semana_wm
    """)

    if not dev_trend.empty:
        dev_trend['fecha'] = dev_trend['semana_wm'].apply(wm_week_to_date_short)
        fig = go.Figure()
        fig.add_trace(go.Bar(x=dev_trend['fecha'], y=dev_trend['q'],
                            name="Claims Q", marker_color=RED, opacity=0.7))
        fig.add_trace(go.Scatter(x=dev_trend['fecha'], y=dev_trend['und'],
                                name="Unidades", line=dict(color=ORANGE, width=2),
                                mode='lines+markers', yaxis='y2'))
        fig.update_layout(
            height=300, margin=dict(l=20, r=20, t=10, b=20),
            yaxis2=dict(overlaying='y', side='right', tickformat=','),
            legend=dict(orientation="h", y=1.1)
        )
        st.plotly_chart(fig, use_container_width=True)


# ═══════════════════════════════════════════
# PAGE: UPLOAD
# ═══════════════════════════════════════════

def _is_raw_walmart_rl(wb):
    """Check if workbook is a raw Walmart Retail Link export (has metadata header rows)."""
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if row:
            text = ' '.join(str(v or '') for v in row[:5]).lower()
            if 'formato nuevo semanal' in text or 'report options' in text:
                return True
    return False


def _find_data_start(ws):
    """Find the header row and first data row in a raw Walmart export."""
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=45, values_only=True), 1):
        if not row:
            continue
        val0 = str(row[0] or '').strip() if row[0] else ''
        # Header row has "Semana Walmart" or "Semana" in col 0
        if 'semana' in val0.lower():
            header_row = i
            continue
        # Data row starts with 6-digit semana number (may be float like 202548.0)
        try:
            num = str(int(float(val0))) if val0 else ''
        except (ValueError, OverflowError):
            num = val0
        if len(num) >= 6 and num[:6].isdigit():
            return header_row or (i - 1), i
    return None, None


def _detect_sheet_type(ws_title, headers_str, first_row, col_count):
    """Auto-detect data type from sheet name, headers, and structure."""
    t = ws_title.lower()
    h = headers_str.lower()

    # Claims Contact Center (SAP) — "Nº documento", "Referencia", "Clase de documento"
    if 'contact' in t or ('documento' in h and 'referencia' in h and 'clase' in h):
        return 'claims_cc'

    # Claims Correo — "No. Vale", "Uuid", "Valor Neto"
    if ('correo' in t and 'vale' in h) or ('uuid' in h and 'valor neto' in h):
        return 'claims_correo'

    # Claims RL — "Reclamo" in header + "Semana Walmart" or similar
    if ('reclamo' in h or 'claim' in h) and ('semana' in h or 'diario' in h):
        return 'claims_rl'

    # Retail Link — "Semana Walmart" + many columns (30+)
    if ('semana' in h and col_count > 25 and
            ('venta' in h or 'item' in h or 'tienda' in h)):
        return 'retail_link'

    # Sell-In flat — starts with CC- codes or has "Código Producto"
    if first_row and str(first_row[0] or '').startswith('CC-'):
        return 'sell_in'
    if 'codigo producto' in h or 'código producto' in h:
        return 'sell_in'

    # Devoluciones (legacy) — "Semana" + "reclamo"/"devolución"
    if 'reclamo' in h or 'devolución' in h or 'devolucion' in h:
        return 'devoluciones'

    # Odoo Pivot Table — very wide (100+ cols), few rows (<15)
    # detected at workbook level, not here

    return None


def _import_claims_cc(conn, ws):
    """Import Claims Contact Center sheet."""
    count = 0
    now = datetime.now().isoformat()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        def pdate(v):
            if v is None: return None
            if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
            return str(v).strip() or None
        def sfloat(v):
            if v is None: return 0.0
            try: return float(v)
            except: return 0.0

        try:
            conn.execute('''
                INSERT OR IGNORE INTO claims_contact_center
                (no_documento, referencia, no_vale, clase_documento, importe_q,
                 fecha_compensacion, fecha_entrada, fecha_documento, vencimiento,
                 cuenta, texto, imported_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                str(row[0]).strip(),
                str(row[7]).strip() if len(row) > 7 and row[7] else None,
                str(row[8]).strip() if len(row) > 8 and row[8] else None,
                str(row[11]).strip() if len(row) > 11 and row[11] else None,
                sfloat(row[15]) if len(row) > 15 else 0.0,
                pdate(row[16]) if len(row) > 16 else None,
                pdate(row[17]) if len(row) > 17 else None,
                pdate(row[18]) if len(row) > 18 else None,
                pdate(row[19]) if len(row) > 19 else None,
                str(row[20]).strip() if len(row) > 20 and row[20] else None,
                str(row[22]).strip() if len(row) > 22 and row[22] else None,
                now,
            ))
            count += 1
        except Exception:
            pass
    conn.commit()
    return count


def _import_claims_correo(conn, ws):
    """Import Claims Correo sheet."""
    count = 0
    now = datetime.now().isoformat()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        def pdate(v):
            if v is None: return None
            if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
            return str(v).strip() or None
        def sfloat(v):
            if v is None: return 0.0
            try: return float(v)
            except: return 0.0
        def sint(v):
            if v is None: return 0
            try: return int(float(v))
            except: return 0

        try:
            conn.execute('''
                INSERT OR IGNORE INTO claims_correo
                (doc, fecha_inicial, fecha_final, no_vale, estado, uuid,
                 codigo_tienda, nombre_tienda, unidades, valor_neto, valor_iva,
                 valor_total, imported_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                str(row[0]).strip(),
                pdate(row[1]),
                pdate(row[2]),
                str(row[3]).strip() if row[3] else None,
                str(row[5]).strip() if len(row) > 5 and row[5] else None,
                str(row[7]).strip() if len(row) > 7 and row[7] else None,
                str(row[11]).strip() if len(row) > 11 and row[11] else None,
                str(row[12]).strip() if len(row) > 12 and row[12] else None,
                sint(row[13]) if len(row) > 13 else 0,
                sfloat(row[14]) if len(row) > 14 else 0.0,
                sfloat(row[15]) if len(row) > 15 else 0.0,
                sfloat(row[16]) if len(row) > 16 else 0.0,
                now,
            ))
            count += 1
        except Exception:
            pass
    conn.commit()
    return count


def _import_claims_rl(conn, ws):
    """Import Claims RL sheet."""
    count = 0
    now = datetime.now().isoformat()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        def pdate(v):
            if v is None: return None
            if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
            return str(v).strip() or None
        def sfloat(v):
            if v is None: return 0.0
            try: return float(v)
            except: return 0.0
        def sint(v):
            if v is None: return 0
            try: return int(float(v))
            except: return 0

        try:
            conn.execute('''
                INSERT OR IGNORE INTO claims_rl
                (diario, semana_wm, pais, item_nbr, descripcion, store_nbr, tienda,
                 cant_reclamo, precio_reclamo_q, precio_reclamo_usd,
                 costo_reclamo_q, costo_reclamo_usd, fecha_tc, tipo_cambio, imported_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                pdate(row[0]),
                str(row[1]).strip() if row[1] else None,
                str(row[2]).strip() if len(row) > 2 and row[2] else None,
                str(row[3]).strip() if len(row) > 3 and row[3] else None,
                str(row[4]).strip() if len(row) > 4 and row[4] else None,
                sint(row[5]) if len(row) > 5 else 0,
                str(row[6]).strip() if len(row) > 6 and row[6] else None,
                sint(row[7]) if len(row) > 7 else 0,
                sfloat(row[8]) if len(row) > 8 else 0.0,
                sfloat(row[9]) if len(row) > 9 else 0.0,
                sfloat(row[10]) if len(row) > 10 else 0.0,
                sfloat(row[11]) if len(row) > 11 else 0.0,
                str(row[12]).strip() if len(row) > 12 and row[12] else None,
                sfloat(row[13]) if len(row) > 13 else 0.0,
                now,
            ))
            count += 1
        except Exception:
            pass
    conn.commit()
    return count


def _import_odoo_pivot(conn, wb, ws):
    """Import Sell-In from Odoo pivot table format."""
    from import_walmart import safe_str, safe_float, safe_int

    rows_raw = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows_raw) < 7:
        return 0

    row_products = rows_raw[0]
    row_dates = rows_raw[2]
    row_orders = rows_raw[3]
    row_values = rows_raw[6]

    # Parse product names and map column blocks
    products = {}
    current_prod = None
    for ci, val in enumerate(row_products):
        if val and str(val).strip().startswith('['):
            current_prod = str(val).strip()
        if current_prod:
            products[ci] = current_prod

    # Spanish months
    es_months = {'ene':'01','feb':'02','mar':'03','abr':'04','may':'05','jun':'06',
                 'jul':'07','ago':'08','sep':'09','oct':'10','nov':'11','dic':'12'}

    def parse_es_date(s):
        if not s: return None
        parts = str(s).strip().lower().split()
        if len(parts) == 3 and parts[1] in es_months:
            return f"{parts[2]}-{es_months[parts[1]]}-{parts[0].zfill(2)}"
        return None

    def extract_code(prod_str):
        if '[' in prod_str and ']' in prod_str:
            return prod_str.split(']')[0].replace('[', '').strip()
        return None

    # Lookup pricing from productos table
    precio_map = {}
    costo_map = {}
    und_map = {}
    for r in conn.execute("SELECT codigo, ingreso_caja_wm, costo_clan_caja, und_x_caja FROM productos").fetchall():
        if r[0]:
            precio_map[r[0]] = r[1] or 0
            costo_map[r[0]] = r[2] or 0
            und_map[r[0]] = r[3] or 1

    records = []
    seen_cols = set()
    for ci in range(len(row_values or [])):
        prod_str = products.get(ci)
        if not prod_str:
            continue
        val = row_values[ci]
        if val is None or ci in seen_cols:
            continue

        # Each block is 4 cols: qty, discount, unit_price, total
        date_str = str(row_dates[ci]).strip() if row_dates[ci] else None
        order_str = str(row_orders[ci]).strip() if row_orders[ci] else None
        fecha = parse_es_date(date_str)
        if not fecha:
            continue

        qty = safe_int(val)
        if qty <= 0:
            continue

        code = extract_code(prod_str)
        name = prod_str.split(']')[1].strip() if ']' in prod_str else prod_str

        # Calculate financials
        und_caja = und_map.get(code, 1) or 1
        cajas = qty / und_caja if und_caja else 0
        precio_caja = precio_map.get(code, 0)
        costo_caja = costo_map.get(code, 0)
        ingreso = cajas * precio_caja
        costo = cajas * costo_caja
        central = ingreso * CONFIG["centralization_rate"]
        neto = ingreso - central
        margen = neto - costo
        margen_pct = (margen / neto * 100) if neto else 0

        records.append((
            code, name, 'WALMART', fecha, qty, None, name,
            und_caja, round(cajas, 2),
            precio_caja, costo_caja,
            round(ingreso, 2), round(costo, 2), round(central, 2),
            round(neto, 2), round(margen, 2), round(margen_pct, 1),
            None, None,
        ))
        seen_cols.add(ci)

    added = 0
    for rec in records:
        try:
            conn.execute("""
                INSERT OR IGNORE INTO sell_in
                (codigo_producto, nombre_producto, cliente, fecha, cantidad_facturada,
                 fecha_esperada, producto_precios, und_x_caja, cajas_facturadas,
                 precio_caja_wm, costo_caja_clan, ingreso_bruto, costo_total,
                 centralizacion, ingreso_neto, margen_bruto, margen_pct, marca, anio_mes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, rec)
            added += 1
        except sqlite3.IntegrityError:
            pass
    conn.commit()
    return added


def _is_odoo_pivot(wb):
    """Check if workbook is an Odoo pivot table (few rows, many columns)."""
    ws = wb.worksheets[0]
    row_count = 0
    max_cols = 0
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row and any(v is not None for v in row):
            row_count += 1
            max_cols = max(max_cols, sum(1 for v in row if v is not None))
    # Odoo pivot: <10 data rows, 50+ columns, first row has [CC-...] product codes
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    has_cc_codes = any(str(v or '').startswith('[CC-') for v in first_row)
    return row_count < 12 and max_cols > 40 and has_cc_codes


def page_upload():
    st.markdown("## Subir Datos")
    st.markdown("""
    Subí **cualquier** archivo — el sistema detecta automáticamente qué tipo de datos es:

    | Tipo | Cómo lo reconoce |
    |------|-----------------|
    | **RL Crudo (Walmart export)** | "Formato Nuevo Semanal" + 50 cols (incluye devoluciones) |
    | **RL Semanal** | Columnas "Semana Walmart", 30+ cols |
    | **Sell-In (Odoo flat)** | Códigos CC-XXX-XX |
    | **Sell-In (Odoo pivot)** | Tabla dinámica con [CC-...] en fila 1 |
    | **Claims Contact Center** | "Nº documento", "Referencia", "Clase de documento" |
    | **Claims Correo** | "No. Vale", "Uuid", "Valor Neto" |
    | **Claims RL** | "Reclamo" + "Semana"/"Diario" |
    | **Devoluciones** | "Reclamo"/"Devolución" sin semana |
    """)

    uploaded = st.file_uploader(
        "Arrastrá o seleccioná archivos Excel",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
    )

    if uploaded and st.button("Procesar archivos", type="primary"):
        conn = sqlite3.connect(DB_PATH)
        sys.path.insert(0, BASE_DIR)
        from import_walmart import (import_retail_link, import_sell_in,
                                     import_devoluciones, safe_str, safe_float)

        total_added = 0
        total_duped = 0
        progress = st.progress(0, text="Procesando...")

        for i, file in enumerate(uploaded):
            progress.progress((i + 1) / len(uploaded), text=f"Procesando {file.name}...")

            tmp_path = os.path.join(tempfile.gettempdir(), file.name)
            with open(tmp_path, 'wb') as f:
                f.write(file.getbuffer())

            try:
                import openpyxl

                # Handle .xls files that are actually xlsx
                load_path = tmp_path
                if tmp_path.lower().endswith('.xls') and not tmp_path.lower().endswith('.xlsx'):
                    xlsx_path = tmp_path + 'x'
                    import shutil
                    shutil.copy2(tmp_path, xlsx_path)
                    load_path = xlsx_path

                wb = openpyxl.load_workbook(load_path, read_only=False, data_only=True)

                # Check if it's a raw Walmart RL export (has metadata header)
                if _is_raw_walmart_rl(wb):
                    ws = wb.worksheets[0]
                    _, data_start = _find_data_start(ws)
                    if data_start:
                        all_rows = list(ws.iter_rows(min_row=data_start, values_only=True))
                        def _semana_val(v):
                            try: return str(int(float(v)))
                            except: return str(v).strip()
                        data_rows = [r for r in all_rows if r and r[0] and
                                     _semana_val(r[0])[:4].isdigit()]
                        a, d, u = import_retail_link(conn, data_rows, f'upload:{file.name}')
                        total_added += a; total_duped += d
                        reclamos = sum(1 for r in data_rows if len(r) > 38 and r[38] and safe_float(r[38]) != 0)
                        upd_msg = f", {u} reclamos actualizados" if u else ""
                        st.success(f"**{file.name}**: RL Crudo (Walmart) — +{a:,}, {d:,} dupes"
                                   f" ({reclamos} con reclamos{upd_msg})")
                    else:
                        st.warning(f"**{file.name}**: Formato Walmart detectado pero no encontré datos")
                    wb.close()
                    continue

                # Check if it's an Odoo pivot table
                if _is_odoo_pivot(wb):
                    a = _import_odoo_pivot(conn, wb, wb.worksheets[0])
                    total_added += a
                    st.success(f"**{file.name}**: Sell-In (Odoo Pivot) — +{a:,} filas")
                    wb.close()
                    continue

                for ws in wb.worksheets:
                    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
                    if not rows:
                        continue

                    first_data = None
                    for r in rows:
                        if r and r[0]:
                            first_data = r
                            break
                    if first_data is None:
                        continue

                    header_str = ' '.join(safe_str(v) for v in first_data[:20] if v)
                    col_count = sum(1 for v in first_data if v is not None)

                    dtype = _detect_sheet_type(ws.title, header_str, first_data, col_count)

                    if dtype == 'retail_link':
                        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
                        data_rows = [r for r in all_rows[1:] if r and r[0] and safe_str(r[0])[:4].isdigit()]
                        a, d, u = import_retail_link(conn, data_rows, f'upload:{file.name}')
                        total_added += a; total_duped += d
                        upd_msg = f", {u} reclamos actualizados" if u else ""
                        st.success(f"**{file.name}** / {ws.title}: RL Semanal — +{a:,}, {d:,} dupes{upd_msg}")

                    elif dtype == 'sell_in':
                        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
                        data_rows = [r for r in all_rows[1:] if r and r[0] and safe_str(r[0]).startswith('CC-')]
                        a, d = import_sell_in(conn, data_rows, f'upload:{file.name}')
                        total_added += a; total_duped += d
                        st.success(f"**{file.name}** / {ws.title}: Sell-In — +{a:,}, {d:,} dupes")

                    elif dtype == 'devoluciones':
                        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
                        data_rows = [r for r in all_rows[1:] if r and r[0] and safe_str(r[0])[:4].isdigit()]
                        a, d = import_devoluciones(conn, data_rows, f'upload:{file.name}')
                        total_added += a; total_duped += d
                        st.success(f"**{file.name}** / {ws.title}: Devoluciones — +{a:,}, {d:,} dupes")

                    elif dtype == 'claims_cc':
                        a = _import_claims_cc(conn, ws)
                        total_added += a
                        st.success(f"**{file.name}** / {ws.title}: Claims Contact Center — +{a:,}")

                    elif dtype == 'claims_correo':
                        a = _import_claims_correo(conn, ws)
                        total_added += a
                        st.success(f"**{file.name}** / {ws.title}: Claims Correo — +{a:,}")

                    elif dtype == 'claims_rl':
                        a = _import_claims_rl(conn, ws)
                        total_added += a
                        st.success(f"**{file.name}** / {ws.title}: Claims RL — +{a:,}")

                    else:
                        st.info(f"**{file.name}** / {ws.title}: Tipo no reconocido (saltado)")

                wb.close()

            except Exception as e:
                st.error(f"Error procesando {file.name}: {e}")
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
                xlsx_tmp = tmp_path + 'x'
                if os.path.exists(xlsx_tmp):
                    os.remove(xlsx_tmp)

        conn.execute("""
            INSERT INTO import_log (source, filename, rows_added, rows_duped)
            VALUES ('webapp', ?, ?, ?)
        """, (datetime.now().strftime('%Y-%m-%d'), total_added, total_duped))
        conn.commit()

        progress.progress(1.0, text="Completado")
        st.balloons()
        st.markdown(f"""
        ### Importacion completada
        - **Filas nuevas:** {total_added:,}
        - **Duplicados ignorados:** {total_duped:,}
        """)
        if total_added == 0 and total_duped == 0:
            st.warning("No se importaron filas. Verificá que el archivo tenga el formato correcto.")
        elif total_added == 0 and total_duped > 0:
            st.info("Todos los datos ya estaban en la base (duplicados).")
        st.cache_resource.clear()

    st.divider()

    # DB Status
    st.markdown("### Estado de la Base de Datos")

    rl_count = query_val("SELECT COUNT(*) FROM retail_link") or 0
    si_count = query_val("SELECT COUNT(*) FROM sell_in") or 0
    dev_count = query_val("SELECT COUNT(*) FROM devoluciones") or 0
    weeks = query_one("SELECT MIN(semana_wm), MAX(semana_wm) FROM retail_link")
    last_import = query_one("SELECT source, filename, rows_added, imported_at FROM import_log ORDER BY id DESC LIMIT 1")

    c1, c2, c3 = st.columns(3)
    c1.metric("Retail Link", fmt_int(rl_count))
    c2.metric("Sell-In", fmt_int(si_count))
    c3.metric("Devoluciones", fmt_int(dev_count))

    if weeks:
        st.caption(f"Rango: {wm_week_to_date(weeks[0])} - {wm_week_to_date(weeks[1])}")
    if last_import:
        st.caption(f"Ultima importacion: {last_import[3]} ({last_import[0]}: +{last_import[2]:,} filas)")

    # Import history
    with st.expander("Historial de importaciones"):
        history = query_df("SELECT * FROM import_log ORDER BY id DESC LIMIT 20")
        if not history.empty:
            st.dataframe(history, use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════
# PAGE: DB STATUS
# ═══════════════════════════════════════════

def page_status():
    st.markdown("## Estado del Sistema")

    db_size = os.path.getsize(DB_PATH) / (1024 * 1024)

    # Tables
    tables = query_df("""
        SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'
        ORDER BY name
    """)

    st.markdown(f"**Base de datos:** `walmart.db` ({db_size:.1f} MB)")

    for _, row in tables.iterrows():
        table = row['name']
        count = query_val(f"SELECT COUNT(*) FROM [{table}]")
        st.markdown(f"- **{table}**: {count:,} filas")

    st.divider()

    # Export
    st.markdown("### Exportar Reporte Excel")
    if st.button("Generar REPORTE-WM.xlsx", type="primary"):
        with st.spinner("Generando reporte..."):
            sys.path.insert(0, BASE_DIR)
            import generate_report
            generate_report.main()
        report_path = os.path.join(REPORT_DIR, "REPORTE-WM_latest.xlsx")
        if os.path.exists(report_path):
            with open(report_path, 'rb') as f:
                st.download_button(
                    "Descargar Reporte",
                    data=f.read(),
                    file_name=f"REPORTE-WM_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            st.success(f"Reporte generado: {report_path}")


# ═══════════════════════════════════════════
# MAIN APP
# ═══════════════════════════════════════════

def main():
    st.set_page_config(
        page_title="ClanTrack — Clan Cervecero",
        page_icon=os.path.join(BASE_DIR, "ClanTrack.png"),
        layout="wide",
        initial_sidebar_state="expanded"
    )

    # Custom CSS
    st.markdown("""
    <style>
        .stMetric { background-color: #f8f9fa; padding: 12px; border-radius: 8px; border-left: 4px solid #2D4A7A; }
        [data-testid="stSidebar"] { background-color: #1B2A4A; }
        [data-testid="stSidebar"] * { color: white !important; }
        [data-testid="stSidebar"] .stRadio label { font-size: 15px; }
        h2 { color: #1B2A4A; border-bottom: 2px solid #2D4A7A; padding-bottom: 8px; }
        h3 { color: #2D4A7A; }
        .stSelectbox label, .stSlider label { color: #666; }
    </style>
    """, unsafe_allow_html=True)

    # Sidebar
    with st.sidebar:
        st.markdown("### CLANTRACK")
        st.markdown("##### Retail Intelligence")
        st.divider()

        nav_pages = ["Alertas", "Dashboard", "Semanal", "Tendencias", "Productos",
             "Tiendas", "Rentabilidad", "Sugerido", "Claims",
             "Upload", "Sistema"] if IS_LOCAL else [
             "Alertas", "Dashboard", "Semanal", "Tendencias", "Productos",
             "Tiendas", "Rentabilidad", "Sugerido", "Claims", "Sistema"]
        page = st.radio(
            "Navegacion",
            nav_pages,
            label_visibility="collapsed"
        )

        st.divider()

        # Quick stats in sidebar
        rl_count = query_val("SELECT COUNT(*) FROM retail_link") or 0
        max_week = query_val("SELECT MAX(semana_wm) FROM retail_link") or "—"
        st.caption(f"{rl_count:,} registros")
        st.caption(f"Hasta: {wm_week_to_date(max_week)}")

        # Sync button — only on local machine
        if IS_LOCAL:
            st.divider()
            if st.button("Sincronizar Cloud", type="primary", use_container_width=True):
                with st.spinner("Subiendo datos..."):
                    try:
                        _git = lambda cmd: subprocess.run(
                            ["git"] + cmd, cwd=BASE_DIR,
                            capture_output=True, text=True, timeout=120
                        )
                        _git(["add", "walmart.db", "app.py", "import_walmart.py",
                              "requirements.txt", "ClanTrack.png"])
                        _git(["commit", "-m", f"sync: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
                        result = _git(["push"])
                        if result.returncode == 0:
                            st.success("Datos sincronizados")
                        else:
                            st.error(f"Error push: {result.stderr}")
                    except Exception as e:
                        st.error(f"Error: {e}")

    # Router
    pages = {
        "Alertas": page_alertas,
        "Dashboard": page_dashboard,
        "Semanal": page_semanal,
        "Tendencias": page_tendencias,
        "Productos": page_productos,
        "Tiendas": page_tiendas,
        "Rentabilidad": page_rentabilidad,
        "Sugerido": page_sugerido,
        "Claims": page_claims,
        "Upload": page_upload,
        "Sistema": page_status,
    }

    pages[page]()


if __name__ == '__main__':
    main()
