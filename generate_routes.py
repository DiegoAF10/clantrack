#!/usr/bin/env python3
"""
RUTAS TRASTIENDA — Generador de hojas de ruta para merchandising
Identifica producto parado en bodega y genera checklists por tienda
"""

import sqlite3
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "walmart.db")
OUTPUT_DIR = r"C:\Users\Diego\OneDrive - Clan Cervecero\CLAN CERVECERO\WALMART\REPORTES"
REPORT_DIR = os.path.join(BASE_DIR, "reports")

# Styles
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2D4A7A", end_color="2D4A7A", fill_type="solid")
BRAND_FONT = Font(bold=True, size=11, color="2D4A7A")
BRAND_FILL = PatternFill(start_color="E8EDF5", end_color="E8EDF5", fill_type="solid")
ZERO_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
TOP5_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
TOP15_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
CHECK_FONT = Font(size=14)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def generate_routes(min_inv=10, min_coverage=60, top_n=15):
    conn = sqlite3.connect(DB_PATH)
    sale_weeks = [w[0] for w in conn.execute(
        "SELECT DISTINCT semana_wm FROM retail_link "
        "WHERE tipo_registro='VENTA' ORDER BY semana_wm DESC LIMIT 4"
    ).fetchall()]
    inv_week = conn.execute(
        "SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO'"
    ).fetchone()[0]

    ph = ",".join(["?"] * len(sale_weeks))
    days = len(sale_weeks) * 7

    rows = conn.execute(f"""
        WITH ventas AS (
            SELECT producto, store_nbr, SUM(venta_und) as venta_4sem
            FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm IN ({ph})
            GROUP BY producto, store_nbr
        ),
        inventario AS (
            SELECT producto, store_nbr, MAX(inv_actual) as inv
            FROM retail_link WHERE tipo_registro='INVENTARIO' AND semana_wm=?
            GROUP BY producto, store_nbr
        ),
        network AS (
            SELECT producto,
                   AVG(CASE WHEN venta_4sem > 0 THEN venta_4sem * 1.0 / {days}
                       ELSE NULL END) as avg_rot,
                   COUNT(CASE WHEN venta_4sem > 0 THEN 1 END) as tiendas_vendiendo
            FROM ventas GROUP BY producto
        )
        SELECT i.producto, i.store_nbr,
               COALESCE(t.nombre, 'T-' || i.store_nbr) as tienda,
               t.formato, t.region,
               i.inv, COALESCE(v.venta_4sem, 0) as venta,
               n.avg_rot, n.tiendas_vendiendo,
               p.ingreso_und_wm, p.und_x_caja, p.marca, p.codigo_barras
        FROM inventario i
        LEFT JOIN ventas v ON i.producto = v.producto AND i.store_nbr = v.store_nbr
        LEFT JOIN network n ON i.producto = n.producto
        LEFT JOIN productos p ON i.producto = p.producto
        LEFT JOIN tiendas t ON i.store_nbr = t.no_tienda
        WHERE p.estado = 'Activo'
          AND i.inv >= {min_inv}
          AND (COALESCE(v.venta_4sem, 0) = 0
               OR i.inv * {days}.0 / COALESCE(v.venta_4sem, 1) > {min_coverage})
          AND n.avg_rot > 0.05
        ORDER BY i.inv * COALESCE(p.ingreso_und_wm, 1) DESC
    """, [*sale_weeks, inv_week]).fetchall()

    # Group by store
    stores = {}
    for r in rows:
        key = (r[1], r[2], r[3] or "?", r[4] or "?")
        if key not in stores:
            stores[key] = {"items": [], "total_und": 0, "total_q": 0}
        stores[key]["items"].append({
            "producto": r[0], "inv": r[5], "venta": r[6],
            "avg_rot": r[7], "vendiendo": r[8], "costo": r[9],
            "uxc": r[10], "marca": r[11], "barcode": r[12],
        })
        stores[key]["total_und"] += r[5]
        stores[key]["total_q"] += r[5] * r[9]

    sorted_stores = sorted(stores.items(), key=lambda x: -x[1]["total_q"])
    print(f"Total stores with backroom alerts: {len(sorted_stores)}")
    print(f"Total units stuck: {sum(d['total_und'] for _, d in sorted_stores):,}")
    print(f"Total value: Q{sum(d['total_q'] for _, d in sorted_stores):,.0f}")

    # === BUILD EXCEL ===
    wb = openpyxl.Workbook()

    # --- RESUMEN sheet ---
    ws = wb.active
    ws.title = "RESUMEN RUTAS"

    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value="RUTAS DE MERCHANDISING - TRASTIENDA")
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1,
            value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, color="666666")
    ws.cell(row=3, column=1,
            value=f"Inventario: Semana WM {inv_week} | Ventas: {days}d").font = Font(italic=True, color="666666")
    ws.cell(row=4, column=1,
            value="INSTRUCCION: Buscar producto en bodega/trastienda y colocarlo en piso de ventas"
            ).font = Font(bold=True, color="C62828")

    sum_headers = ["#", "Tienda", "Formato", "Region", "Productos",
                   "Und Paradas", "Valor Q", "Sin Ventas 28d"]
    for i, h in enumerate(sum_headers, 1):
        c = ws.cell(row=6, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN

    for rank, ((sn, name, fmt, reg), data) in enumerate(sorted_stores, 1):
        r = 6 + rank
        zero = sum(1 for it in data["items"] if it["venta"] == 0)
        vals = [rank, f"{name} (#{sn})", fmt, reg, len(data["items"]),
                data["total_und"], round(data["total_q"]), f"{zero}/{len(data['items'])}"]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = THIN
            if rank <= 5:
                c.fill = TOP5_FILL
            elif rank <= 15:
                c.fill = TOP15_FILL

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    for col in "EFGH":
        ws.column_dimensions[col].width = 14
    ws.freeze_panes = "A7"

    # --- Individual route sheets ---
    for rank, ((sn, name, fmt, reg), data) in enumerate(sorted_stores[:top_n], 1):
        safe_name = name.replace("/", "-")[:25]
        sheet_name = f"R{rank:02d} {safe_name}"
        ws = wb.create_sheet(title=sheet_name)

        # Title
        ws.merge_cells("A1:H1")
        c = ws.cell(row=1, column=1, value=f"RUTA #{rank} - {name} (#{sn})")
        c.font = TITLE_FONT
        c.fill = TITLE_FILL
        c.alignment = Alignment(horizontal="center")

        ws.cell(row=2, column=1, value=f"Formato: {fmt}").font = Font(bold=True)
        ws.cell(row=2, column=3, value=f"Region: {reg}").font = Font(bold=True)

        zero = sum(1 for it in data["items"] if it["venta"] == 0)
        ws.cell(row=3, column=1, value=f"Productos: {len(data['items'])}").font = Font(bold=True)
        ws.cell(row=3, column=3,
                value=f"Und paradas: {data['total_und']}").font = Font(bold=True, color="C62828")
        ws.cell(row=3, column=5,
                value=f"Valor: Q{data['total_q']:,.0f}").font = Font(bold=True, color="C62828")
        ws.cell(row=3, column=7,
                value=f"Sin ventas: {zero}/{len(data['items'])}").font = Font(bold=True, color="C62828")

        ws.cell(row=4, column=1,
                value="BUSCAR en bodega/trastienda. Marcar OK si se coloco en piso."
                ).font = Font(bold=True, italic=True, color="2D4A7A")

        # Column headers
        headers = ["OK", "Marca", "Producto", "Buscar (cajas aprox)",
                   "Und en Sistema", "Venta 28d", "Barcode", "Notas Campo"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=6, column=i, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.border = THIN
            c.alignment = Alignment(horizontal="center", wrap_text=True)

        # Group by brand
        brands = {}
        for item in data["items"]:
            m = item["marca"]
            if m not in brands:
                brands[m] = []
            brands[m].append(item)

        row = 7
        for marca, prods in sorted(brands.items(), key=lambda x: -sum(p["inv"] for p in x[1])):
            total_m_und = sum(p["inv"] for p in prods)
            total_m_cajas = sum(round(p["inv"] / (p["uxc"] or 20), 1) for p in prods)
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row=row, column=1,
                        value=f"{marca.upper()} - {total_m_und} und (~{total_m_cajas:.0f} cajas)")
            c.font = BRAND_FONT
            c.fill = BRAND_FILL
            row += 1

            for p in sorted(prods, key=lambda x: -x["inv"]):
                uxc = p["uxc"] or 20
                cajas = round(p["inv"] / uxc, 1)
                cajas_txt = f"~{cajas:.0f} cajas" if cajas >= 1 else f"{p['inv']} und sueltas"
                status = "SIN VENTAS" if p["venta"] == 0 else str(p["venta"])
                fill = ZERO_FILL if p["venta"] == 0 else LOW_FILL

                ws.cell(row=row, column=1, value="[ ]").font = CHECK_FONT
                ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=2, value=p["marca"])
                ws.cell(row=row, column=3, value=p["producto"])
                ws.cell(row=row, column=4, value=cajas_txt).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=5, value=p["inv"]).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=6, value=status).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=7, value=str(p["barcode"]))
                ws.cell(row=row, column=8, value="")

                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = THIN
                    ws.cell(row=row, column=col).fill = fill

                row += 1

            row += 1  # Space between brands

        # Signature block
        row += 1
        ws.cell(row=row, column=1, value="Visitado por:").font = Font(bold=True)
        ws.cell(row=row, column=3, value="_________________________")
        row += 1
        ws.cell(row=row, column=1, value="Fecha visita:").font = Font(bold=True)
        ws.cell(row=row, column=3, value="_________________________")
        row += 1
        ws.cell(row=row, column=1, value="Contacto tienda:").font = Font(bold=True)
        ws.cell(row=row, column=3, value="_________________________")
        row += 2
        ws.cell(row=row, column=1, value="Observaciones:").font = Font(bold=True)
        ws.merge_cells(f"A{row+1}:H{row+3}")
        for rr in range(row + 1, row + 4):
            for cc in range(1, 9):
                ws.cell(row=rr, column=cc).border = THIN

        # Column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 32
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 20
        ws.freeze_panes = "A7"

    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(REPORT_DIR, exist_ok=True)
    filename = f"RUTAS_TRASTIENDA_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    print(f"\nSaved: {filepath}")

    local = os.path.join(REPORT_DIR, filename)
    wb.save(local)
    print(f"Also: {local}")

    conn.close()
    return filepath


if __name__ == "__main__":
    generate_routes()
