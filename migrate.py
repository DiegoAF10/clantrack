#!/usr/bin/env python3
"""
Migrate MASTER WALMART.xlsx → walmart.db (SQLite)
One-time migration script. Reads all sheets and populates the database.
"""

import sqlite3
import os
import sys
from datetime import datetime

# Fix encoding for Windows console
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

MASTER_PATH = r"C:\Users\Diego\OneDrive - Clan Cervecero\CLAN CERVECERO\WALMART\REPORTES\MASTER WALMART.xlsx"
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "walmart.db")
SCHEMA_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "schema.sql")


def safe_int(v, default=0):
    if v is None:
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def safe_float(v, default=0.0):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def safe_str(v, default=''):
    if v is None:
        return default
    return str(v).strip()


def parse_date(v):
    """Parse various date formats to YYYY-MM-DD string."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s:
        return None
    # Try common formats
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y/%m/%d', '%Y-%m-%d', '%d %b %Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s.split('.')[0], fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    # If it looks like "14 ene 2026" etc, try Spanish months
    spanish_months = {
        'ene': '01', 'feb': '02', 'mar': '03', 'abr': '04',
        'may': '05', 'jun': '06', 'jul': '07', 'ago': '08',
        'sep': '09', 'oct': '10', 'nov': '11', 'dic': '12'
    }
    parts = s.lower().split()
    if len(parts) == 3 and parts[1] in spanish_months:
        try:
            return f"{parts[2]}-{spanish_months[parts[1]]}-{parts[0].zfill(2)}"
        except:
            pass
    return s  # Return as-is if can't parse


def extract_year_month(semana_wm, diario):
    """Extract year and month from semana_walmart or diario."""
    # semana_wm format: "202423" -> year=2024, week=23
    s = safe_str(semana_wm)
    if len(s) == 6 and s.isdigit():
        year = int(s[:4])
        # Approximate month from week number
        week = int(s[4:])
        month = min(12, max(1, (week * 7) // 30 + 1))
        # If we have a diario date, prefer that
        d = parse_date(diario)
        if d and len(d) >= 7:
            try:
                year = int(d[:4])
                month = int(d[5:7])
            except:
                pass
        return year, month
    return None, None


def init_db():
    """Create database and schema."""
    print(f"Creating database: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
        conn.executescript(f.read())
    conn.commit()
    return conn


def migrate_catalogo_productos(wb, conn):
    """Migrate ⚙️ Catálogo_Productos."""
    ws = wb['⚙️ Catálogo_Productos']
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # Skip header rows 1-2
    count = 0
    for row in rows:
        if not row or not row[0]:
            continue
        producto = safe_str(row[0])
        if not producto or producto.startswith('Nota:'):
            continue
        conn.execute("""
            INSERT OR REPLACE INTO productos
            (producto, ingreso_wm_date, precio_publico, und_x_caja, litros,
             ingreso_caja_wm, ingreso_und_wm, marca, codigo, costo_clan_caja,
             costo_clan_und, item_wm, item_wm_ant, codigo_barras, costo_und_usd,
             estado, temporada_ini, temporada_fin)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            producto,
            parse_date(row[1]),
            safe_float(row[2]),
            safe_int(row[4]),
            safe_float(row[5]),
            safe_float(row[6]),
            safe_float(row[7]),
            safe_str(row[8]),
            safe_str(row[9]),
            safe_float(row[10]),
            safe_float(row[11]),
            safe_str(row[12]),
            safe_str(row[13]),
            safe_str(row[14]),
            safe_float(row[15]),
            safe_str(row[16]) or 'Activo',
            parse_date(row[17]) if len(row) > 17 else None,
            parse_date(row[18]) if len(row) > 18 else None,
        ))
        count += 1
    conn.commit()
    print(f"  Productos: {count} imported")
    return count


def migrate_catalogo_tiendas(wb, conn):
    """Migrate ⚙️ Catálogo_Tiendas."""
    ws = wb['⚙️ Catálogo_Tiendas']
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header row 1
    count = 0
    for row in rows:
        if not row or not row[0]:
            continue
        conn.execute("""
            INSERT OR REPLACE INTO tiendas (no_tienda, nombre, formato, region, estado)
            VALUES (?,?,?,?,?)
        """, (
            safe_int(row[0]),
            safe_str(row[1]),
            safe_str(row[2]),
            safe_str(row[3]),
            safe_str(row[4]) or 'Activo',
        ))
        count += 1
    conn.commit()
    print(f"  Tiendas: {count} imported")
    return count


def migrate_rl_master(wb, conn):
    """Migrate 📥 RL_Master (the big one: 242K rows)."""
    ws = wb['📥 RL_Master']
    count = 0
    duped = 0
    batch = []
    BATCH_SIZE = 5000

    print("  RL_Master: reading (242K+ rows, this takes a minute)...")

    for i, row in enumerate(ws.iter_rows(min_row=6, values_only=True)):  # Skip 5 header rows
        if not row or not row[0]:
            continue

        semana = safe_str(row[0])
        if not semana or not semana[:4].isdigit():
            continue

        item_nbr = safe_str(row[3])
        store_nbr = safe_int(row[5])
        diario = safe_str(row[1])

        year, month = extract_year_month(semana, diario)

        # RL_Master doesn't have tipo_registro — infer from diario
        tipo = 'VENTA' if diario and '/' in diario else 'INVENTARIO'

        batch.append((
            'rl_master', tipo, semana, diario,
            item_nbr, safe_str(row[4]),
            store_nbr, safe_str(row[6]),
            safe_float(row[7]), safe_float(row[8]),
            safe_float(row[9]), safe_float(row[10]),
            safe_int(row[11]), safe_float(row[12]), safe_float(row[13]),
            safe_float(row[14]), safe_float(row[15]), safe_int(row[16]),
            safe_int(row[17]), safe_float(row[18]),
            safe_int(row[19]), safe_float(row[20]),
            safe_float(row[21]), safe_float(row[22]),
            safe_float(row[23]), safe_float(row[24]),
            safe_int(row[25]),
            safe_str(row[26]), safe_int(row[27]), safe_str(row[28]),
            safe_int(row[29]), safe_int(row[30]),
            safe_str(row[31]), safe_float(row[32]),
            None, None, None, None, None,  # FormatoRepSemanal extras
            safe_str(row[33]), safe_str(row[34]),
            year, month,
        ))

        if len(batch) >= BATCH_SIZE:
            result = flush_rl_batch(conn, batch)
            count += result[0]
            duped += result[1]
            batch = []
            if count % 50000 == 0:
                print(f"    ...{count:,} rows processed")

    if batch:
        result = flush_rl_batch(conn, batch)
        count += result[0]
        duped += result[1]

    conn.commit()
    print(f"  RL_Master: {count:,} imported, {duped:,} duplicates skipped")
    return count


def flush_rl_batch(conn, batch):
    added = 0
    duped = 0
    for row in batch:
        try:
            conn.execute("""
                INSERT OR IGNORE INTO retail_link
                (source, tipo_registro, semana_wm, diario,
                 item_nbr, producto, store_nbr, tienda,
                 costo_q, costo_usd, precio_q, precio_usd,
                 venta_und, venta_q, venta_usd,
                 venta_comp_q, venta_comp_usd, venta_comp_und,
                 inv_actual, sem_abasto, pedido_actual, tasa_venta,
                 venta_costo_q, venta_costo_usd,
                 instock_act_pct, instock_prom_pct, tiendas_validas,
                 cod_faltante, cant_faltante, fecha_faltante,
                 pronostico_52s, inv_historico,
                 fecha_tc, tipo_cambio,
                 margen_pct, cod_barras, max_estante, pedido_transito, inv_liquidacion,
                 codigo_interno, estado_producto,
                 anio, mes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, row)
            if conn.total_changes:
                added += 1
        except sqlite3.IntegrityError:
            duped += 1
    return (added, duped)


def migrate_formato_semanal(wb, conn):
    """Migrate FormatoRepSemanal_RL (recent weekly data with INVENTARIO rows)."""
    ws = wb['FormatoRepSemanal_RL']
    count = 0
    duped = 0
    batch = []

    print("  FormatoRepSemanal: reading...")

    for row in ws.iter_rows(min_row=6, values_only=True):  # Skip 5 header rows
        if not row or not row[0]:
            continue

        semana = safe_str(row[0])
        if not semana or not semana[:4].isdigit():
            continue

        diario = safe_str(row[1])
        tipo = safe_str(row[40]) if len(row) > 40 else ('VENTA' if diario and '/' in diario else 'INVENTARIO')

        year, month = extract_year_month(semana, diario)

        batch.append((
            'formato_semanal', tipo, semana, diario,
            safe_str(row[2]),   # item_nbr (col 2 in FormatoSemanal)
            safe_str(row[3]),   # producto
            safe_int(row[12]),  # store_nbr (col 12)
            safe_str(row[13]),  # tienda
            safe_float(row[5]), safe_float(row[6]),   # costo_q, costo_usd
            safe_float(row[7]), safe_float(row[8]),   # precio_q, precio_usd
            safe_int(row[16]),  # venta_und
            safe_float(row[14]), safe_float(row[15]), # venta_q, venta_usd
            safe_float(row[17]), safe_float(row[18]), safe_int(row[19]),  # venta_comp
            safe_int(row[27]),  # inv_actual
            None,               # sem_abasto (not in FormatoSemanal)
            safe_int(row[32]),  # pedido_actual
            safe_float(row[23]),  # tasa_venta
            safe_float(row[20]), safe_float(row[21]),  # venta_costo
            safe_float(row[30]), safe_float(row[29]),  # instock_act, instock_prom
            safe_int(row[28]),   # tiendas_validas
            safe_str(row[35]),   # cod_faltante
            safe_int(row[36]),   # cant_faltante
            safe_str(row[37]),   # fecha_faltante
            safe_int(row[34]),   # pronostico_52s
            None,                # inv_historico (not in FormatoSemanal)
            safe_str(row[38]), safe_float(row[39]),  # fecha_tc, tipo_cambio
            # Extra FormatoSemanal columns
            safe_float(row[9]),   # margen_pct
            safe_str(row[10]),    # cod_barras
            safe_int(row[22]),    # max_estante
            safe_int(row[31]),    # pedido_transito
            safe_int(row[33]),    # inv_liquidacion
            safe_str(row[41]) if len(row) > 41 else None,  # codigo_interno
            safe_str(row[42]) if len(row) > 42 else None,  # estado_producto
            year, month,
        ))

    # Flush all at once (only ~9K rows)
    for r in batch:
        try:
            conn.execute("""
                INSERT OR IGNORE INTO retail_link
                (source, tipo_registro, semana_wm, diario,
                 item_nbr, producto, store_nbr, tienda,
                 costo_q, costo_usd, precio_q, precio_usd,
                 venta_und, venta_q, venta_usd,
                 venta_comp_q, venta_comp_usd, venta_comp_und,
                 inv_actual, sem_abasto, pedido_actual, tasa_venta,
                 venta_costo_q, venta_costo_usd,
                 instock_act_pct, instock_prom_pct, tiendas_validas,
                 cod_faltante, cant_faltante, fecha_faltante,
                 pronostico_52s, inv_historico,
                 fecha_tc, tipo_cambio,
                 margen_pct, cod_barras, max_estante, pedido_transito, inv_liquidacion,
                 codigo_interno, estado_producto,
                 anio, mes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, r)
            count += 1
        except sqlite3.IntegrityError:
            duped += 1

    conn.commit()
    print(f"  FormatoRepSemanal: {count:,} imported, {duped:,} duplicates skipped")
    return count


def migrate_sell_in(wb, conn):
    """Migrate 📝 Sell_In_Data."""
    ws = wb['📝 Sell_In_Data']
    count = 0

    for row in ws.iter_rows(min_row=5, values_only=True):  # Skip header rows 1-4
        if not row or not row[0]:
            continue
        codigo = safe_str(row[0])
        if not codigo or codigo.startswith('INSTRUCCIONES'):
            continue

        try:
            conn.execute("""
                INSERT OR IGNORE INTO sell_in
                (codigo_producto, nombre_producto, cliente, fecha, cantidad_facturada,
                 fecha_esperada, producto_precios, und_x_caja, cajas_facturadas,
                 precio_caja_wm, costo_caja_clan, ingreso_bruto, costo_total,
                 centralizacion, ingreso_neto, margen_bruto, margen_pct, marca, anio_mes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                codigo,
                safe_str(row[1]),       # nombre_producto
                safe_str(row[2]),       # cliente
                parse_date(row[3]),     # fecha
                safe_int(row[4]),       # cantidad_facturada
                parse_date(row[5]),     # fecha_esperada
                safe_str(row[6]),       # producto_precios
                # row[7] = Fecha Num (skip — internal calc)
                safe_int(row[8]) if len(row) > 8 else None,    # und_x_caja
                safe_float(row[9]) if len(row) > 9 else None,  # cajas_facturadas
                safe_float(row[10]) if len(row) > 10 else None, # precio_caja_wm
                safe_float(row[11]) if len(row) > 11 else None, # costo_caja_clan
                safe_float(row[12]) if len(row) > 12 else None, # ingreso_bruto
                safe_float(row[13]) if len(row) > 13 else None, # costo_total
                safe_float(row[14]) if len(row) > 14 else None, # centralizacion 4%
                safe_float(row[15]) if len(row) > 15 else None, # ingreso_neto
                safe_float(row[16]) if len(row) > 16 else None, # margen_bruto
                safe_float(row[17]) if len(row) > 17 else None, # % margen
                safe_str(row[18]) if len(row) > 18 else None,   # marca
                safe_str(row[19]) if len(row) > 19 else None,   # anio_mes
            ))
            count += 1
        except sqlite3.IntegrityError:
            pass

    conn.commit()
    print(f"  Sell_In_Data: {count} imported")
    return count


def migrate_devoluciones(wb, conn):
    """Migrate Devoluciones."""
    ws = wb['Devoluciones']
    count = 0

    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip 3 header rows
        if not row or not row[0]:
            continue
        semana = safe_str(row[0])
        if not semana or not semana[:4].isdigit():
            continue

        # Some rows have sidebar data (conciliation dashboard) — skip those
        item_nbr = safe_str(row[2]) if len(row) > 2 else ''
        if not item_nbr or not item_nbr.isdigit():
            continue

        conn.execute("""
            INSERT INTO devoluciones
            (semana_wm, diario, item_nbr, descripcion, pais,
             cant_reclamo_tienda, precio_reclamo_q, precio_reclamo_usd,
             costo_reclamo_q, costo_reclamo_usd,
             devolucion_cliente_und, costo_devolucion_q, costo_devolucion_usd,
             precio_devolucion_q, precio_devolucion_usd,
             fecha_tc, tipo_cambio, estado)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            semana,
            safe_str(row[1]),
            item_nbr,
            safe_str(row[3]),
            safe_str(row[4]),
            safe_float(row[5]),
            safe_float(row[6]),
            safe_float(row[7]),
            safe_float(row[8]),
            safe_float(row[9]),
            safe_float(row[10]),
            safe_float(row[11]),
            safe_float(row[12]),
            safe_float(row[13]),
            safe_float(row[14]),
            safe_str(row[15]),
            safe_float(row[16]),
            safe_str(row[17]) if len(row) > 17 else None,
        ))
        count += 1

    conn.commit()
    print(f"  Devoluciones: {count} imported")
    return count


def print_summary(conn):
    """Print migration summary."""
    print("\n═══════════════════════════════════════════")
    print("  MIGRATION COMPLETE")
    print("═══════════════════════════════════════════")

    tables = ['retail_link', 'sell_in', 'devoluciones', 'productos', 'tiendas']
    for table in tables:
        count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        print(f"  {table:20s} {count:>10,} rows")

    # Additional stats
    rl_weeks = conn.execute("SELECT MIN(semana_wm), MAX(semana_wm) FROM retail_link").fetchone()
    print(f"\n  Retail Link period: {rl_weeks[0]} → {rl_weeks[1]}")

    rl_types = conn.execute("""
        SELECT tipo_registro, COUNT(*) FROM retail_link
        GROUP BY tipo_registro
    """).fetchall()
    for tipo, cnt in rl_types:
        print(f"    {tipo or 'NULL':15s} {cnt:>10,}")

    db_size = os.path.getsize(DB_PATH) / (1024 * 1024)
    print(f"\n  Database size: {db_size:.1f} MB")
    print("═══════════════════════════════════════════")


def main():
    print("╔═══════════════════════════════════════════╗")
    print("║  MASTER WALMART → SQLite Migration        ║")
    print("╚═══════════════════════════════════════════╝")
    print()

    # Remove existing DB for fresh migration
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print(f"Removed existing {DB_PATH}")

    conn = init_db()

    print(f"\nOpening {MASTER_PATH}...")
    print("(This may take 1-2 minutes for 242K+ rows)")
    wb = openpyxl.load_workbook(MASTER_PATH, read_only=True, data_only=True)

    print("\nMigrating...")
    migrate_catalogo_productos(wb, conn)
    migrate_catalogo_tiendas(wb, conn)
    migrate_rl_master(wb, conn)
    migrate_formato_semanal(wb, conn)
    migrate_sell_in(wb, conn)
    migrate_devoluciones(wb, conn)

    # Log the migration
    conn.execute("""
        INSERT INTO import_log (source, filename, rows_added)
        VALUES ('migration', 'MASTER WALMART.xlsx',
                (SELECT COUNT(*) FROM retail_link))
    """)
    conn.commit()

    print_summary(conn)

    wb.close()
    conn.close()


if __name__ == '__main__':
    main()
