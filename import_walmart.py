#!/usr/bin/env python3
"""
Import new Walmart data into walmart.db
Supports two input modes:
  1. PASTE MODE: Read from WM-INPUT.xlsx template (Diego pastes data there)
  2. DROP MODE:  Read from files dropped in imports/ folder

Usage:
  python import_walmart.py              # Checks both sources
  python import_walmart.py --template   # Only read from WM-INPUT.xlsx
  python import_walmart.py --drop       # Only read from imports/ folder
  python import_walmart.py --report     # Generate report after import
"""

import sqlite3
import os
import sys
import glob
import argparse
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "walmart.db")
TEMPLATE_PATH = os.path.join(BASE_DIR, "templates", "WM-INPUT.xlsx")
IMPORTS_DIR = os.path.join(BASE_DIR, "imports")
ARCHIVE_DIR = os.path.join(BASE_DIR, "imports", "_processed")


def safe_int(v, default=0):
    if v is None: return default
    try: return int(float(v))
    except: return default

def safe_float(v, default=0.0):
    if v is None: return default
    try: return float(v)
    except: return default

def safe_str(v, default=''):
    if v is None: return default
    return str(v).strip()

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s: return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y/%m/%d', '%Y-%m-%d', '%m/%d/%Y'):
        try: return datetime.strptime(s.split('.')[0], fmt).strftime('%Y-%m-%d')
        except: continue
    spanish = {'ene':'01','feb':'02','mar':'03','abr':'04','may':'05','jun':'06',
               'jul':'07','ago':'08','sep':'09','oct':'10','nov':'11','dic':'12'}
    parts = s.lower().split()
    if len(parts) == 3 and parts[1] in spanish:
        try: return f"{parts[2]}-{spanish[parts[1]]}-{parts[0].zfill(2)}"
        except: pass
    return s

def extract_year_month(semana_wm, diario):
    s = safe_str(semana_wm)
    if len(s) == 6 and s.isdigit():
        year = int(s[:4])
        week = int(s[4:])
        month = min(12, max(1, (week * 7) // 30 + 1))
        d = parse_date(diario)
        if d and len(d) >= 7:
            try:
                year = int(d[:4])
                month = int(d[5:7])
            except: pass
        return year, month
    return None, None


def import_retail_link(conn, rows, source_name):
    """Import Retail Link rows. Handles both old (41-col) and new (50-col) formats.
    Two-pass strategy:
      Pass 1: INSERT OR IGNORE (new rows)
      Pass 2: UPDATE reclamo/devolución columns on existing rows (new-format only)
    """
    before_count = conn.execute("SELECT COUNT(*) FROM retail_link").fetchone()[0]
    updated = 0
    processed = 0

    for row in rows:
        if not row or len(row) < 28:
            continue
        # Semana may come as float (202548.0) from openpyxl — normalize to int string
        try:
            semana = str(int(float(row[0])))
        except (ValueError, TypeError):
            semana = safe_str(row[0])
        if not semana or not semana[:4].isdigit():
            continue
        processed += 1

        diario = safe_str(row[1])

        # Detect format: new (50-col) has TC at [48-49], old (41-col) at [38-39]
        is_new_fmt = len(row) > 47

        # Determine tipo_registro
        tipo = None
        if not is_new_fmt and len(row) > 40 and row[40]:
            tipo = safe_str(row[40])
        if not tipo:
            tipo = 'VENTA' if diario and '/' in diario else 'INVENTARIO'

        year, month = extract_year_month(semana, diario)

        # Column indices that differ between formats
        if is_new_fmt:
            fecha_tc_idx, tc_idx = 48, 49
        else:
            fecha_tc_idx, tc_idx = 38, 39

        # item_nbr may come as float (75214930.0) — normalize
        try:
            item_nbr = str(int(float(row[2])))
        except (ValueError, TypeError):
            item_nbr = safe_str(row[2])
        store_nbr = safe_int(row[12])

        # --- Pass 1: INSERT new rows ---
        conn.execute("""
            INSERT OR IGNORE INTO retail_link
            (source, tipo_registro, semana_wm, diario,
             item_nbr, producto, store_nbr, tienda,
             costo_q, costo_usd, precio_q, precio_usd,
             venta_und, venta_q, venta_usd,
             venta_comp_q, venta_comp_usd, venta_comp_und,
             inv_actual, sem_abasto, pedido_actual, tasa_venta,
             venta_costo_q, venta_costo_usd,
             instock_act_pct, instock_prom_pct, tiendas_validas,
             cod_faltante, cant_faltante, fecha_faltante,
             pronostico_52s, inv_historico,
             fecha_tc, tipo_cambio,
             margen_pct, cod_barras, max_estante, pedido_transito, inv_liquidacion,
             codigo_interno, estado_producto,
             cant_reclamo_proveedor, precio_reclamo_proveedor_q, precio_reclamo_proveedor_usd,
             costo_reclamo_proveedor_q, costo_reclamo_proveedor_usd,
             devolucion_cliente_total, precio_devolucion_total_q, precio_devolucion_total_usd,
             costo_devolucion_total_q, costo_devolucion_total_usd,
             anio, mes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            source_name, tipo, semana, diario,
            item_nbr,
            safe_str(row[3]),    # producto
            store_nbr,
            safe_str(row[13]),   # tienda
            safe_float(row[5]), safe_float(row[6]),    # costo
            safe_float(row[7]), safe_float(row[8]),    # precio
            safe_int(row[16]),                          # venta_und
            safe_float(row[14]), safe_float(row[15]),  # venta_q/usd
            safe_float(row[17]), safe_float(row[18]), safe_int(row[19]),  # comp
            safe_int(row[27]),   # inv_actual
            None,                # sem_abasto
            safe_int(row[32]) if len(row) > 32 else 0,  # pedido_actual
            safe_float(row[23]) if len(row) > 23 else 0, # tasa_venta
            safe_float(row[20]), safe_float(row[21]),  # venta_costo
            safe_float(row[30]) if len(row) > 30 else None,  # instock_act
            safe_float(row[29]) if len(row) > 29 else None,  # instock_prom
            safe_int(row[28]) if len(row) > 28 else None,    # tiendas_validas
            safe_str(row[35]) if len(row) > 35 else None,
            safe_int(row[36]) if len(row) > 36 else None,
            safe_str(row[37]) if len(row) > 37 else None,
            safe_int(row[34]) if len(row) > 34 else None,
            None,  # inv_historico
            safe_str(row[fecha_tc_idx]) if len(row) > fecha_tc_idx else None,
            safe_float(row[tc_idx]) if len(row) > tc_idx else None,
            safe_float(row[9]) if len(row) > 9 else None,   # margen_pct
            safe_str(row[10]) if len(row) > 10 else None,   # cod_barras
            safe_int(row[22]) if len(row) > 22 else None,   # max_estante
            safe_int(row[31]) if len(row) > 31 else None,   # pedido_transito
            safe_int(row[33]) if len(row) > 33 else None,   # inv_liquidacion
            None, None,  # codigo_interno, estado_producto
            # Reclamo/devolución (new format only)
            safe_float(row[38]) if is_new_fmt and len(row) > 38 else 0,
            safe_float(row[39]) if is_new_fmt and len(row) > 39 else 0,
            safe_float(row[40]) if is_new_fmt and len(row) > 40 else 0,
            safe_float(row[41]) if is_new_fmt and len(row) > 41 else 0,
            safe_float(row[42]) if is_new_fmt and len(row) > 42 else 0,
            safe_float(row[43]) if is_new_fmt and len(row) > 43 else 0,
            safe_float(row[44]) if is_new_fmt and len(row) > 44 else 0,
            safe_float(row[45]) if is_new_fmt and len(row) > 45 else 0,
            safe_float(row[46]) if is_new_fmt and len(row) > 46 else 0,
            safe_float(row[47]) if is_new_fmt and len(row) > 47 else 0,
            year, month,
        ))

        # --- Pass 2: UPDATE reclamo columns on existing rows ---
        if is_new_fmt:
            reclamo_vals = [safe_float(row[i]) for i in range(38, 48)]
            has_reclamo = any(v != 0 for v in reclamo_vals)
            if has_reclamo:
                cur = conn.execute("""
                    UPDATE retail_link SET
                        cant_reclamo_proveedor = ?,
                        precio_reclamo_proveedor_q = ?,
                        precio_reclamo_proveedor_usd = ?,
                        costo_reclamo_proveedor_q = ?,
                        costo_reclamo_proveedor_usd = ?,
                        devolucion_cliente_total = ?,
                        precio_devolucion_total_q = ?,
                        precio_devolucion_total_usd = ?,
                        costo_devolucion_total_q = ?,
                        costo_devolucion_total_usd = ?
                    WHERE semana_wm = ? AND item_nbr = ? AND store_nbr = ?
                      AND tipo_registro = ? AND diario = ?
                """, (*reclamo_vals,
                       semana, item_nbr, store_nbr, tipo, diario))
                if cur.rowcount > 0:
                    updated += cur.rowcount

    after_count = conn.execute("SELECT COUNT(*) FROM retail_link").fetchone()[0]
    added = after_count - before_count
    duped = processed - added

    return added, duped, updated


def import_sell_in(conn, rows, source_name):
    """Import Sell-In rows (Odoo invoice data)."""
    added = 0
    duped = 0

    for row in rows:
        if not row or not row[0]:
            continue
        codigo = safe_str(row[0])
        if not codigo or not codigo.startswith('CC-'):
            continue

        try:
            conn.execute("""
                INSERT OR IGNORE INTO sell_in
                (codigo_producto, nombre_producto, cliente, fecha, cantidad_facturada,
                 fecha_esperada, producto_precios, und_x_caja, cajas_facturadas,
                 precio_caja_wm, costo_caja_clan, ingreso_bruto, costo_total,
                 centralizacion, ingreso_neto, margen_bruto, margen_pct, marca, anio_mes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                codigo,
                safe_str(row[1]),
                safe_str(row[2]),
                parse_date(row[3]),
                safe_int(row[4]),
                parse_date(row[5]) if len(row) > 5 else None,
                safe_str(row[6]) if len(row) > 6 else None,
                safe_int(row[8]) if len(row) > 8 else None,    # und_x_caja (skip row[7] = Fecha Num)
                safe_float(row[9]) if len(row) > 9 else None,
                safe_float(row[10]) if len(row) > 10 else None,
                safe_float(row[11]) if len(row) > 11 else None,
                safe_float(row[12]) if len(row) > 12 else None,
                safe_float(row[13]) if len(row) > 13 else None,
                safe_float(row[14]) if len(row) > 14 else None,
                safe_float(row[15]) if len(row) > 15 else None,
                safe_float(row[16]) if len(row) > 16 else None,
                safe_float(row[17]) if len(row) > 17 else None,
                safe_str(row[18]) if len(row) > 18 else None,
                safe_str(row[19]) if len(row) > 19 else None,
            ))
            added += 1
        except sqlite3.IntegrityError:
            duped += 1

    return added, duped


def import_devoluciones(conn, rows, source_name):
    """Import Devoluciones rows."""
    added = 0
    duped = 0

    for row in rows:
        if not row or not row[0]:
            continue
        semana = safe_str(row[0])
        if not semana or not semana[:4].isdigit():
            continue
        item_nbr = safe_str(row[2]) if len(row) > 2 else ''
        if not item_nbr or not item_nbr.isdigit():
            continue

        try:
            conn.execute("""
                INSERT OR IGNORE INTO devoluciones
                (semana_wm, diario, item_nbr, descripcion, pais,
                 cant_reclamo_tienda, precio_reclamo_q, precio_reclamo_usd,
                 costo_reclamo_q, costo_reclamo_usd,
                 devolucion_cliente_und, costo_devolucion_q, costo_devolucion_usd,
                 precio_devolucion_q, precio_devolucion_usd,
                 fecha_tc, tipo_cambio, estado)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                semana, safe_str(row[1]), item_nbr, safe_str(row[3]), safe_str(row[4]),
                safe_float(row[5]), safe_float(row[6]), safe_float(row[7]),
                safe_float(row[8]), safe_float(row[9]),
                safe_float(row[10]), safe_float(row[11]), safe_float(row[12]),
                safe_float(row[13]), safe_float(row[14]),
                safe_str(row[15]), safe_float(row[16]),
                safe_str(row[17]) if len(row) > 17 else None,
            ))
            added += 1
        except sqlite3.IntegrityError:
            duped += 1

    return added, duped


def process_template():
    """Read from WM-INPUT.xlsx template."""
    if not os.path.exists(TEMPLATE_PATH):
        print(f"  Template not found: {TEMPLATE_PATH}")
        print("  Run: python create_template.py to generate it")
        return None

    wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    results = {}

    # Sheet 1: RL Semanal
    if 'RL Semanal' in wb.sheetnames:
        ws = wb['RL Semanal']
        rows = [row for row in ws.iter_rows(min_row=2, values_only=True)]
        data_rows = [r for r in rows if r and r[0] and safe_str(r[0])[:4].isdigit()]
        results['rl'] = data_rows
        print(f"  Template RL Semanal: {len(data_rows)} data rows found")

    # Sheet 2: Sell-In
    if 'Sell-In' in wb.sheetnames:
        ws = wb['Sell-In']
        rows = [row for row in ws.iter_rows(min_row=2, values_only=True)]
        data_rows = [r for r in rows if r and r[0] and safe_str(r[0]).startswith('CC-')]
        results['sell_in'] = data_rows
        print(f"  Template Sell-In: {len(data_rows)} data rows found")

    # Sheet 3: Devoluciones
    if 'Devoluciones' in wb.sheetnames:
        ws = wb['Devoluciones']
        rows = [row for row in ws.iter_rows(min_row=2, values_only=True)]
        data_rows = [r for r in rows if r and r[0] and safe_str(r[0])[:4].isdigit()]
        results['dev'] = data_rows
        print(f"  Template Devoluciones: {len(data_rows)} data rows found")

    wb.close()
    return results


def process_dropped_files():
    """Read from files in imports/ folder."""
    results = {'rl': [], 'sell_in': [], 'dev': []}
    os.makedirs(ARCHIVE_DIR, exist_ok=True)

    files = glob.glob(os.path.join(IMPORTS_DIR, "*.xlsx")) + glob.glob(os.path.join(IMPORTS_DIR, "*.csv"))
    # Exclude files in _processed
    files = [f for f in files if '_processed' not in f]

    if not files:
        print("  No files in imports/ folder")
        return results

    for filepath in files:
        fname = os.path.basename(filepath)
        print(f"  Processing: {fname}")

        if filepath.endswith('.xlsx'):
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

            for ws in wb.worksheets:
                rows = list(ws.iter_rows(min_row=1, values_only=True))
                if not rows:
                    continue

                # Auto-detect content type from headers or first data row
                first_data = None
                for r in rows:
                    if r and r[0]:
                        first_data = r
                        break

                if first_data is None:
                    continue

                header_str = ' '.join(safe_str(v) for v in first_data[:5]).lower()

                if 'semana' in header_str and ('venta' in header_str or 'item' in header_str):
                    # Retail Link data - skip header row
                    data_rows = [r for r in rows[1:] if r and r[0] and safe_str(r[0])[:4].isdigit()]
                    results['rl'].extend(data_rows)
                    print(f"    → Detected as Retail Link: {len(data_rows)} rows")

                elif safe_str(first_data[0]).startswith('CC-') or 'código producto' in header_str:
                    data_rows = [r for r in rows[1:] if r and r[0] and safe_str(r[0]).startswith('CC-')]
                    results['sell_in'].extend(data_rows)
                    print(f"    → Detected as Sell-In: {len(data_rows)} rows")

                elif 'reclamo' in header_str or 'devolución' in header_str:
                    data_rows = [r for r in rows[1:] if r and r[0] and safe_str(r[0])[:4].isdigit()]
                    results['dev'].extend(data_rows)
                    print(f"    → Detected as Devoluciones: {len(data_rows)} rows")

                else:
                    print(f"    → Could not auto-detect type for sheet '{ws.title}'")

            wb.close()

        # Move to processed
        dest = os.path.join(ARCHIVE_DIR, f"{datetime.now().strftime('%Y%m%d_%H%M')}_{fname}")
        os.rename(filepath, dest)
        print(f"    → Moved to _processed/")

    return results


def main():
    parser = argparse.ArgumentParser(description='Import Walmart data')
    parser.add_argument('--template', action='store_true', help='Only read from WM-INPUT.xlsx')
    parser.add_argument('--drop', action='store_true', help='Only read from imports/ folder')
    parser.add_argument('--report', action='store_true', help='Generate report after import')
    args = parser.parse_args()

    print("╔═══════════════════════════════════════════╗")
    print("║  Walmart Data Import                      ║")
    print("╚═══════════════════════════════════════════╝")

    conn = sqlite3.connect(DB_PATH)

    # Before counts
    before = {
        'rl': conn.execute("SELECT COUNT(*) FROM retail_link").fetchone()[0],
        'si': conn.execute("SELECT COUNT(*) FROM sell_in").fetchone()[0],
        'dev': conn.execute("SELECT COUNT(*) FROM devoluciones").fetchone()[0],
    }

    all_data = {'rl': [], 'sell_in': [], 'dev': []}

    # Process sources
    if not args.drop:
        print("\n[Template Mode]")
        template_data = process_template()
        if template_data:
            for k in all_data:
                if k in template_data:
                    all_data[k].extend(template_data[k])

    if not args.template:
        print("\n[Drop Mode]")
        drop_data = process_dropped_files()
        for k in all_data:
            if k in drop_data:
                all_data[k].extend(drop_data[k])

    # Import
    print("\n[Importing]")
    total_added = 0
    total_duped = 0

    if all_data['rl']:
        a, d, u = import_retail_link(conn, all_data['rl'], 'import')
        total_added += a
        total_duped += d
        upd_msg = f", {u} reclamos updated" if u else ""
        print(f"  Retail Link: +{a} new, {d} duplicates{upd_msg}")
    else:
        print("  Retail Link: no new data")

    if all_data['sell_in']:
        a, d = import_sell_in(conn, all_data['sell_in'], 'import')
        total_added += a
        total_duped += d
        print(f"  Sell-In: +{a} new, {d} duplicates")
    else:
        print("  Sell-In: no new data")

    if all_data['dev']:
        a, d = import_devoluciones(conn, all_data['dev'], 'import')
        total_added += a
        total_duped += d
        print(f"  Devoluciones: +{a} new, {d} duplicates")
    else:
        print("  Devoluciones: no new data")

    # Log
    conn.execute("""
        INSERT INTO import_log (source, filename, rows_added, rows_duped)
        VALUES ('import', ?, ?, ?)
    """, (datetime.now().strftime('%Y-%m-%d'), total_added, total_duped))
    conn.commit()

    # After counts
    after = {
        'rl': conn.execute("SELECT COUNT(*) FROM retail_link").fetchone()[0],
        'si': conn.execute("SELECT COUNT(*) FROM sell_in").fetchone()[0],
        'dev': conn.execute("SELECT COUNT(*) FROM devoluciones").fetchone()[0],
    }

    print(f"\n═══════════════════════════════════════════")
    print(f"  IMPORT COMPLETE")
    print(f"  Retail Link:   {before['rl']:>10,} → {after['rl']:>10,}  (+{after['rl']-before['rl']:,})")
    print(f"  Sell-In:       {before['si']:>10,} → {after['si']:>10,}  (+{after['si']-before['si']:,})")
    print(f"  Devoluciones:  {before['dev']:>10,} → {after['dev']:>10,}  (+{after['dev']-before['dev']:,})")
    print(f"═══════════════════════════════════════════")

    conn.close()

    # Auto-generate report if requested
    if args.report:
        print("\n[Generating Report]")
        import generate_report
        generate_report.main()


if __name__ == '__main__':
    main()
