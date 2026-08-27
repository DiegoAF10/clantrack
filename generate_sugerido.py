#!/usr/bin/env python3
"""
SUGERIDO WALMART — Generador de Orden de Compra
Replica el formato SUGERIDO_EXPORT del Master Walmart.xlsx
Thresholds: 15 días objetivo, prioridades por cobertura
"""

import sqlite3
import math
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === CONFIG ===
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "walmart.db")
OUTPUT_DIR = r"C:\Users\Diego\OneDrive - Clan Cervecero\CLAN CERVECERO\WALMART\REPORTES"

TARGET_DAYS = 15
SALES_LOOKBACK_DAYS = 28  # 4 weeks
PROVEEDOR = "CLAN CERVECERO SA"
NO_PROVEEDOR = 258977
PAIS = "GT"

# Priority fills
P1_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
P2_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
P3_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2D4A7A", end_color="2D4A7A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
PRIORITY_FILLS = {1: P1_FILL, 2: P2_FILL, 3: P3_FILL}


def generate_sugerido(delivery_date=None):
    conn = sqlite3.connect(DB_PATH)

    # Sale weeks (last 4 with VENTA data)
    sale_weeks = [w[0] for w in conn.execute(
        "SELECT DISTINCT semana_wm FROM retail_link "
        "WHERE tipo_registro='VENTA' ORDER BY semana_wm DESC LIMIT 4"
    ).fetchall()]
    print(f"Sale weeks: {sale_weeks}")

    # Inventory week (latest with INVENTARIO data)
    inv_week = conn.execute(
        "SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO'"
    ).fetchone()[0]
    print(f"Inventory week: {inv_week}")

    # --- 1. Active products ---
    products = {}
    for r in conn.execute(
        "SELECT producto, item_wm, und_x_caja, ingreso_und_wm, codigo_barras, "
        "costo_und_usd, estado, temporada_ini, temporada_fin "
        "FROM productos WHERE estado IN ('Activo', 'Temporada')"
    ).fetchall():
        if r[6] == "Temporada":
            today = datetime.now()
            ini = datetime.strptime(r[7], "%Y-%m-%d") if r[7] else None
            fin = datetime.strptime(r[8], "%Y-%m-%d") if r[8] else None
            if ini and fin and not (ini <= today <= fin):
                continue
        products[r[0]] = {
            "item_wm": r[1],
            "und_x_caja": r[2] or 20,
            "ingreso_und_wm": r[3],
            "codigo_barras": r[4],
            "costo_und_usd": r[5],
        }
    print(f"Active products: {len(products)}")

    # --- 2. Sales last 4 weeks ---
    placeholders = ",".join("?" * len(sale_weeks))
    sales = {}
    for r in conn.execute(
        f"SELECT producto, store_nbr, SUM(venta_und) "
        f"FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm IN ({placeholders}) "
        f"GROUP BY producto, store_nbr",
        sale_weeks,
    ).fetchall():
        sales[(r[0], r[1])] = r[2] or 0

    # --- 3. Latest inventory (INVENTARIO) ---
    inventory = {}
    for r in conn.execute(
        "SELECT producto, store_nbr, MAX(inv_actual) "
        "FROM retail_link WHERE tipo_registro='INVENTARIO' AND semana_wm = ? "
        "GROUP BY producto, store_nbr",
        (inv_week,),
    ).fetchall():
        inventory[(r[0], r[1])] = r[2] or 0

    # Store names
    store_names = {}
    for r in conn.execute("SELECT no_tienda, nombre FROM tiendas"):
        store_names[r[0]] = r[1]
    for r in conn.execute(
        "SELECT DISTINCT store_nbr, tienda FROM retail_link "
        "WHERE tienda IS NOT NULL AND tienda != '' AND tienda != 'nan'"
    ):
        if r[0] not in store_names:
            store_names[r[0]] = r[1]

    print(f"Sales combos: {len(sales)}, Inventory combos: {len(inventory)}")

    # --- 4. Calculate sugerido ---
    all_combos = set(list(sales.keys()) + list(inventory.keys()))
    results = []
    stats = {"p1": 0, "p2": 0, "p3": 0, "p4": 0}

    for prod, store in sorted(all_combos):
        if prod not in products:
            continue

        p = products[prod]
        venta_4sem = sales.get((prod, store), 0)
        inv = inventory.get((prod, store), 0)

        rotacion_diaria = round(venta_4sem / SALES_LOOKBACK_DAYS, 2)

        if rotacion_diaria > 0:
            cobertura = round(inv / rotacion_diaria, 1)
        elif inv > 0:
            cobertura = 999
        else:
            cobertura = 0

        inv_objetivo = round(rotacion_diaria * TARGET_DAYS, 0)
        sugerido_und = max(0, int(inv_objetivo - inv))

        und_x_caja = p["und_x_caja"]
        if sugerido_und > 0:
            sugerido_cajas = max(1, math.ceil(sugerido_und / und_x_caja))
            sugerido_und_final = sugerido_cajas * und_x_caja
        else:
            sugerido_cajas = 0
            sugerido_und_final = 0

        if sugerido_cajas == 0:
            prioridad = 4
            justificacion = "Stock OK"
            stats["p4"] += 1
        elif inv <= 0 or cobertura <= 5:
            prioridad = 1
            justificacion = f"URGENTE - Cobertura {cobertura}d -> objetivo {TARGET_DAYS}d"
            stats["p1"] += 1
        elif cobertura <= 10:
            prioridad = 2
            justificacion = f"Cobertura {cobertura}d -> objetivo {TARGET_DAYS}d"
            stats["p2"] += 1
        else:
            prioridad = 3
            justificacion = f"Cobertura {cobertura}d -> objetivo {TARGET_DAYS}d"
            stats["p3"] += 1

        if sugerido_cajas > 0:
            results.append([
                prod,                                   # A
                p["item_wm"],                           # B
                PAIS,                                   # C
                p["ingreso_und_wm"],                    # D
                p["costo_und_usd"],                     # E
                p["codigo_barras"],                     # F
                PROVEEDOR,                              # G
                NO_PROVEEDOR,                           # H
                und_x_caja,                             # I
                sugerido_cajas,                         # J
                sugerido_und_final,                     # K
                store,                                  # L
                store_names.get(store, f"T-{store}"),   # M
                inv,                                    # N
                prioridad,                              # O
                justificacion,                          # P
            ])

    results.sort(key=lambda x: (x[14], x[0], x[11]))

    print(f"\nP1={stats['p1']}, P2={stats['p2']}, P3={stats['p3']}, P4(OK)={stats['p4']}")
    print(f"Lines to order: {len(results)}")
    total_cajas = sum(r[9] for r in results)
    total_und = sum(r[10] for r in results)
    total_q = sum(r[9] * r[8] * r[3] for r in results)
    print(f"Total cajas: {total_cajas}, Total und: {total_und}, Costo: Q{total_q:,.2f}")

    # --- 5. Build Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SUGERIDO_EXPORT"

    headers = [
        "Descripcion Articulo", "Item No.", "Pais", "Costo Unit Q", "Costo Unit USD",
        "Codigo Barras", "Proveedor", "No. Proveedor", "Und x Caja", "Cajas Sugeridas",
        "Und Sugeridas", "No. Tienda", "Tienda", "Inventario Actual", "Prioridad",
        "Justificacion",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(results, 2):
        priority = row_data[14]
        fill = PRIORITY_FILLS.get(priority)
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill

    widths = [35, 12, 5, 12, 12, 16, 22, 12, 10, 14, 12, 10, 30, 14, 10, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{len(results) + 1}"

    # --- RESUMEN sheet ---
    ws2 = wb.create_sheet("RESUMEN")
    ws2.cell(row=1, column=1, value="SUGERIDO WALMART").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    if delivery_date:
        ws2.cell(row=3, column=1, value=f"Entrega: {delivery_date}")
    ws2.cell(row=4, column=1, value=f"Inventario semana WM: {inv_week}")
    ws2.cell(row=5, column=1, value=f"Ventas base: semanas {', '.join(str(w) for w in sale_weeks)}")

    ws2.cell(row=7, column=1, value="RESUMEN POR PRIORIDAD").font = Font(bold=True, size=12)
    sum_headers = ["Prioridad", "Descripcion", "Lineas", "Cajas", "Unidades", "Costo Est. Q"]
    for i, h in enumerate(sum_headers, 1):
        ws2.cell(row=8, column=i, value=h).font = Font(bold=True)

    for p, desc, fill in [
        (1, "Urgente (<=5d cobertura)", P1_FILL),
        (2, "Media (<=10d)", P2_FILL),
        (3, "Baja (>10d)", P3_FILL),
    ]:
        lines = [r for r in results if r[14] == p]
        row = 8 + p
        ws2.cell(row=row, column=1, value=p)
        ws2.cell(row=row, column=2, value=desc)
        ws2.cell(row=row, column=3, value=len(lines))
        ws2.cell(row=row, column=4, value=sum(r[9] for r in lines))
        ws2.cell(row=row, column=5, value=sum(r[10] for r in lines))
        ws2.cell(row=row, column=6, value=round(sum(r[9] * r[8] * r[3] for r in lines), 2))
        for c in range(1, 7):
            ws2.cell(row=row, column=c).fill = fill

    row = 12
    ws2.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=row, column=3, value=len(results)).font = Font(bold=True)
    ws2.cell(row=row, column=4, value=total_cajas).font = Font(bold=True)
    ws2.cell(row=row, column=5, value=total_und).font = Font(bold=True)
    ws2.cell(row=row, column=6, value=round(total_q, 2)).font = Font(bold=True)

    # Product summary
    ws2.cell(row=14, column=1, value="POR PRODUCTO").font = Font(bold=True, size=12)
    prod_headers = ["Producto", "Tiendas", "Cajas", "Unidades", "Costo Est. Q"]
    for i, h in enumerate(prod_headers, 1):
        ws2.cell(row=15, column=i, value=h).font = Font(bold=True)

    prod_summary = {}
    for r in results:
        prod = r[0]
        if prod not in prod_summary:
            prod_summary[prod] = {"tiendas": 0, "cajas": 0, "und": 0, "costo": 0}
        prod_summary[prod]["tiendas"] += 1
        prod_summary[prod]["cajas"] += r[9]
        prod_summary[prod]["und"] += r[10]
        prod_summary[prod]["costo"] += r[9] * r[8] * r[3]

    for i, (prod, data) in enumerate(sorted(prod_summary.items()), 16):
        ws2.cell(row=i, column=1, value=prod)
        ws2.cell(row=i, column=2, value=data["tiendas"])
        ws2.cell(row=i, column=3, value=data["cajas"])
        ws2.cell(row=i, column=4, value=data["und"])
        ws2.cell(row=i, column=5, value=round(data["costo"], 2))

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 30
    for c in "CDEF":
        ws2.column_dimensions[c].width = 15

    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"SUGERIDO_WALMART_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    print(f"\nSaved: {filepath}")

    conn.close()
    return filepath, results, stats


if __name__ == "__main__":
    generate_sugerido(delivery_date="Viernes 7 Mar 2026")
