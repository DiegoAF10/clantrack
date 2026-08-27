#!/usr/bin/env python3
"""Create WM-INPUT.xlsx template for weekly data input."""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "templates", "WM-INPUT.xlsx")

# Styles
PASTE_HERE = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type='solid')  # Light yellow
HEADER_FILL = PatternFill(start_color="2D4A7A", end_color="2D4A7A", fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
INSTRUCTION_FONT = Font(name='Calibri', size=11, bold=True, color="1B2A4A")
NOTE_FONT = Font(name='Calibri', size=9, italic=True, color="888888")
CENTER = Alignment(horizontal='center', vertical='center')


def create_rl_sheet(wb):
    """Sheet 1: Retail Link semanal."""
    ws = wb.active
    ws.title = "RL Semanal"
    ws.sheet_properties.tabColor = "2D4A7A"

    # Instructions
    ws.merge_cells('A1:J1')
    ws['A1'] = "PEGAR DATOS DE RETAIL LINK ACÁ ABAJO (sin encabezados)"
    ws['A1'].font = INSTRUCTION_FONT

    # Headers matching FormatoRepSemanal_RL format
    headers = [
        'Semana Walmart', 'Diario', 'Item_Nbr', 'Producto', 'País',
        'Costo Q', 'Costo USD', 'Precio Q', 'Precio USD', 'Margen %',
        'Cod Barras', 'Fecha Exp', 'Store_Nbr', 'Tienda',
        'Venta Q', 'Venta USD', 'Venta Und',
        'Venta Comp Q', 'Venta Comp USD', 'Venta Comp Und',
        'Venta Costo Q', 'Venta Costo USD',
        'Max Estante', 'Tasa Venta', 'Und/Tienda/Sem',
        '$/Tienda/Sem Q', '$/Tienda/Sem USD',
        'Inv Actual', 'Tiendas Válidas',
        'Instock Prom %', 'Instock Act %',
        'Pedido Tránsito', 'Pedido Actual', 'Inv Liquidación',
        'Pronóstico 52s', 'Cód Faltante', 'Cant Faltante', 'Fecha Faltante',
        'Fecha TC', 'Tipo Cambio', 'Tipo_Registro',
    ]

    widths = [14, 12, 12, 24, 5, 10, 10, 10, 10, 10,
              14, 12, 10, 24, 10, 10, 10, 10, 10, 10,
              12, 12, 10, 10, 12, 12, 12, 10, 10, 10,
              10, 10, 10, 10, 10, 12, 10, 12, 12, 10, 14]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Yellow "paste here" zone (rows 3-1003)
    for row in range(3, 20):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = PASTE_HERE

    ws.cell(row=3, column=1, value="← Pegá los datos acá").font = NOTE_FONT

    ws.freeze_panes = 'A3'
    return ws


def create_sellin_sheet(wb):
    """Sheet 2: Sell-In (Odoo invoices)."""
    ws = wb.create_sheet("Sell-In")
    ws.sheet_properties.tabColor = "27AE60"

    ws.merge_cells('A1:E1')
    ws['A1'] = "PEGAR FACTURAS DE ODOO ACÁ ABAJO (5 columnas mínimo)"
    ws['A1'].font = INSTRUCTION_FONT

    # Headers: raw input columns from Odoo + calculated columns
    headers = [
        'Código Producto', 'Nombre Producto', 'Cliente', 'Fecha', 'Cantidad Facturada',
        'Fecha Esperada Tienda', 'Producto PRECIOS', 'Fecha Num',
        'Und x Caja', 'Cajas Facturadas',
        'Precio Caja WM', 'Costo Caja Clan',
        'Ingreso Bruto', 'Costo Total', 'Centralización 4%',
        'Ingreso Neto', 'Margen Bruto', '% Margen', 'Marca', 'Año-Mes'
    ]
    widths = [16, 20, 28, 14, 16, 16, 20, 12, 10, 14, 14, 14, 14, 12, 14, 14, 14, 10, 12, 10]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for row in range(3, 20):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = PASTE_HERE

    ws.cell(row=3, column=1, value="← Pegá los datos acá").font = NOTE_FONT

    ws.freeze_panes = 'A3'
    return ws


def create_dev_sheet(wb):
    """Sheet 3: Devoluciones."""
    ws = wb.create_sheet("Devoluciones")
    ws.sheet_properties.tabColor = "E74C3C"

    ws.merge_cells('A1:E1')
    ws['A1'] = "PEGAR DATOS DE DEVOLUCIONES/CLAIMS ACÁ ABAJO"
    ws['A1'].font = INSTRUCTION_FONT

    headers = [
        'Semana Walmart', 'Diario', 'Número Artículo', 'Descripción Artículo', 'País',
        'Cant. Reclamo Tienda', 'Precio Reclamo Q', 'Precio Reclamo USD',
        'Costo Reclamo Q', 'Costo Reclamo USD',
        'Dev. Cliente Total', 'Costo Dev. Total Q', 'Costo Dev. Total USD',
        'Precio Dev. Total Q', 'Precio Dev. Total USD',
        'Fecha TC', 'Tipo Cambio', 'Estado'
    ]
    widths = [14, 12, 14, 24, 5, 16, 14, 14, 14, 14, 14, 14, 14, 14, 14, 12, 10, 14]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for row in range(3, 20):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = PASTE_HERE

    ws.cell(row=3, column=1, value="← Pegá los datos acá").font = NOTE_FONT

    ws.freeze_panes = 'A3'
    return ws


def main():
    os.makedirs(os.path.dirname(TEMPLATE_PATH), exist_ok=True)

    wb = openpyxl.Workbook()
    create_rl_sheet(wb)
    create_sellin_sheet(wb)
    create_dev_sheet(wb)

    wb.save(TEMPLATE_PATH)
    print(f"Template created: {TEMPLATE_PATH}")
    print("3 sheets: RL Semanal | Sell-In | Devoluciones")
    print("Open, paste data, save. Then run: python import_walmart.py --template --report")


if __name__ == '__main__':
    main()
