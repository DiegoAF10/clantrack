# -*- coding: utf-8 -*-
"""Plan FINAL conservador (Diego 31/08 noche):
1 caja por tienda donde NO hay existencias en RL 202630 · 2 cajas máx en las mejores
tiendas históricas de cada ítem · 11 tiendas con problemas BLOQUEADAS · VN excluida.
Salida: Excel en formato Cristian + datos para el PDF + plan para el manifiesto."""
import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

dd = pd.read_csv('t_combos_dedupe.csv', dtype={'item': str, 'tienda_num': str})
final = json.load(open('datos_final.json', encoding='utf-8'))
base = json.load(open('datos.json', encoding='utf-8'))
cat = {c['item']: c for c in json.load(open('cat_export.json', encoding='utf-8'))}
top = json.load(open('top_historico.json', encoding='utf-8'))
FIS = {(f['tienda'], f['item']): int(f['fisico']) for f in base['matriz']
       if f.get('fisico') == f.get('fisico') and f.get('fisico') is not None}

NOMBRE = {'75437615': 'HB Helles Lager', '75437613': 'NC Old Rasputin', '75437612': 'NC Scrimshaw',
          '75437617': 'HB Session Lager', '75437618': 'Belhaven Best', '75437619': 'Belhaven Black',
          '75437620': 'Belhaven McCallums'}
LANZ = {'75437617': 26.1, '75437619': 19.4, '75437618': 17.8, '75437620': 16.5,
        '75437615': 15.1, '75437613': 10.6, '75437612': 2.8}
BLOQUEADAS = {'414', '4176', '24', '23', '121', '459', '4433', '419', '33', '46', '948'}
UCAJA = 20
tiendas40 = {t['num']: t for t in final['tiendas']}
exist = {(r['tienda_num'], r['item']): int(r['existencia']) for _, r in dd.iterrows()}
combos = set(exist.keys())
tops = {it: {t[0]: t[2] for t in v} for it, v in top.items()}

plan, lineas = [], []
for it in NOMBRE:
    mejores = tops.get(it, {})
    for tn, t in tiendas40.items():
        if tn in BLOQUEADAS:
            continue
        key = (tn, it)
        rl = exist.get(key, 0)
        fis = FIS.get(key)
        real = fis if fis is not None else rl
        if rl > 0 or (fis is not None and fis > 0):
            continue                      # hay existencias (o dice haber): conservador, no va
        tiene_combo = key in combos
        es_top = tn in mejores
        cajas = 2 if es_top else 1
        if es_top:
            pri = 1
            just = f'MEDICIÓN+ · mejor tienda histórica del ítem ({mejores[tn]} u vendidas en lanzamiento) · sin existencias' + ('' if tiene_combo else ' · requiere habilitar combo')
        elif tiene_combo:
            pri, just = 2, 'MEDICIÓN · sin existencias en RL sem 202630 · piso de 1 caja para lectura de demanda'
        else:
            pri, just = 3, 'MEDICIÓN · sin combo ítem-tienda en maestro — requiere habilitación · piso de 1 caja'
        plan.append(dict(tienda=tn, item=it, cajas=cajas,
                         motivo=just + (' · evaluación diaria, reposición según comportamiento' if pri < 3 else '')))
        lineas.append(dict(it=it, tn=tn, cajas=cajas, pri=pri, just=just, combo=tiene_combo))

# ---- totales
res = {}
for l in lineas:
    r = res.setdefault(l['it'], dict(cajas=0, tiendas=0, dos=0, sincombo=0))
    r['cajas'] += l['cajas']; r['tiendas'] += 1
    if l['cajas'] == 2: r['dos'] += 1
    if not l['combo']: r['sincombo'] += 1
tc = sum(l['cajas'] for l in lineas); tu = tc * UCAJA
tq = sum(l['cajas'] * UCAJA * float(cat[l['it']]['costoq']) for l in lineas)
print(f"{'ítem':<20}{'cajas':>6}{'tiendas':>8}{'2-cajas':>8}{'s/combo':>8}{'Q':>11}")
for it, nom in NOMBRE.items():
    r = res.get(it, dict(cajas=0, tiendas=0, dos=0, sincombo=0))
    q = r['cajas'] * UCAJA * float(cat[it]['costoq'])
    print(f"{nom:<20}{r['cajas']:>6}{r['tiendas']:>8}{r['dos']:>8}{r['sincombo']:>8}{q:>11,.2f}")
print(f"{'TOTAL':<20}{tc:>6}{'':>8}{'':>8}{sum(1 for l in lineas if not l['combo']):>8}{tq:>11,.2f}  ({tu:,} u)")

# ---- Excel formato Cristian
wb = Workbook()
ws = wb.active; ws.title = 'SUGERIDO_EXPORT'
HDR = ['Descripcion Articulo', 'Item No.', 'Pais', 'Costo Unit Q', 'Costo Unit USD', 'Codigo Barras',
       'Proveedor', 'No. Proveedor', 'Und x Caja', 'Cajas Sugeridas', 'Und Sugeridas', 'No. Tienda',
       'Tienda', 'Inventario Actual', 'Prioridad', 'Justificacion']
ws.append(HDR)
azul = PatternFill('solid', fgColor='0D211C')
oro = Font(color='C9A227', bold=True)
for c in ws[1]:
    c.fill = azul; c.font = oro; c.alignment = Alignment(horizontal='center')
for l in sorted(lineas, key=lambda x: (x['pri'], NOMBRE[x['it']], tiendas40[x['tn']]['nombre'])):
    c = cat[l['it']]
    ws.append([c['desc'], int(l['it']), 'GT', float(c['costoq']), round(float(c['costousd']), 6),
               int(str(c['barcode']).lstrip('0')), 'CLAN CERVECERO SA', 258977, UCAJA,
               l['cajas'], l['cajas'] * UCAJA, int(l['tn']), tiendas40[l['tn']]['nombre'], 0,
               l['pri'], l['just']])
for col, w in zip('ABCDEFGHIJKLMNOP', [22, 10, 6, 11, 13, 14, 19, 12, 10, 13, 12, 10, 26, 15, 9, 74]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'

r2 = wb.create_sheet('RESUMEN')
r2.append(['SUGERIDO WALMART — PEDIDO DE MEDICIÓN · NO MODULADOS'])
r2.append(['Generado: 2026-08-31 · Clan Cervecero'])
r2.append(['Entrega: por confirmar con Cristian'])
r2.append(['Inventario semana WM: 202630 (Retail Link corte 30/08) + conteos físicos GSP 26-29/08'])
r2.append(['Ventas base: rotación de lanzamiento 202550-202613 (justificación) · vigencias renovadas 31/08'])
r2.append([])
r2.append(['RESUMEN POR PRIORIDAD'])
r2.append(['Prioridad', 'Descripcion', 'Lineas', 'Cajas', 'Unidades', 'Costo Est. Q'])
for pri, descr in [(1, 'Mejor tienda histórica del ítem (2 cajas)'),
                   (2, 'Piso de medición — sin existencias (1 caja)'),
                   (3, 'Piso — requiere habilitar combo ítem-tienda (1 caja)')]:
    ls = [l for l in lineas if l['pri'] == pri]
    cj = sum(l['cajas'] for l in ls)
    q = sum(l['cajas'] * UCAJA * float(cat[l['it']]['costoq']) for l in ls)
    r2.append([pri, descr, len(ls), cj, cj * UCAJA, round(q)])
r2.append(['TOTAL', '', len(lineas), tc, tu, round(tq)])
r2.append([])
r2.append(['NOTA: 11 tiendas con incidencias abiertas (ajustes de inventario, planograma, reclamos) quedan'])
r2.append(['fuera de este pedido hasta resolverse — detalle en el documento "Hallazgos y apoyos solicitados".'])
r2.append(['Regla del pedido: 1 caja por tienda sin existencias · máximo 2 en las mejores tiendas históricas ·'])
r2.append(['evaluación diaria del comportamiento y reposición según demanda observada.'])
r2['A1'].font = Font(bold=True, size=13)
for c in r2[8]:
    c.font = Font(bold=True)
xlout = r'C:\Users\Diego\projects\clan\coo\walmart\expediente-walmart-2026\pedido-refuerzo\SUGERIDO_WALMART_2026-08-31.xlsx'
wb.save(xlout)
print('\nExcel guardado:', xlout.split('\\')[-1], '·', len(lineas), 'líneas')

# ---- actualizar manifiesto data + guardar para PDF
final['planes']['recomendado'] = plan
final['kpis'][0] = dict(label='Pedido de medición (final)', valor=f'{tc} cajas',
                        sub=f'{tu:,} u · 1 caja donde no hay + 2 en mejores históricas', tono='ok')
final['kpis'][1]['valor'] = f'Q {tq:,.0f}'
final['medicion_meta'].update(dict(cajas=tc, unidades=tu, q_costo=round(tq, 2),
    regla='FINAL: 1 caja/tienda sin existencias · 2 máx en mejores históricas · 11 tiendas bloqueadas · eval diaria'))
json.dump(final, open('datos_final.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
json.dump(dict(lineas=lineas, res=res, tot=dict(c=tc, u=tu, q=round(tq, 2)), lanz=LANZ),
          open('plan_final_pdf.json', 'w', encoding='utf-8'), ensure_ascii=False)
print('datos_final actualizado (manifiesto) · plan_final_pdf.json listo')
