#!/usr/bin/env python3
"""
Generate REPORTE-WM.xlsx from walmart.db
Clean, focused, professional report. 5 sheets, values only, opens instantly.
"""

import sqlite3
import os
import sys
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "walmart.db")
REPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")

# ═══════════════════════════════════════════
# STYLES
# ═══════════════════════════════════════════

# Colors
DARK_BLUE = "1B2A4A"
MEDIUM_BLUE = "2D4A7A"
LIGHT_BLUE = "E8EDF5"
ACCENT_GREEN = "27AE60"
ACCENT_RED = "E74C3C"
ACCENT_ORANGE = "F39C12"
HEADER_BG = "2D4A7A"
SUBHEADER_BG = "4A6FA5"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
BORDER_COLOR = "D0D5DD"

# Fonts
title_font = Font(name='Calibri', size=16, bold=True, color=WHITE)
section_font = Font(name='Calibri', size=13, bold=True, color=WHITE)
header_font = Font(name='Calibri', size=10, bold=True, color=WHITE)
kpi_label_font = Font(name='Calibri', size=10, color="666666")
kpi_value_font = Font(name='Calibri', size=14, bold=True, color=DARK_BLUE)
kpi_small_font = Font(name='Calibri', size=9, color="999999")
data_font = Font(name='Calibri', size=10)
data_bold = Font(name='Calibri', size=10, bold=True)
pct_font_green = Font(name='Calibri', size=10, color=ACCENT_GREEN)
pct_font_red = Font(name='Calibri', size=10, color=ACCENT_RED)

# Fills
title_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
subheader_fill = PatternFill(start_color=SUBHEADER_BG, end_color=SUBHEADER_BG, fill_type='solid')
light_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type='solid')
red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type='solid')
orange_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type='solid')

# Alignment
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center')
right = Alignment(horizontal='right', vertical='center')
wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Border
thin_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR)
)


def style_title_row(ws, row, text, max_col=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 36


def style_section_header(ws, row, text, max_col=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 28


def style_headers(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        if widths and i <= len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.row_dimensions[row].height = 24


def write_data_row(ws, row, values, formats=None, bold_first=False):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = data_bold if (bold_first and i == 1) else data_font
        cell.border = thin_border
        cell.alignment = right if isinstance(v, (int, float)) else left

        if formats and i <= len(formats) and v is not None:
            fmt = formats[i-1]
            if fmt == 'q':
                cell.number_format = '#,##0'
            elif fmt == 'q2':
                cell.number_format = '#,##0.00'
            elif fmt == 'pct':
                cell.number_format = '0.0%'
            elif fmt == 'int':
                cell.number_format = '#,##0'

        # Alternating row color
        if row % 2 == 0:
            cell.fill = alt_fill


def write_kpi_card(ws, row, col, label, value, subtitle=None, fmt='q2'):
    ws.cell(row=row, column=col, value=label).font = kpi_label_font
    cell = ws.cell(row=row+1, column=col, value=value)
    cell.font = kpi_value_font
    if fmt == 'q2':
        cell.number_format = 'Q#,##0.00'
    elif fmt == 'int':
        cell.number_format = '#,##0'
    elif fmt == 'pct':
        cell.number_format = '0.0%'
    elif fmt == 'gmroi':
        cell.number_format = '0.00x'
    if subtitle:
        ws.cell(row=row+2, column=col, value=subtitle).font = kpi_small_font


# ═══════════════════════════════════════════
# SHEET GENERATORS
# ═══════════════════════════════════════════

def build_dashboard(wb, conn):
    """Sheet 1: Resumen Ejecutivo."""
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_properties.tabColor = DARK_BLUE

    # Get key data
    max_week = conn.execute("SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='VENTA'").fetchone()[0]
    min_week = conn.execute("SELECT MIN(semana_wm) FROM retail_link WHERE tipo_registro='VENTA'").fetchone()[0]

    # All-time sales
    all_sales = conn.execute("""
        SELECT SUM(venta_und), SUM(venta_q), SUM(venta_costo_q)
        FROM retail_link WHERE tipo_registro='VENTA'
    """).fetchone()

    # Last 30 days sales (approx last 4 weeks)
    recent_weeks = conn.execute("""
        SELECT DISTINCT semana_wm FROM retail_link
        WHERE tipo_registro='VENTA' ORDER BY semana_wm DESC LIMIT 4
    """).fetchall()
    recent_week_list = [w[0] for w in recent_weeks]
    placeholders = ','.join(['?'] * len(recent_week_list))

    recent_sales = conn.execute(f"""
        SELECT SUM(venta_und), SUM(venta_q), SUM(venta_costo_q),
               COUNT(DISTINCT store_nbr)
        FROM retail_link
        WHERE tipo_registro='VENTA' AND semana_wm IN ({placeholders})
    """, recent_week_list).fetchone()

    # Previous 4 weeks for comparison
    prev_weeks = conn.execute("""
        SELECT DISTINCT semana_wm FROM retail_link
        WHERE tipo_registro='VENTA' ORDER BY semana_wm DESC LIMIT 8
    """).fetchall()
    prev_week_list = [w[0] for w in prev_weeks[4:8]]
    if prev_week_list:
        ph2 = ','.join(['?'] * len(prev_week_list))
        prev_sales = conn.execute(f"""
            SELECT SUM(venta_und), SUM(venta_q)
            FROM retail_link
            WHERE tipo_registro='VENTA' AND semana_wm IN ({ph2})
        """, prev_week_list).fetchone()
    else:
        prev_sales = (0, 0)

    # Current inventory
    inv = conn.execute("""
        SELECT SUM(inv_actual), COUNT(DISTINCT item_nbr), COUNT(DISTINCT store_nbr)
        FROM retail_link
        WHERE tipo_registro='INVENTARIO'
        AND semana_wm = (SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO')
    """).fetchone()

    # Sell-In totals
    si = conn.execute("""
        SELECT SUM(ingreso_bruto), SUM(costo_total), SUM(centralizacion),
               SUM(ingreso_neto), SUM(margen_bruto), SUM(cantidad_facturada)
    FROM sell_in
    """).fetchone()

    # Devoluciones totals
    dev = conn.execute("""
        SELECT SUM(cant_reclamo_tienda + devolucion_cliente_und),
               SUM(precio_reclamo_q + precio_devolucion_q)
        FROM devoluciones
    """).fetchone()

    # Unique products and stores
    n_prod = conn.execute("SELECT COUNT(DISTINCT item_nbr) FROM retail_link WHERE tipo_registro='VENTA' AND venta_und > 0").fetchone()[0]
    n_stores = conn.execute("SELECT COUNT(DISTINCT store_nbr) FROM retail_link WHERE tipo_registro='VENTA' AND venta_und > 0").fetchone()[0]

    # === BUILD SHEET ===
    max_col = 10

    style_title_row(ws, 1, "CLAN CERVECERO x WALMART — Resumen Ejecutivo", max_col)

    # Metadata row
    ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = kpi_small_font
    ws.cell(row=2, column=4, value=f"Período: {min_week} → {max_week}").font = kpi_small_font
    ws.cell(row=2, column=7, value=f"{n_prod} productos | {n_stores} tiendas").font = kpi_small_font

    # KPI Cards Row 1: SELL-OUT (Walmart sells to consumer)
    row = 4
    style_section_header(ws, row, "SELL-OUT — Lo que Walmart vende al consumidor", max_col)

    row = 5
    write_kpi_card(ws, row, 1, "Ventas Totales (Und)", all_sales[0] or 0, f"Último mes: {recent_sales[0] or 0:,}", 'int')
    write_kpi_card(ws, row, 3, "Ventas Totales (Q)", all_sales[1] or 0, f"Último mes: Q{recent_sales[1] or 0:,.0f}", 'q2')

    margen_wm = (all_sales[1] - all_sales[2]) if all_sales[1] and all_sales[2] else 0
    margen_wm_pct = margen_wm / all_sales[1] if all_sales[1] else 0
    write_kpi_card(ws, row, 5, "Margen WM (%)", margen_wm_pct, f"Q{margen_wm:,.0f} total", 'pct')

    # Variation vs previous period
    var_und = ((recent_sales[0] - prev_sales[0]) / prev_sales[0]) if prev_sales[0] else 0
    write_kpi_card(ws, row, 7, "Tendencia (vs mes ant.)", var_und, "Unidades vendidas", 'pct')

    inv_total = inv[0] or 0
    write_kpi_card(ws, row, 9, "Inventario Actual (Und)", inv_total, f"{inv[1] or 0} SKUs en {inv[2] or 0} tiendas", 'int')

    # KPI Cards Row 2: SELL-IN (Clan sells to Walmart)
    row = 9
    style_section_header(ws, row, "SELL-IN — Lo que Clan Cervecero factura a Walmart", max_col)

    row = 10
    write_kpi_card(ws, row, 1, "Ingreso Bruto", si[0] or 0, "Facturado antes de centralización", 'q2')
    write_kpi_card(ws, row, 3, "Centralización 4%", si[2] or 0, "Descuento WM", 'q2')
    write_kpi_card(ws, row, 5, "Ingreso Neto", si[3] or 0, "Lo que realmente cobramos", 'q2')
    write_kpi_card(ws, row, 7, "Margen Clan", si[4] or 0, f"{si[4]/(si[0] or 1)*100:.1f}% margen", 'q2')

    # GMROI Clan
    inv_costo = conn.execute("""
        SELECT SUM(rl.inv_actual * p.costo_clan_und)
        FROM retail_link rl
        JOIN productos p ON rl.producto = p.producto
        WHERE rl.tipo_registro='INVENTARIO'
        AND rl.semana_wm = (SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO')
    """).fetchone()[0] or 1
    gmroi = (si[4] or 0) / inv_costo
    write_kpi_card(ws, row, 9, "GMROI Clan", gmroi, "Margen / Inv. valorizado", 'gmroi')

    # Devoluciones alert
    row = 14
    style_section_header(ws, row, "DEVOLUCIONES & ALERTAS", max_col)

    row = 15
    dev_und = dev[0] or 0
    dev_q = dev[1] or 0
    dev_pct = dev_und / (si[5] or 1)
    write_kpi_card(ws, row, 1, "Devoluciones (Und)", dev_und, f"Q{dev_q:,.0f} total", 'int')
    write_kpi_card(ws, row, 3, "% Devuelto", dev_pct, "Meta: < 5%", 'pct')

    # OOS Alert
    oos = conn.execute("""
        SELECT COUNT(DISTINCT item_nbr || '-' || store_nbr)
        FROM retail_link
        WHERE tipo_registro='INVENTARIO'
        AND semana_wm = (SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO')
        AND inv_actual = 0
    """).fetchone()[0]
    total_combos = conn.execute("""
        SELECT COUNT(DISTINCT item_nbr || '-' || store_nbr)
        FROM retail_link
        WHERE tipo_registro='INVENTARIO'
        AND semana_wm = (SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO')
    """).fetchone()[0] or 1
    oos_rate = oos / total_combos
    write_kpi_card(ws, row, 5, "OOS Rate", oos_rate, f"{oos} de {total_combos} combinaciones sin stock", 'pct')

    # Top 5 Products (quick reference)
    row = 19
    style_section_header(ws, row, "TOP 5 PRODUCTOS (Últimas 4 semanas)", max_col)
    row = 20
    headers = ['#', 'Producto', 'Marca', 'Venta Und', 'Venta Q', '% Total', 'Inv. Actual', 'Cobertura']
    style_headers(ws, row, headers, [4, 28, 12, 12, 14, 10, 12, 12])

    top_prods = conn.execute(f"""
        SELECT rl.producto, p.marca,
               SUM(rl.venta_und), SUM(rl.venta_q)
        FROM retail_link rl
        LEFT JOIN productos p ON rl.producto = p.producto
        WHERE rl.tipo_registro='VENTA' AND rl.semana_wm IN ({placeholders})
        GROUP BY rl.producto
        ORDER BY SUM(rl.venta_und) DESC
        LIMIT 5
    """, recent_week_list).fetchall()

    total_recent = sum(r[2] for r in top_prods)
    for i, (prod, marca, und, q) in enumerate(top_prods, 1):
        # Get current inventory for this product
        inv_prod = conn.execute("""
            SELECT SUM(inv_actual) FROM retail_link
            WHERE tipo_registro='INVENTARIO' AND producto=?
            AND semana_wm = (SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO')
        """, (prod,)).fetchone()[0] or 0
        daily_rate = und / 28 if und else 0
        coverage = inv_prod / daily_rate if daily_rate > 0 else 999
        row += 1
        pct = und / total_recent if total_recent else 0
        write_data_row(ws, row, [i, prod, marca or '', und, q, pct, inv_prod, f"{coverage:.0f}d"],
                      ['int', None, None, 'int', 'q2', 'pct', 'int', None], bold_first=True)

    # Top 5 Stores
    row += 2
    style_section_header(ws, row, "TOP 5 TIENDAS (Últimas 4 semanas)", max_col)
    row += 1
    headers = ['#', 'Tienda', 'Formato', 'Venta Und', 'Venta Q', '% Total', 'SKUs', 'Región']
    style_headers(ws, row, headers, [4, 28, 10, 12, 14, 10, 8, 12])

    top_stores = conn.execute(f"""
        SELECT rl.tienda, t.formato, t.region,
               SUM(rl.venta_und), SUM(rl.venta_q),
               COUNT(DISTINCT rl.item_nbr)
        FROM retail_link rl
        LEFT JOIN tiendas t ON rl.store_nbr = t.no_tienda
        WHERE rl.tipo_registro='VENTA' AND rl.semana_wm IN ({placeholders})
        GROUP BY rl.tienda
        ORDER BY SUM(rl.venta_und) DESC
        LIMIT 5
    """, recent_week_list).fetchall()

    total_store = sum(r[3] for r in top_stores)
    for i, (tienda, fmt, region, und, q, skus) in enumerate(top_stores, 1):
        row += 1
        pct = und / total_store if total_store else 0
        write_data_row(ws, row, [i, tienda, fmt or '', und, q, pct, skus, region or ''],
                      ['int', None, None, 'int', 'q2', 'pct', 'int', None], bold_first=True)

    print("  [1/5] Resumen Ejecutivo")
    return ws


def build_semanal(wb, conn):
    """Sheet 2: Reporte Semanal."""
    ws = wb.create_sheet("Semanal")
    ws.sheet_properties.tabColor = MEDIUM_BLUE
    max_col = 14

    # Get latest 2 weeks for comparison
    weeks = conn.execute("""
        SELECT DISTINCT semana_wm FROM retail_link
        WHERE tipo_registro='VENTA'
        ORDER BY semana_wm DESC LIMIT 2
    """).fetchall()
    curr_week = weeks[0][0] if weeks else None
    prev_week = weeks[1][0] if len(weeks) > 1 else None

    style_title_row(ws, 1, f"REPORTE SEMANAL — Semana {curr_week}", max_col)

    # Summary KPIs
    for wk, label, start_row in [(curr_week, "SEMANA ACTUAL", 3), (prev_week, "SEMANA ANTERIOR", 3)]:
        if not wk:
            continue
        data = conn.execute("""
            SELECT SUM(venta_und), SUM(venta_q), SUM(venta_costo_q),
                   COUNT(DISTINCT store_nbr), COUNT(DISTINCT item_nbr)
            FROM retail_link WHERE tipo_registro='VENTA' AND semana_wm=?
        """, (wk,)).fetchone()
        if label == "SEMANA ACTUAL":
            curr_data = data
        else:
            prev_data = data

    row = 3
    style_section_header(ws, row, f"Semana {curr_week} vs {prev_week}", max_col)

    row = 4
    cd = curr_data if 'curr_data' in dir() else (0,0,0,0,0)
    pd_ = prev_data if 'prev_data' in dir() else (0,0,0,0,0)

    headers = ['Métrica', 'Semana Actual', 'Semana Anterior', 'Variación', 'Var %']
    style_headers(ws, row, headers, [20, 16, 16, 14, 10])

    metrics = [
        ('Ventas (Und)', cd[0] or 0, pd_[0] or 0, 'int'),
        ('Ventas (Q)', cd[1] or 0, pd_[1] or 0, 'q2'),
        ('Ticket Prom. (Q)', (cd[1] or 0)/(cd[0] or 1), (pd_[1] or 0)/(pd_[0] or 1), 'q2'),
        ('Tiendas Activas', cd[3] or 0, pd_[3] or 0, 'int'),
        ('SKUs Vendidos', cd[4] or 0, pd_[4] or 0, 'int'),
    ]

    for name, curr, prev, fmt in metrics:
        row += 1
        diff = curr - prev
        var_pct = diff / prev if prev else 0
        write_data_row(ws, row, [name, curr, prev, diff, var_pct],
                      [None, fmt, fmt, fmt, 'pct'], bold_first=True)

    # Sales by Product this week
    row += 2
    style_section_header(ws, row, f"VENTAS POR PRODUCTO — Semana {curr_week}", max_col)
    row += 1
    headers = ['Producto', 'Marca', 'Venta Und', 'Venta Q', 'Sem. Anterior', 'Var %', 'Inv. Actual']
    style_headers(ws, row, headers, [28, 12, 12, 14, 14, 10, 12])

    prods = conn.execute("""
        SELECT rl.producto, p.marca,
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_und ELSE 0 END) as curr_und,
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_q ELSE 0 END) as curr_q,
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_und ELSE 0 END) as prev_und
        FROM retail_link rl
        LEFT JOIN productos p ON rl.producto = p.producto
        WHERE rl.tipo_registro='VENTA' AND rl.semana_wm IN (?,?)
        GROUP BY rl.producto
        ORDER BY curr_und DESC
    """, (curr_week, curr_week, prev_week, curr_week, prev_week)).fetchall()

    for prod, marca, c_und, c_q, p_und in prods:
        if c_und == 0 and p_und == 0:
            continue
        row += 1
        var = (c_und - p_und) / p_und if p_und else 0
        inv_p = conn.execute("""
            SELECT SUM(inv_actual) FROM retail_link
            WHERE tipo_registro='INVENTARIO' AND producto=?
            AND semana_wm=(SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO')
        """, (prod,)).fetchone()[0] or 0
        write_data_row(ws, row, [prod, marca or '', c_und, c_q, p_und, var, inv_p],
                      [None, None, 'int', 'q2', 'int', 'pct', 'int'], bold_first=True)

    # Sales by Store this week (top 15)
    row += 2
    style_section_header(ws, row, f"TOP 15 TIENDAS — Semana {curr_week}", max_col)
    row += 1
    headers = ['#', 'Tienda', 'Formato', 'Venta Und', 'Venta Q', 'Sem. Anterior', 'Var %']
    style_headers(ws, row, headers, [4, 28, 10, 12, 14, 14, 10])

    stores = conn.execute("""
        SELECT rl.tienda, t.formato,
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_und ELSE 0 END) as curr_und,
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_q ELSE 0 END) as curr_q,
               SUM(CASE WHEN rl.semana_wm=? THEN rl.venta_und ELSE 0 END) as prev_und
        FROM retail_link rl
        LEFT JOIN tiendas t ON rl.store_nbr = t.no_tienda
        WHERE rl.tipo_registro='VENTA' AND rl.semana_wm IN (?,?)
        GROUP BY rl.tienda
        ORDER BY curr_und DESC
        LIMIT 15
    """, (curr_week, curr_week, prev_week, curr_week, prev_week)).fetchall()

    for i, (tienda, fmt, c_und, c_q, p_und) in enumerate(stores, 1):
        row += 1
        var = (c_und - p_und) / p_und if p_und else 0
        write_data_row(ws, row, [i, tienda, fmt or '', c_und, c_q, p_und, var],
                      ['int', None, None, 'int', 'q2', 'int', 'pct'], bold_first=True)

    print("  [2/5] Reporte Semanal")
    return ws


def build_sugerido(wb, conn):
    """Sheet 3: Pedido Sugerido."""
    ws = wb.create_sheet("Sugerido")
    ws.sheet_properties.tabColor = ACCENT_GREEN
    max_col = 16

    style_title_row(ws, 1, "PEDIDO SUGERIDO — Clan Cervecero x Walmart", max_col)

    # Configuration
    DIAS_ROTACION = 30  # Period to analyze sales
    DIAS_INV_OBJETIVO = 15  # Target days of inventory

    # Get latest inventory week
    inv_week = conn.execute("""
        SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO'
    """).fetchone()[0]

    # Get sales period (last N days of data)
    sale_weeks = conn.execute("""
        SELECT DISTINCT semana_wm FROM retail_link
        WHERE tipo_registro='VENTA' ORDER BY semana_wm DESC LIMIT 5
    """).fetchall()
    sale_week_list = [w[0] for w in sale_weeks]
    ph = ','.join(['?'] * len(sale_week_list))

    ws.cell(row=2, column=1, value=f"Inventario: Semana {inv_week}").font = kpi_small_font
    ws.cell(row=2, column=4, value=f"Ventas: Últimas {len(sale_week_list)} semanas").font = kpi_small_font
    ws.cell(row=2, column=7, value=f"Objetivo: {DIAS_INV_OBJETIVO} días de cobertura").font = kpi_small_font

    # Calculate suggested orders
    suggestions = conn.execute(f"""
        WITH ventas AS (
            SELECT item_nbr, producto, store_nbr, tienda,
                   SUM(venta_und) as venta_total,
                   COUNT(DISTINCT semana_wm) as semanas_venta
            FROM retail_link
            WHERE tipo_registro='VENTA' AND semana_wm IN ({ph})
            GROUP BY item_nbr, producto, store_nbr, tienda
        ),
        inventario AS (
            SELECT item_nbr, producto, store_nbr, tienda, inv_actual
            FROM retail_link
            WHERE tipo_registro='INVENTARIO' AND semana_wm=?
        )
        SELECT
            COALESCE(v.producto, i.producto) as producto,
            p.marca,
            p.item_wm,
            p.und_x_caja,
            p.costo_clan_und,
            p.codigo_barras,
            COALESCE(v.store_nbr, i.store_nbr) as store_nbr,
            COALESCE(v.tienda, i.tienda) as tienda,
            COALESCE(v.venta_total, 0) as venta_total,
            COALESCE(v.semanas_venta, 0) as semanas_venta,
            COALESCE(i.inv_actual, 0) as inv_actual,
            CASE
                WHEN COALESCE(v.venta_total, 0) = 0 THEN 999
                ELSE COALESCE(i.inv_actual, 0) * 1.0 / (COALESCE(v.venta_total, 0) * 1.0 / ({DIAS_ROTACION}))
            END as dias_cobertura
        FROM ventas v
        FULL OUTER JOIN inventario i ON v.item_nbr = i.item_nbr AND v.store_nbr = i.store_nbr
        LEFT JOIN productos p ON COALESCE(v.producto, i.producto) = p.producto
        WHERE p.estado = 'Activo'
        ORDER BY dias_cobertura ASC
    """, (*sale_week_list, inv_week)).fetchall()

    # Summary by priority
    urgent = []     # P1: 0 days coverage, has sales history
    medium = []     # P2: < target coverage
    low = []        # P3: approaching target

    for row_data in suggestions:
        prod, marca, item_wm, und_caja, costo, barcode, store, tienda, venta, sem, inv, cob = row_data
        if cob >= DIAS_INV_OBJETIVO * 1.5:
            continue  # Sufficient coverage
        if und_caja is None or und_caja == 0:
            und_caja = 1

        # Calculate suggested boxes
        daily_rate = venta / DIAS_ROTACION if venta > 0 else 0.5  # Assume 0.5/day for new products
        needed = max(0, (DIAS_INV_OBJETIVO * daily_rate) - inv)
        cajas = max(1, int((needed + und_caja - 1) // und_caja)) if needed > 0 else 0

        if cajas == 0:
            continue

        entry = (prod, marca, item_wm, und_caja, costo, barcode, store, tienda, venta, inv, cob, cajas, daily_rate)

        if cob <= 0 or inv == 0:
            if venta > 0:
                urgent.append(entry)
        elif cob < DIAS_INV_OBJETIVO * 0.7:
            medium.append(entry)
        else:
            low.append(entry)

    # Resumen
    row = 4
    style_section_header(ws, row, "RESUMEN", max_col)
    row = 5
    total_cajas = sum(e[11] for e in urgent + medium + low)
    ws.cell(row=row, column=1, value=f"P1 Urgente: {len(urgent)} items ({sum(e[11] for e in urgent)} cajas)").font = data_bold
    ws.cell(row=row, column=4, value=f"P2 Media: {len(medium)} items ({sum(e[11] for e in medium)} cajas)").font = data_bold
    ws.cell(row=row, column=7, value=f"P3 Baja: {len(low)} items ({sum(e[11] for e in low)} cajas)").font = data_bold
    ws.cell(row=row, column=10, value=f"TOTAL: {total_cajas} cajas").font = Font(name='Calibri', size=12, bold=True, color=ACCENT_RED)

    # Detail table
    row = 7
    headers = ['Pri', 'Producto', 'Item WM', 'Tienda', 'Store#', 'Und/Caja',
               'Cajas Sug.', 'Und Sug.', 'Inv. Actual', 'Venta 30d', 'Cobertura', 'Justificación']
    style_headers(ws, row, headers, [5, 24, 12, 24, 8, 10, 10, 10, 10, 10, 10, 28])

    all_entries = [(1, urgent), (2, medium), (3, low)]
    for pri, entries in all_entries:
        for entry in sorted(entries, key=lambda x: x[10]):  # Sort by coverage
            prod, marca, item_wm, und_caja, costo, barcode, store, tienda, venta, inv, cob, cajas, daily = entry
            row += 1
            und_sug = cajas * und_caja

            if cob <= 0 and inv == 0 and venta > 0:
                justif = f"Sin stock, venta {venta} und/30d"
            elif cob <= 0 and venta == 0:
                justif = "Producto nuevo"
            else:
                justif = f"Cobertura {cob:.1f}d → objetivo {DIAS_INV_OBJETIVO}d"

            pri_label = {1: 'P1', 2: 'P2', 3: 'P3'}[pri]
            write_data_row(ws, row,
                [pri_label, prod, item_wm or '', tienda, store, und_caja,
                 cajas, und_sug, inv, venta, f"{cob:.1f}d" if cob < 999 else "N/A", justif],
                [None, None, None, None, 'int', 'int', 'int', 'int', 'int', 'int', None, None])

            # Color code priority
            if pri == 1:
                ws.cell(row=row, column=1).fill = red_fill
            elif pri == 2:
                ws.cell(row=row, column=1).fill = orange_fill
            else:
                ws.cell(row=row, column=1).fill = green_fill

    print(f"  [3/5] Sugerido ({len(urgent)+len(medium)+len(low)} items, {total_cajas} cajas)")
    return ws


def build_rankings(wb, conn):
    """Sheet 4: Rankings / Scorecard."""
    ws = wb.create_sheet("Rankings")
    ws.sheet_properties.tabColor = ACCENT_ORANGE
    max_col = 12

    style_title_row(ws, 1, "RANKINGS — Clan Cervecero x Walmart", max_col)

    # Top 15 Products (all time)
    row = 3
    style_section_header(ws, row, "TOP 15 PRODUCTOS (Histórico Completo)", max_col)
    row += 1
    headers = ['#', 'Producto', 'Marca', 'Venta Und', 'Venta Q', '% Total', 'Margen WM Q', 'Tiendas', 'Semanas']
    style_headers(ws, row, headers, [4, 28, 12, 12, 14, 10, 14, 8, 8])

    total_q = conn.execute("SELECT SUM(venta_q) FROM retail_link WHERE tipo_registro='VENTA'").fetchone()[0] or 1

    prods = conn.execute("""
        SELECT rl.producto, p.marca,
               SUM(rl.venta_und), SUM(rl.venta_q), SUM(rl.venta_costo_q),
               COUNT(DISTINCT rl.store_nbr), COUNT(DISTINCT rl.semana_wm)
        FROM retail_link rl
        LEFT JOIN productos p ON rl.producto = p.producto
        WHERE rl.tipo_registro='VENTA' AND rl.venta_und > 0
        GROUP BY rl.producto
        ORDER BY SUM(rl.venta_q) DESC
        LIMIT 15
    """).fetchall()

    for i, (prod, marca, und, q, costo_q, stores, weeks) in enumerate(prods, 1):
        row += 1
        margen = q - (costo_q or 0)
        write_data_row(ws, row, [i, prod, marca or '', und, q, q/total_q, margen, stores, weeks],
                      ['int', None, None, 'int', 'q2', 'pct', 'q2', 'int', 'int'], bold_first=True)

    # Top 15 Stores (all time)
    row += 2
    style_section_header(ws, row, "TOP 15 TIENDAS (Histórico Completo)", max_col)
    row += 1
    headers = ['#', 'Tienda', 'Formato', 'Región', 'Venta Und', 'Venta Q', '% Total', 'SKUs', 'Semanas']
    style_headers(ws, row, headers, [4, 28, 10, 12, 12, 14, 10, 8, 8])

    stores = conn.execute("""
        SELECT rl.tienda, t.formato, t.region,
               SUM(rl.venta_und), SUM(rl.venta_q),
               COUNT(DISTINCT rl.item_nbr), COUNT(DISTINCT rl.semana_wm)
        FROM retail_link rl
        LEFT JOIN tiendas t ON rl.store_nbr = t.no_tienda
        WHERE rl.tipo_registro='VENTA' AND rl.venta_und > 0
        GROUP BY rl.tienda
        ORDER BY SUM(rl.venta_q) DESC
        LIMIT 15
    """).fetchall()

    for i, (tienda, fmt, region, und, q, skus, weeks) in enumerate(stores, 1):
        row += 1
        write_data_row(ws, row, [i, tienda, fmt or '', region or '', und, q, q/total_q, skus, weeks],
                      ['int', None, None, None, 'int', 'q2', 'pct', 'int', 'int'], bold_first=True)

    # Brand Performance
    row += 2
    style_section_header(ws, row, "RENDIMIENTO POR MARCA", max_col)
    row += 1
    headers = ['Marca', 'Productos', 'Venta Und', 'Venta Q', '% Total', 'Ticket Prom.', 'Tiendas']
    style_headers(ws, row, headers, [16, 10, 12, 14, 10, 12, 10])

    brands = conn.execute("""
        SELECT p.marca,
               COUNT(DISTINCT rl.item_nbr), SUM(rl.venta_und), SUM(rl.venta_q),
               AVG(CASE WHEN rl.venta_und > 0 THEN rl.venta_q / rl.venta_und ELSE NULL END),
               COUNT(DISTINCT rl.store_nbr)
        FROM retail_link rl
        LEFT JOIN productos p ON rl.producto = p.producto
        WHERE rl.tipo_registro='VENTA' AND p.marca IS NOT NULL
        GROUP BY p.marca
        ORDER BY SUM(rl.venta_q) DESC
    """).fetchall()

    for marca, n_prod, und, q, ticket, stores in brands:
        row += 1
        write_data_row(ws, row, [marca or 'Sin marca', n_prod, und, q, q/total_q, ticket or 0, stores],
                      [None, 'int', 'int', 'q2', 'pct', 'q2', 'int'], bold_first=True)

    # Monthly Trend (last 12 months)
    row += 2
    style_section_header(ws, row, "TENDENCIA MENSUAL (Últimos 12 meses)", max_col)
    row += 1
    headers = ['Año-Mes', 'Semanas', 'Venta Und', 'Venta Q', 'Prom/Semana Und', 'Prom/Semana Q', 'Var %']
    style_headers(ws, row, headers, [12, 10, 12, 14, 14, 14, 10])

    monthly = conn.execute("""
        SELECT anio, mes, COUNT(DISTINCT semana_wm),
               SUM(venta_und), SUM(venta_q)
        FROM retail_link
        WHERE tipo_registro='VENTA' AND anio IS NOT NULL
        GROUP BY anio, mes
        ORDER BY anio DESC, mes DESC
        LIMIT 12
    """).fetchall()

    prev_q = None
    for y, m, sem, und, q in reversed(monthly):
        row += 1
        avg_und = und / sem if sem else 0
        avg_q = q / sem if sem else 0
        var = (q - prev_q) / prev_q if prev_q else 0
        write_data_row(ws, row, [f"{y}-{m:02d}", sem, und, q, avg_und, avg_q, var],
                      [None, 'int', 'int', 'q2', 'q2', 'q2', 'pct'], bold_first=True)
        prev_q = q

    print("  [4/5] Rankings")
    return ws


def build_rentabilidad(wb, conn):
    """Sheet 5: Rentabilidad."""
    ws = wb.create_sheet("Rentabilidad")
    ws.sheet_properties.tabColor = "8E24AA"
    max_col = 12

    style_title_row(ws, 1, "RENTABILIDAD — Clan Cervecero x Walmart", max_col)

    # P&L Simplificado
    row = 3
    style_section_header(ws, row, "P&L SIMPLIFICADO — CLAN CERVECERO", max_col)

    si = conn.execute("""
        SELECT SUM(ingreso_bruto), SUM(costo_total), SUM(centralizacion),
               SUM(ingreso_neto), SUM(margen_bruto), SUM(cantidad_facturada)
        FROM sell_in
    """).fetchone()

    dev_q = conn.execute("SELECT SUM(precio_reclamo_q + precio_devolucion_q) FROM devoluciones").fetchone()[0] or 0

    pnl = [
        ("Ingreso Bruto (Facturado a WM)", si[0] or 0),
        ("(-) Centralización 4%", -(si[2] or 0)),
        ("= Ingreso Neto", si[3] or 0),
        ("(-) Costo de Producto", -(si[1] or 0)),
        ("= Margen Bruto", si[4] or 0),
        ("(-) Devoluciones", -dev_q),
        ("= Margen Neto (estimado)", (si[4] or 0) - dev_q),
    ]

    row = 4
    headers = ['Concepto', 'Monto (Q)']
    style_headers(ws, row, headers, [36, 18])

    for label, val in pnl:
        row += 1
        cell_l = ws.cell(row=row, column=1, value=label)
        cell_v = ws.cell(row=row, column=2, value=val)
        cell_v.number_format = '#,##0.00'
        cell_l.border = thin_border
        cell_v.border = thin_border
        if label.startswith('='):
            cell_l.font = data_bold
            cell_v.font = data_bold
            if 'Neto (estimado)' in label:
                cell_v.fill = green_fill if val > 0 else red_fill
        else:
            cell_l.font = data_font
            cell_v.font = data_font

    # Margen %
    row += 1
    ws.cell(row=row, column=1, value="Margen Bruto %").font = data_bold
    ws.cell(row=row, column=1).border = thin_border
    c = ws.cell(row=row, column=2, value=(si[4] or 0) / (si[0] or 1))
    c.number_format = '0.0%'
    c.font = data_bold
    c.border = thin_border

    # Rentabilidad por Producto
    row += 2
    style_section_header(ws, row, "RENTABILIDAD POR PRODUCTO (Sell-In)", max_col)
    row += 1
    headers = ['Producto', 'Marca', 'Und Fact.', 'Ingreso Bruto', 'Costo', 'Centraliz.', 'Margen', 'Margen %', 'GMROI']
    style_headers(ws, row, headers, [24, 12, 10, 14, 14, 12, 14, 10, 10])

    prods = conn.execute("""
        SELECT si.nombre_producto, si.marca,
               SUM(si.cantidad_facturada), SUM(si.ingreso_bruto),
               SUM(si.costo_total), SUM(si.centralizacion), SUM(si.margen_bruto)
        FROM sell_in si
        GROUP BY si.codigo_producto
        ORDER BY SUM(si.margen_bruto) DESC
    """).fetchall()

    for prod, marca, und, ib, cost, centr, margen in prods:
        row += 1
        mpct = margen / ib if ib else 0
        # GMROI: margen / inventory cost for this product
        inv_cost = conn.execute("""
            SELECT SUM(rl.inv_actual) * p.costo_clan_und
            FROM retail_link rl
            JOIN productos p ON rl.producto = p.producto
            WHERE rl.tipo_registro='INVENTARIO' AND p.producto LIKE ?
            AND rl.semana_wm = (SELECT MAX(semana_wm) FROM retail_link WHERE tipo_registro='INVENTARIO')
        """, (f"%{(prod or '')[:10]}%",)).fetchone()[0] or 1
        gmroi = margen / inv_cost if inv_cost > 0 else 0

        write_data_row(ws, row, [prod, marca or '', und, ib, cost, centr, margen, mpct, gmroi],
                      [None, None, 'int', 'q2', 'q2', 'q2', 'q2', 'pct', 'gmroi'], bold_first=True)

    # Rentabilidad por Marca
    row += 2
    style_section_header(ws, row, "RENTABILIDAD POR MARCA", max_col)
    row += 1
    headers = ['Marca', 'Und Fact.', 'Ingreso Bruto', 'Costo', 'Margen', 'Margen %', '% del Total']
    style_headers(ws, row, headers, [16, 12, 14, 14, 14, 10, 10])

    total_margen = si[4] or 1
    brands = conn.execute("""
        SELECT marca, SUM(cantidad_facturada), SUM(ingreso_bruto),
               SUM(costo_total), SUM(margen_bruto)
        FROM sell_in
        GROUP BY marca
        ORDER BY SUM(margen_bruto) DESC
    """).fetchall()

    for marca, und, ib, cost, margen in brands:
        row += 1
        mpct = margen / ib if ib else 0
        contrib = margen / total_margen
        write_data_row(ws, row, [marca or '', und, ib, cost, margen, mpct, contrib],
                      [None, 'int', 'q2', 'q2', 'q2', 'pct', 'pct'], bold_first=True)

    # Devoluciones por producto
    row += 2
    style_section_header(ws, row, "DEVOLUCIONES POR PRODUCTO", max_col)
    row += 1
    headers = ['Producto', 'Reclamos Und', 'Reclamos Q', 'Dev. Cliente Und', 'Dev. Cliente Q', 'Total Q']
    style_headers(ws, row, headers, [24, 14, 14, 14, 14, 14])

    devs = conn.execute("""
        SELECT descripcion,
               SUM(cant_reclamo_tienda), SUM(precio_reclamo_q),
               SUM(devolucion_cliente_und), SUM(precio_devolucion_q),
               SUM(precio_reclamo_q + precio_devolucion_q) as total
        FROM devoluciones
        GROUP BY item_nbr
        ORDER BY total DESC
    """).fetchall()

    for desc, r_und, r_q, d_und, d_q, total in devs:
        row += 1
        write_data_row(ws, row, [desc, r_und, r_q, d_und, d_q, total],
                      [None, 'int', 'q2', 'int', 'q2', 'q2'], bold_first=True)

    print("  [5/5] Rentabilidad")
    return ws


# ═══════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════

def main():
    print("╔═══════════════════════════════════════════╗")
    print("║  Generating REPORTE-WM.xlsx               ║")
    print("╚═══════════════════════════════════════════╝")

    conn = sqlite3.connect(DB_PATH)
    wb = openpyxl.Workbook()

    build_dashboard(wb, conn)
    build_semanal(wb, conn)
    build_sugerido(wb, conn)
    build_rankings(wb, conn)
    build_rentabilidad(wb, conn)

    # Save
    os.makedirs(REPORT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime('%Y-%m-%d')
    filename = f"REPORTE-WM_{timestamp}.xlsx"
    filepath = os.path.join(REPORT_DIR, filename)
    wb.save(filepath)

    # Also save as "latest" for easy access
    latest_path = os.path.join(REPORT_DIR, "REPORTE-WM_latest.xlsx")
    wb.save(latest_path)

    size_kb = os.path.getsize(filepath) / 1024
    print(f"\n  Saved: {filepath}")
    print(f"  Also:  {latest_path}")
    print(f"  Size:  {size_kb:.0f} KB")
    print(f"\n  (Original MASTER WALMART was ~15MB with 21 sheets)")
    print(f"  (This report: {size_kb:.0f} KB with 5 sheets)")

    conn.close()


if __name__ == '__main__':
    main()
